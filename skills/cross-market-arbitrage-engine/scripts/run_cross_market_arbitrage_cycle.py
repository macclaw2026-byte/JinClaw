#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import time
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable
from urllib.parse import quote_plus
from urllib.request import urlopen
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mac_claw/.openclaw/workspace")
CONTROL_CENTER_ROOT = ROOT / "tools/openmoss/control_center"
STATE_DIR = ROOT / ".state"
STATE_DIR.mkdir(parents=True, exist_ok=True)
STATE_PATH = STATE_DIR / "cross-market-arbitrage-engine.json"
ADAPTIVE_HISTORY_PATH = STATE_DIR / "cross-market-arbitrage-adaptive-history.json"
QUERY_HISTORY_PATH = STATE_DIR / "cross-market-arbitrage-query-history.json"
MARKET_HISTORY_PATH = STATE_DIR / "cross-market-arbitrage-market-history.json"
THRESHOLD_HISTORY_PATH = STATE_DIR / "cross-market-arbitrage-threshold-history.json"
OUT_DIR = ROOT / "output" / "cross-market-arbitrage-engine"
OUT_DIR.mkdir(parents=True, exist_ok=True)
LATEST_REPORT_PATH = OUT_DIR / "latest-report.json"
FX_CACHE_PATH = STATE_DIR / "usd_cny_midrate.json"

TOOLS_ROOT = ROOT / "tools"
CRAWL4AI = TOOLS_ROOT / "bin" / "crawl4ai"
AGENT_BROWSER = TOOLS_ROOT / "agent-browser-local" / "node_modules" / "agent-browser" / "bin" / "agent-browser-darwin-arm64"
VENV_PY = TOOLS_ROOT / "matrix-venv" / "bin" / "python"
SITE_PREFS_PATH = ROOT / "tools/openmoss/runtime/autonomy/learning/crawler_site_preferences.json"
OPENCLAW_BIN = "/opt/homebrew/bin/openclaw"

import sys

if str(CONTROL_CENTER_ROOT) not in sys.path:
    sys.path.insert(0, str(CONTROL_CENTER_ROOT))

from control_plane_builder import build_control_plane
from memory_writeback_runtime import record_memory_writeback
from paths import CROSS_MARKET_ARBITRAGE_SCHEDULER_STATE_PATH

DISCOVERY_INTERVAL_SECONDS = 30 * 60
MATCH_INTERVAL_SECONDS = 60 * 60
REPORT_WINDOW_HOURS = 24
REPORT_TIMEZONE = ZoneInfo("America/New_York")
REPORT_HOUR = 18
DEFAULT_TELEGRAM_TARGET = "8528973600"
USD_TO_CNY_FALLBACK = 7.2
_USD_TO_CNY_CACHE: dict[str, Any] | None = None
DEFAULT_PLATFORM_FEE = 0.15
SELLERSPRITE_FETCHER = ROOT / "skills/cross-market-arbitrage-engine/scripts/sellersprite_session_fetch.py"
SELLERSPRITE_KEYWORD_SUBMIT_PROBE = ROOT / "skills/cross-market-arbitrage-engine/scripts/sellersprite_keyword_submit_probe.py"
SELLERSPRITE_PRODUCT_SUBMIT_PROBE = ROOT / "skills/cross-market-arbitrage-engine/scripts/sellersprite_product_submit_probe.py"
SELLERSPRITE_PRODUCT_RESEARCH_URL = "https://www.sellersprite.com/v3/product-research"
SELLERSPRITE_KEYWORD_RESEARCH_URL = "https://www.sellersprite.com/v2/keyword-research"

PLATFORM_FEE_TABLE = {
    "amazon": 0.15,
    "walmart": 0.15,
    "temu": 0.15,
}

DEFAULT_DISCOVERY_QUERIES = [
    "drawer organizer",
    "cable clips",
    "wall hooks",
    "storage basket",
    "under sink organizer",
]

SELL_PLATFORM_SEARCH = {
    "temu": lambda q: f"https://www.temu.com/search_result.html?search_key={quote_plus(q)}",
    "amazon": lambda q: f"https://www.amazon.com/s?k={quote_plus(q)}",
    "walmart": lambda q: f"https://www.walmart.com/search?q={quote_plus(q)}",
}

SOURCE_PLATFORM_SEARCH = {
    "1688": lambda q: f"https://s.1688.com/selloffer/offer_search.htm?keywords={quote_plus(q)}",
    "yiwugo": lambda q: f"https://en.yiwugo.com/search/s.html?queryKey={quote_plus(q)}",
    "made_in_china": lambda q: (
        "https://www.made-in-china.com/productdirectory.do?"
        f"subaction=hunt&style=b&mode=and&code=0&comProvince=nolimit&order=0&isOpenCorrection=1&org=top&keyword=&file=&searchType=0&word={quote_plus(q)}"
    ),
}

BLOCK_PATTERNS = {
    "temu": [r"verify", r"captcha", r"login", r"access denied", r"unusual traffic"],
    "amazon": [r"captcha", r"robot check", r"automated access", r"sorry, we just need to make sure"],
    "walmart": [r"robot or human", r"confirm that you.?re human", r"activate and hold the button"],
    "1688": [r"登录", r"请按住滑块", r"短信登录", r"密码登录", r"punish", r"x5secdata", r"nocaptcha", r"unusual traffic"],
    "yiwugo": [r"captcha", r"forbidden", r"access denied"],
    "made_in_china": [r"captcha", r"access denied", r"forbidden"],
}

RESTRICTED_PATTERNS = [
    r"\bbattery\b",
    r"\blithium\b",
    r"\bli[- ]?ion\b",
    r"\brechargeable\b",
    r"\bliquid\b",
    r"\bgel\b",
    r"\bspray\b",
    r"\bcosmetic\b",
    r"\bserum\b",
    r"\bcream\b",
    r"\btoner\b",
    r"\bfood\b",
    r"\bsnack\b",
    r"\bdrink\b",
    r"\bmedicine\b",
    r"\bdrug\b",
    r"\bcapsule\b",
    r"化妆",
    r"食品",
    r"药",
    r"锂电",
    r"液体",
]

PAIN_POINT_PATTERNS = [
    r"broke",
    r"broken",
    r"cheap",
    r"flimsy",
    r"too small",
    r"too thin",
    r"doesn't fit",
    r"poor quality",
    r"fell apart",
    r"leak",
    r"damaged",
    r"not sturdy",
    r"坏了",
    r"太小",
    r"太薄",
    r"质量差",
    r"不结实",
    r"容易坏",
]

MONTHLY_ORDER_PATTERNS = [
    r"(\d[\d,]*)\+?\s+bought in past month",
    r"(\d[\d,]*)\+?\s+purchased in the last month",
    r"(\d[\d,]*)\+?\s+sold in past month",
    r"(\d[\d,]*(?:\.\d+)?)\s*k\+?\s+(?:bought|purchased|sold)(?:\s+in\s+(?:the\s+)?)?(?:past|last)?\s*(?:30 days|month)",
    r"(\d[\d,]*(?:\.\d+)?)\s*k\+?\s+(?:bought|purchased|sold)(?:\s+in\s+(?:the\s+)?)?(?:past|last)?\s*month",
    r"(\d[\d,]*)\+?\s+(?:bought|purchased|sold)(?:\s+in\s+(?:the\s+)?)?(?:past|last)?\s*(?:month|30 days)",
    r'"(?:monthlySold|unitsSold|monthly_orders|monthlySales)"\s*[:=]\s*"?(?:\s*)(\d[\d,]*)\+?"?',
    r'"(?:monthlySold|unitsSold|monthly_orders|monthlySales)"\s*:\s*(\d[\d,]*)',
]

SOLD_PATTERNS = [
    r"(\d+(?:\.\d+)?)k\+?\s+sold",
    r"(\d[\d,]*)\+?\s+sold",
]

WEEKLY_ORDER_PATTERNS = [
    r"(\d[\d,]*(?:\.\d+)?)\s*k\+?\s+(?:bought|purchased|sold)(?:\s+in\s+(?:the\s+)?)?(?:past|last)?\s*(?:week|7 days)",
    r"(\d[\d,]*)\+?\s+(?:bought|purchased|sold)(?:\s+in\s+(?:the\s+)?)?(?:past|last)?\s*(?:week|7 days)",
]

RATING_PATTERNS = [
    r"(\d(?:\.\d)?)\s+out of 5 stars",
    r"(\d(?:\.\d)?)\s*\/\s*5",
]

REVIEW_COUNT_PATTERNS = [
    r"(\d[\d,]*)\s+ratings",
    r"(\d[\d,]*)\s+reviews",
]

AMAZON_TOP_CATEGORIES = {
    "home & kitchen",
    "kitchen & dining",
    "tools & home improvement",
    "patio, lawn & garden",
    "office products",
    "industrial & scientific",
    "sports & outdoors",
    "arts, crafts & sewing",
    "home improvement",
    "pet supplies",
}

LISTING_AGE_PATTERNS = [
    r"Date First Available[^A-Za-z0-9]{0,20}([A-Za-z]+ \d{1,2}, \d{4})",
    r"First available[^A-Za-z0-9]{0,20}([A-Za-z]+ \d{1,2}, \d{4})",
    r"Available on[^A-Za-z0-9]{0,20}([A-Za-z]+ \d{1,2}, \d{4})",
    r"Date First Available[^0-9]{0,20}(\d{1,2}/\d{1,2}/\d{4})",
    r"First available[^0-9]{0,20}(\d{1,2}/\d{1,2}/\d{4})",
    r"Available on[^0-9]{0,20}(\d{1,2}/\d{1,2}/\d{4})",
    r'"(?:dateFirstAvailable|firstAvailableDate|datePublished|releaseDate)"\s*[:=]\s*"(\d{4}-\d{2}-\d{2})"',
    r'"(?:dateFirstAvailable|firstAvailableDate|datePublished|releaseDate)"\s*[:=]\s*"(\d{4}/\d{2}/\d{2})"',
]

WEIGHT_PATTERNS = [
    r"(\d+(?:\.\d+)?)\s*(kg|公斤)",
    r"(\d+(?:\.\d+)?)\s*(g|克)",
    r"weight[^\\d]{0,10}(\d+(?:\.\d+)?)\s*(kg|g)",
    r"净重[^\\d]{0,10}(\d+(?:\.\d+)?)\s*(kg|g|公斤|克)",
    r"毛重[^\\d]{0,10}(\d+(?:\.\d+)?)\s*(kg|g|公斤|克)",
]

PRICE_PATTERNS = [
    r"(?:US\$|USD\s*)(\d+(?:\.\d+)?)",
    r"\$(\d+(?:\.\d+)?)",
    r"(?:¥|RMB\s*|￥)(\d+(?:\.\d+)?)",
    r"(\d+(?:\.\d+)?)\s*元",
]

TITLE_STOPWORDS = {
    "for",
    "the",
    "with",
    "and",
    "pack",
    "pcs",
    "piece",
    "pieces",
    "set",
    "sale",
    "new",
    "hot",
    "portable",
    "wireless",
    "home",
}

SOURCE_QUERY_STOPWORDS = {
    "organizer",
    "storage",
    "household",
    "home",
    "kitchen",
    "portable",
    "adjustable",
    "multifunction",
    "multi",
    "tool",
}

COLOR_TOKENS = {
    "black", "white", "gray", "grey", "silver", "gold", "beige", "brown", "pink", "blue", "green",
    "red", "purple", "orange", "yellow", "clear", "transparent", "wood", "wooden", "natural",
}

MATERIAL_TOKENS = {
    "plastic", "acrylic", "metal", "steel", "iron", "wood", "wooden", "bamboo", "silicone", "fabric",
    "cotton", "linen", "canvas", "polyester", "mesh", "glass", "ceramic", "paper", "leather",
}

FEATURE_TOKENS = {
    "wall", "mounted", "hanging", "drawer", "stackable", "foldable", "portable", "magnetic", "adhesive",
    "waterproof", "desktop", "desk", "under", "sink", "closet", "shelf", "tier", "tray", "basket", "box",
    "hook", "rack", "holder", "container", "dispenser", "pill", "cable", "bag", "tote", "insert", "bin",
}

CHINESE_TOKEN_NORMALIZERS: dict[str, tuple[str, ...]] = {
    "透明": ("clear", "transparent"),
    "黑色": ("black",),
    "白色": ("white",),
    "灰色": ("gray",),
    "银色": ("silver",),
    "金色": ("gold",),
    "米色": ("beige",),
    "棕色": ("brown",),
    "粉色": ("pink",),
    "蓝色": ("blue",),
    "绿色": ("green",),
    "红色": ("red",),
    "紫色": ("purple",),
    "橙色": ("orange",),
    "黄色": ("yellow",),
    "塑料": ("plastic",),
    "亚克力": ("acrylic",),
    "金属": ("metal",),
    "钢制": ("steel",),
    "铁艺": ("iron",),
    "木质": ("wooden",),
    "竹制": ("bamboo",),
    "硅胶": ("silicone",),
    "布艺": ("fabric",),
    "棉质": ("cotton",),
    "亚麻": ("linen",),
    "帆布": ("canvas",),
    "网格": ("mesh",),
    "玻璃": ("glass",),
    "陶瓷": ("ceramic",),
    "皮革": ("leather",),
    "壁挂": ("wall", "mounted"),
    "挂墙": ("wall", "mounted"),
    "悬挂": ("hanging",),
    "抽屉": ("drawer",),
    "可叠加": ("stackable",),
    "可折叠": ("foldable",),
    "便携": ("portable",),
    "磁吸": ("magnetic",),
    "粘贴": ("adhesive",),
    "防水": ("waterproof",),
    "桌面": ("desktop", "desk"),
    "台面": ("desktop",),
    "台下": ("under",),
    "水槽": ("sink",),
    "衣柜": ("closet",),
    "置物架": ("shelf", "rack"),
    "分层": ("tier",),
    "托盘": ("tray",),
    "收纳篮": ("basket",),
    "收纳盒": ("box", "bin", "container"),
    "挂钩": ("hook",),
    "支架": ("holder", "rack"),
    "容器": ("container",),
    "分配器": ("dispenser",),
    "药盒": ("pill", "box"),
    "线缆": ("cable",),
    "卡扣": ("clip",),
    "托特包": ("tote", "bag"),
    "内胆": ("insert",),
}

MATCH_TOKEN_STOPWORDS = {
    *TITLE_STOPWORDS,
    *SOURCE_QUERY_STOPWORDS,
    "amazon",
    "walmart",
    "temu",
    "basics",
    "regular",
    "conventional",
    "organic",
    "premium",
    "wholesale",
}

SOURCE_QUERY_BANNED_TOKENS = {
    "amazon",
    "walmart",
    "temu",
    "com",
    "www",
    "http",
    "https",
    "dp",
    "ip",
    "html",
    "shtml",
}

SOURCE_QUERY_TRANSLATIONS = {
    "drawer": "抽屉",
    "organizer": "收纳盒",
    "organizers": "收纳盒",
    "storage": "收纳",
    "basket": "收纳篮",
    "bin": "收纳盒",
    "bins": "收纳盒",
    "box": "收纳盒",
    "boxes": "收纳盒",
    "tray": "托盘",
    "shelf": "置物架",
    "rack": "置物架",
    "holder": "支架",
    "hook": "挂钩",
    "hooks": "挂钩",
    "wall": "墙面",
    "mounted": "壁挂",
    "under": "下方",
    "sink": "水槽",
    "desk": "桌面",
    "desktop": "桌面",
    "cable": "线缆",
    "clips": "卡扣",
    "clip": "卡扣",
    "pill": "药盒",
    "plastic": "塑料",
    "clear": "透明",
    "acrylic": "亚克力",
    "bamboo": "竹制",
    "wood": "木质",
    "wooden": "木质",
    "metal": "金属",
    "steel": "钢制",
    "black": "黑色",
    "white": "白色",
    "gray": "灰色",
    "grey": "灰色",
    "bag": "包",
    "tote": "托特包",
    "insert": "内胆",
    "container": "容器",
    "dispenser": "分配器",
    "kitchen": "厨房",
    "closet": "衣柜",
    "tier": "分层",
    "stackable": "可叠加",
    "foldable": "可折叠",
    "magnetic": "磁吸",
    "adhesive": "粘贴",
}


@dataclass
class FetchResult:
    platform: str
    tool: str
    status: str
    score: int
    text: str
    notes: str = ""


@dataclass
class DemandCandidate:
    candidate_id: str
    title: str
    sell_platform: str
    sell_link: str
    sell_price_cny: float | None
    query: str
    extracted_at: str
    estimated_daily_orders: float | None = None
    listing_age_days: int | None = None
    rating_value: float | None = None
    review_count: int | None = None
    demand_confidence: float = 0.0
    demand_refresh_priority: float = 0.0
    raw_signals: dict[str, Any] = field(default_factory=dict)


@dataclass
class SourceCandidate:
    platform: str
    link: str
    title: str
    price_cny: float | None
    weight_kg: float | None
    weight_grade: str
    fetch_tool: str
    match_score: float
    keyword_similarity: float = 0.0
    attribute_similarity: float = 0.0
    price_band_similarity: float = 0.0
    supplier_similarity_score: float = 0.0
    supplier_similarity_grade: str = "none"
    supplier_similarity_reasons: list[str] = field(default_factory=list)
    weight_proxy_used: bool = False
    blocked: bool = False
    notes: str = ""


@dataclass
class ArbitrageDecision:
    product_name: str
    buy_platform: str
    buy_link: str
    sell_platform: str
    sell_link: str
    sell_price_cny: float | None
    purchase_cost_cny: float | None
    weight_kg: float | None
    gross_profit_amount: float | None
    gross_margin_rate: float | None
    conservative_margin_rate: float | None
    platform_fee_rate: float | None
    estimated_daily_orders: float | None
    listing_age_days: int | None
    demand_score: float
    competition_score: float
    differentiation_score: float
    price_stability_score: float
    platform_fit_score: float
    platform_fit_label: str
    platform_recommendation: str
    launchability_score: float
    confidence_score: float
    weight_grade: str
    qualified: bool
    selection_score: float = 0.0
    selection_grade: str = "watch"
    selection_thesis: str = ""
    query_source_kind: str = "base"
    trend_source_score: float = 0.0
    execution_resilience_score: float = 0.0
    supplier_confidence_score: float = 0.0
    supplier_similarity_score: float = 0.0
    supplier_similarity_grade: str = "none"
    head_monopoly_score: float = 0.0
    decision_bucket: str = "watchlist"
    priority_reason: str = ""
    reasons: list[str] = field(default_factory=list)


DECISION_DEFAULTS: dict[str, Any] = {
    "platform_fee_rate": None,
    "estimated_daily_orders": None,
    "listing_age_days": None,
    "demand_score": 0.0,
    "competition_score": 0.0,
    "differentiation_score": 0.0,
    "price_stability_score": 0.0,
    "platform_fit_score": 0.0,
    "platform_fit_label": "watch",
    "platform_recommendation": "watchlist_only",
    "launchability_score": 0.0,
    "confidence_score": 0.0,
    "weight_grade": "D",
    "qualified": False,
    "selection_score": 0.0,
    "selection_grade": "watch",
    "selection_thesis": "",
    "query_source_kind": "base",
    "trend_source_score": 0.0,
    "execution_resilience_score": 0.0,
    "supplier_confidence_score": 0.0,
    "supplier_similarity_score": 0.0,
    "supplier_similarity_grade": "none",
    "head_monopoly_score": 0.0,
    "decision_bucket": "watchlist",
    "priority_reason": "",
    "reasons": [],
}


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _utc_now_iso() -> str:
    return _utc_now().isoformat()


def _load_fx_cache() -> dict[str, Any] | None:
    if not FX_CACHE_PATH.exists():
        return None
    try:
        payload = json.loads(FX_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _save_fx_cache(payload: dict[str, Any]) -> None:
    FX_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _fetch_usd_to_cny_midrate() -> dict[str, Any] | None:
    try:
        with urlopen("https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml", timeout=20) as resp:
            raw = resp.read()
    except Exception:
        return None
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return None
    ns = {"e": "http://www.ecb.int/vocabulary/2002-08-01/eurofxref"}
    time_node = root.find(".//e:Cube[@time]", ns)
    if time_node is None:
        return None
    as_of = time_node.attrib.get("time")
    eur_usd = None
    eur_cny = None
    for cube in time_node.findall("e:Cube", ns):
        currency = (cube.attrib.get("currency") or "").upper()
        rate_raw = cube.attrib.get("rate")
        try:
            rate = float(rate_raw)
        except Exception:
            continue
        if currency == "USD":
            eur_usd = rate
        elif currency == "CNY":
            eur_cny = rate
    if not eur_usd or not eur_cny or eur_usd <= 0:
        return None
    return {
        "source": "ECB euro foreign exchange reference rates",
        "as_of": as_of,
        "eur_usd": eur_usd,
        "eur_cny": eur_cny,
        "usd_cny_midrate": round(eur_cny / eur_usd, 6),
        "fetched_at": _utc_now_iso(),
    }


def _usd_to_cny_rate() -> float:
    global _USD_TO_CNY_CACHE
    today = _utc_now().date().isoformat()
    if _USD_TO_CNY_CACHE and _USD_TO_CNY_CACHE.get("as_of") == today:
        return float(_USD_TO_CNY_CACHE.get("usd_cny_midrate") or USD_TO_CNY_FALLBACK)
    cached = _load_fx_cache()
    if cached and cached.get("as_of") == today:
        _USD_TO_CNY_CACHE = cached
        return float(cached.get("usd_cny_midrate") or USD_TO_CNY_FALLBACK)
    fresh = _fetch_usd_to_cny_midrate()
    if fresh:
        _USD_TO_CNY_CACHE = fresh
        try:
            _save_fx_cache(fresh)
        except OSError:
            pass
        return float(fresh.get("usd_cny_midrate") or USD_TO_CNY_FALLBACK)
    if cached and cached.get("usd_cny_midrate") is not None:
        _USD_TO_CNY_CACHE = cached
        return float(cached["usd_cny_midrate"])
    return USD_TO_CNY_FALLBACK


def _ny_now() -> datetime:
    return _utc_now().astimezone(REPORT_TIMEZONE)


def _read_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default


def _write_json_file(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _parse_iso(value: str) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None


def _seconds_since(value: str) -> int | None:
    dt = _parse_iso(value)
    if not dt:
        return None
    return max(0, int((_utc_now() - dt).total_seconds()))


def _load_scheduler_policy() -> dict[str, Any]:
    try:
        control_plane = build_control_plane()
    except Exception:
        return {}
    return (control_plane.get("project_scheduler_policy", {}) or {}).get("cross_market_arbitrage", {}) or {}


def _load_scheduler_state() -> dict[str, Any]:
    return _read_json(CROSS_MARKET_ARBITRAGE_SCHEDULER_STATE_PATH, {})


def _write_scheduler_state(payload: dict[str, Any]) -> None:
    _write_json_file(CROSS_MARKET_ARBITRAGE_SCHEDULER_STATE_PATH, payload)


def _execution_flags_from_scheduler_policy(scheduler_policy: dict[str, Any] | None) -> dict[str, bool]:
    scheduler_policy = scheduler_policy or {}
    repair_mode = str(scheduler_policy.get("repair_mode", "")).strip()
    start_tasks = bool(scheduler_policy.get("start_tasks", True))
    flags = {
        "allow_discovery": start_tasks,
        "allow_match": start_tasks,
        "allow_report": True,
    }
    if not start_tasks:
        flags["allow_discovery"] = False
        flags["allow_match"] = False
        return flags
    if repair_mode == "crawler_hold":
        flags["allow_discovery"] = False
        flags["allow_match"] = False
    elif repair_mode == "repair_backpressure":
        flags["allow_discovery"] = False
        flags["allow_match"] = True
    elif repair_mode == "repair_observe":
        flags["allow_discovery"] = False
        flags["allow_match"] = True
    return flags


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-") or "item"


def _run(cmd: list[str], *, timeout: int = 45) -> subprocess.CompletedProcess:
    timeout_cap = os.environ.get("CROSS_MARKET_TIMEOUT_CAP", "").strip()
    if timeout_cap.isdigit():
        timeout = min(timeout, max(3, int(timeout_cap)))
    try:
        return subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout or ""
        stderr = exc.stderr or ""
        if isinstance(stdout, bytes):
            stdout = stdout.decode("utf-8", "ignore")
        if isinstance(stderr, bytes):
            stderr = stderr.decode("utf-8", "ignore")
        return subprocess.CompletedProcess(
            cmd,
            124,
            stdout=stdout,
            stderr=(stderr + f"\nTIMEOUT after {timeout}s").strip(),
        )


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _normalize_title(text: str) -> str:
    tokens = re.findall(r"[a-z0-9]+", str(text or "").lower())
    return " ".join(token for token in tokens if token not in TITLE_STOPWORDS)


def _title_from_link(link: str, fallback: str) -> str:
    text = str(link or "")
    if "temu.com" in text:
        match = re.search(r"/([a-z0-9-]+)-s\.html", text, flags=re.I)
        if match:
            return match.group(1).replace("-", " ").strip().title()
    if "amazon.com" in text:
        return fallback
    if "walmart.com" in text:
        match = re.search(r"/ip/([^/?]+)", text, flags=re.I)
        if match:
            return match.group(1).replace("-", " ").strip().title()
    return fallback


def _extract_page_title(text: str) -> str | None:
    match = re.search(r"<title>([^<]+)</title>", text, flags=re.I)
    if match:
        title = _clean_text(match.group(1))
        if title:
            return re.sub(r"\s*[|\-]\s*(Amazon|Temu|Walmart).*$", "", title, flags=re.I).strip()
    snapshot_title = re.search(r"^([^\n]{10,160})$", text.strip(), flags=re.M)
    if snapshot_title:
        title = _clean_text(snapshot_title.group(1))
        if title and "http" not in title.lower():
            return title
    return None


def _source_query_variants(candidate: DemandCandidate) -> list[str]:
    query_tokens = set(_normalize_title(candidate.query).split())
    generic_source_tokens = (
        COLOR_TOKENS
        | MATERIAL_TOKENS
        | FEATURE_TOKENS
        | SOURCE_QUERY_STOPWORDS
        | SOURCE_QUERY_BANNED_TOKENS
        | {"divider", "dividers", "organizer", "organizers", "storage", "holder", "holders"}
    )

    def _strip_brand_like_prefix(tokens: list[str]) -> list[str]:
        trimmed = list(tokens)
        while len(trimmed) >= 3:
            head = trimmed[0]
            if head in query_tokens or head in generic_source_tokens:
                break
            trimmed = trimmed[1:]
        return trimmed

    def _compact(text: str) -> str:
        text = re.sub(r"https?://\S+", " ", str(text or ""), flags=re.I)
        text = re.sub(r"\b(?:www\.)?(?:amazon|walmart|temu)\.com\b", " ", text, flags=re.I)
        text = re.sub(r"\b(?:dp|ip|offer|product|detail|html|shtml|ref)\b", " ", text, flags=re.I)
        text = _normalize_title(text)
        text = re.sub(r"\b\d+(?:\.\d+)?\s*(?:pack|pcs?|pieces?|count|ct|inch|in|cm|mm|oz|ml|lb|lbs|g|kg)\b", " ", text)
        text = re.sub(r"\b(?:set of|pack of|for women|for men|for kids|amazon basics|walmart|temu)\b", " ", text)
        text = re.sub(r"\b\d+\b", " ", text)
        tokens = [token for token in text.split() if token not in SOURCE_QUERY_BANNED_TOKENS]
        tokens = _strip_brand_like_prefix(tokens)
        text = " ".join(tokens)
        text = re.sub(r"\s+", " ", text).strip()
        return text

    primary = _compact(candidate.title) or _compact(candidate.query) or candidate.query.strip().lower()
    fallback = _normalize_title(candidate.query) or candidate.query.strip().lower()
    token_pool = [
        token
        for token in primary.split()
        if token not in SOURCE_QUERY_STOPWORDS and token not in SOURCE_QUERY_BANNED_TOKENS
    ]
    variants: list[str] = []

    for text in [
        " ".join(token_pool[:6]).strip(),
        " ".join(token_pool[:5]).strip(),
        " ".join(token_pool[:4]).strip(),
        " ".join(token_pool[:3]).strip(),
        " ".join(token_pool[:2]).strip(),
        primary,
        " ".join(token for token in fallback.split() if token not in SOURCE_QUERY_BANNED_TOKENS).strip(),
        _compact(candidate.query),
        _compact(candidate.title),
        _compact(candidate.query.strip().lower()),
        _compact(candidate.title.strip().lower()),
    ]:
        text = re.sub(r"\s+", " ", text).strip()
        if len(text.split()) < 2:
            continue
        if text and text not in variants:
            variants.append(text)
    return variants[:6]


def _to_chinese_source_queries(english_variants: list[str]) -> list[str]:
    chinese_variants: list[str] = []
    for variant in english_variants:
        tokens = [token for token in re.findall(r"[a-z0-9]+", str(variant or "").lower()) if token]
        translated = [SOURCE_QUERY_TRANSLATIONS.get(token, "") for token in tokens]
        translated = [token for token in translated if token]
        if len(translated) < 2:
            continue
        compact_translated: list[str] = []
        for token in translated[:6]:
            if compact_translated and compact_translated[-1] == token:
                continue
            compact_translated.append(token)
        phrase = "".join(compact_translated).strip()
        if phrase and phrase not in chinese_variants:
            chinese_variants.append(phrase)
    return chinese_variants[:4]


def _source_query_variants_for_platform(candidate: DemandCandidate, platform: str) -> list[str]:
    base_variants = _source_query_variants(candidate)
    if platform not in {"1688", "yiwugo"}:
        return base_variants
    chinese_variants = _to_chinese_source_queries(base_variants)
    merged: list[str] = []
    for text in [*chinese_variants, *base_variants]:
        normalized = re.sub(r"\s+", " ", str(text or "")).strip()
        if not normalized:
            continue
        if normalized not in merged:
            merged.append(normalized)
    return merged[:8]


def _load_site_preferences() -> dict[str, Any]:
    if not SITE_PREFS_PATH.exists():
        return {}
    try:
        return json.loads(SITE_PREFS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _fetch_direct_http(url: str) -> tuple[str, str]:
    timeout = 30
    timeout_cap = os.environ.get("CROSS_MARKET_TIMEOUT_CAP", "").strip()
    if timeout_cap.isdigit():
        timeout = min(timeout, max(3, int(timeout_cap)))
    code = (
        "import sys,urllib.request; "
        "u=sys.argv[1]; "
        f"t={timeout}; "
        "req=urllib.request.Request(u, headers={'User-Agent':'Mozilla/5.0'}); "
        "resp=urllib.request.urlopen(req, timeout=t); "
        "print(resp.read().decode('utf-8','ignore')[:400000])"
    )
    proc = _run(["python3", "-c", code, url])
    return proc.stdout, proc.stderr


def _fetch_curl_cffi(url: str) -> tuple[str, str]:
    timeout = 30
    timeout_cap = os.environ.get("CROSS_MARKET_TIMEOUT_CAP", "").strip()
    if timeout_cap.isdigit():
        timeout = min(timeout, max(3, int(timeout_cap)))
    code = (
        "from curl_cffi import requests; import sys; "
        f"r=requests.get(sys.argv[1], impersonate='chrome124', timeout={timeout}); "
        "print(r.text[:400000])"
    )
    proc = _run([str(VENV_PY), "-c", code, url])
    return proc.stdout, proc.stderr


def _fetch_playwright(url: str, stealth: bool = False) -> tuple[str, str]:
    preamble = "from playwright_stealth import Stealth\n" if stealth else ""
    apply = "    Stealth().apply_stealth_sync(page)\n" if stealth else ""
    code = (
        "from playwright.sync_api import sync_playwright\n"
        f"{preamble}"
        "import sys\n"
        "u = sys.argv[1]\n"
        "with sync_playwright() as p:\n"
        "    browser = p.chromium.launch(headless=True)\n"
        "    page = browser.new_page()\n"
        f"{apply}"
        "    page.goto(u, wait_until='load', timeout=45000)\n"
        "    html = page.content()[:400000]\n"
        "    body_text = ''\n"
        "    try:\n"
        "        body_text = page.locator('body').inner_text(timeout=10000)[:200000]\n"
        "    except Exception:\n"
        "        body_text = ''\n"
        "    print(html)\n"
        "    if body_text:\n"
        "        print('\\n\\n__VISIBLE_TEXT__\\n')\n"
        "        print(body_text)\n"
        "    browser.close()\n"
    )
    proc = _run([str(VENV_PY), "-c", code, url])
    return proc.stdout, proc.stderr


def _fetch_crawl4ai(url: str) -> tuple[str, str]:
    proc = _run([str(CRAWL4AI), url, "-o", "markdown", "--bypass-cache", "-c", "wait_until=load"])
    return proc.stdout, proc.stderr


def _fetch_agent_browser(url: str) -> tuple[str, str]:
    _run([str(AGENT_BROWSER), "connect", "http://127.0.0.1:18800"], timeout=20)
    _run([str(AGENT_BROWSER), "open", url], timeout=25)
    current_url = _run([str(AGENT_BROWSER), "get", "url"], timeout=10).stdout.strip()
    title = _run([str(AGENT_BROWSER), "get", "title"], timeout=10).stdout.strip()
    snapshot = _run([str(AGENT_BROWSER), "snapshot"], timeout=20).stdout
    return f"opened_url={current_url}\n{title}\n{snapshot}", ""


TOOL_FETCHERS: dict[str, Callable[[str], tuple[str, str]]] = {
    "direct_http": _fetch_direct_http,
    "curl_cffi": _fetch_curl_cffi,
    "playwright": lambda url: _fetch_playwright(url, stealth=False),
    "playwright_stealth": lambda url: _fetch_playwright(url, stealth=True),
    "crawl4ai": _fetch_crawl4ai,
    "agent_browser": _fetch_agent_browser,
}


def _tool_order(platform: str) -> list[str]:
    prefs = (_load_site_preferences().get("sites", {}) or {}).get(platform, {}) or {}
    learned = []
    for raw in prefs.get("preferred_tool_order", []) or []:
        raw = str(raw).strip().lower()
        if raw in {"curl-cffi", "curl_cffi"}:
            learned.append("curl_cffi")
        elif raw in {"crawl4ai-cli", "crawl4ai"}:
            learned.append("crawl4ai")
        elif raw in {"scrapy-cffi", "scrapy_cffi"}:
            learned.append("curl_cffi")
        elif raw in {"local-agent-browser-cli", "agent-browser", "agent_browser"}:
            learned.append("agent_browser")
        elif raw in {"playwright"}:
            learned.append("playwright")
        elif raw in {"playwright-stealth", "playwright_stealth"}:
            learned.append("playwright_stealth")
        elif raw in {"direct-http-html", "direct_http"}:
            learned.append("direct_http")
    defaults = {
        "temu": ["agent_browser", "crawl4ai", "curl_cffi", "playwright_stealth", "direct_http"],
        "amazon": ["curl_cffi", "agent_browser", "playwright", "crawl4ai"],
        "walmart": ["agent_browser", "curl_cffi", "playwright_stealth", "crawl4ai"],
        "1688": ["agent_browser", "playwright_stealth", "playwright", "curl_cffi", "crawl4ai", "direct_http"],
        "yiwugo": ["curl_cffi", "playwright_stealth", "playwright", "crawl4ai", "direct_http"],
        "made_in_china": ["curl_cffi", "direct_http", "crawl4ai", "playwright_stealth"],
    }.get(platform, ["curl_cffi", "crawl4ai", "direct_http"])
    ordered = []
    for item in [*learned, *defaults]:
        if item not in ordered:
            ordered.append(item)
    return ordered


def _score_fetch(platform: str, tool: str, text: str, stderr: str) -> FetchResult:
    lowered = text.lower()
    blocked = sum(len(re.findall(p, lowered, flags=re.I)) for p in BLOCK_PATTERNS.get(platform, []))
    score = 0
    if len(text) > 500:
        score += 20
    if len(text) > 5000:
        score += 15
    if "price" in lowered or "¥" in text or "$" in text:
        score += 15
    if "href" in lowered or "/goods.html" in lowered or "/dp/" in lowered or "/ip/" in lowered or "productdirectory" in lowered or "offer_search" in lowered:
        score += 20
    if blocked:
        score -= min(60, blocked * 12)
    status = "blocked" if blocked else ("usable" if score >= 50 else "partial")
    return FetchResult(platform=platform, tool=tool, status=status, score=max(0, min(100, score)), text=text, notes=stderr[:400])


def fetch_best(platform: str, url: str, *, max_tools: int | None = None, fast: bool = False) -> FetchResult:
    best: FetchResult | None = None
    ordered_tools = _tool_order(platform)
    if fast:
        preferred_fast = [tool for tool in ordered_tools if tool in {"curl_cffi", "direct_http"}]
        ordered_tools = preferred_fast or ordered_tools[:1]
    if max_tools is not None:
        ordered_tools = ordered_tools[:max_tools]
    for tool in ordered_tools:
        fetcher = TOOL_FETCHERS.get(tool)
        if not fetcher:
            continue
        try:
            text, stderr = fetcher(url)
        except Exception as exc:
            candidate = FetchResult(platform=platform, tool=tool, status="failed", score=0, text="", notes=str(exc))
        else:
            candidate = _score_fetch(platform, tool, text, stderr)
        if best is None or candidate.score > best.score:
            best = candidate
        if candidate.status == "usable":
            return candidate
    return best or FetchResult(platform=platform, tool="none", status="failed", score=0, text="", notes="no fetcher")


def _extract_prices(text: str) -> list[float]:
    prices = []
    for pattern in PRICE_PATTERNS:
        for raw in re.findall(pattern, text, flags=re.I):
            try:
                prices.append(float(raw))
            except Exception:
                pass
    return prices


def _extract_prices_detailed(text: str) -> list[tuple[float, str]]:
    rows: list[tuple[float, str]] = []
    for raw in re.findall(r"(?:US\$|USD\s*)(\d+(?:\.\d+)?)", text, flags=re.I):
        try:
            rows.append((float(raw), "USD"))
        except Exception:
            pass
    for raw in re.findall(r"\$(\d+(?:\.\d+)?)", text, flags=re.I):
        try:
            rows.append((float(raw), "USD"))
        except Exception:
            pass
    for raw in re.findall(r"(?:¥|RMB\s*|￥)(\d+(?:\.\d+)?)", text, flags=re.I):
        try:
            rows.append((float(raw), "CNY"))
        except Exception:
            pass
    for raw in re.findall(r"(\d+(?:\.\d+)?)\s*元", text, flags=re.I):
        try:
            rows.append((float(raw), "CNY"))
        except Exception:
            pass
    return rows


def _extract_amazon_price_usd(text: str) -> float | None:
    patterns = [
        r'id="apex-pricetopay-accessibility-label"[^>]*>\s*\$?(\d+(?:\.\d+)?)',
        r'class="[^"]*priceToPay[^"]*"[^>]*>\s*<span class="a-offscreen">\s*\$?(\d+(?:\.\d+)?)',
        r'class="[^"]*apex-pricetopay-value[^"]*"[^>]*>.*?<span class="a-offscreen">\s*\$?(\d+(?:\.\d+)?)',
        r'id="corePriceDisplay_desktop_feature_div".*?<span class="a-offscreen">\s*\$?(\d+(?:\.\d+)?)',
    ]
    for pattern in patterns:
        found = re.search(pattern, text, flags=re.I | re.S)
        if not found:
            continue
        try:
            value = float(found.group(1))
        except Exception:
            continue
        if value > 1.0:
            return value
    return None


def _price_to_cny(value: float, currency: str, platform: str) -> float:
    usd_to_cny = _usd_to_cny_rate()
    if currency == "USD":
        return value * usd_to_cny
    if platform in {"amazon", "walmart", "temu"}:
        return value * usd_to_cny
    if platform == "made_in_china" and currency != "CNY":
        return value * usd_to_cny
    return value


def _best_price_cny(text: str, platform: str) -> float | None:
    detailed = _extract_prices_detailed(text[:120000])
    if detailed:
        converted = [_price_to_cny(value, currency, platform) for value, currency in detailed if value > 0]
        if converted:
            return round(min(converted), 2)
    simple = _extract_prices(text[:120000])
    if not simple:
        return None
    inferred = min(simple)
    if platform in {"amazon", "walmart", "temu", "made_in_china"}:
        inferred *= _usd_to_cny_rate()
    return round(inferred, 2)


def _extract_platform_price_cny(text: str, platform: str) -> float | None:
    if platform == "amazon":
        amazon_price = _extract_amazon_price_usd(text)
        if amazon_price is not None:
            return round(amazon_price * _usd_to_cny_rate(), 2)
    if platform == "yiwugo":
        prices: list[float] = []
        for raw in re.findall(r'sellprice="(\d+(?:\.\d+)?)"', text, flags=re.I):
            try:
                value = float(raw)
            except Exception:
                continue
            if value > 0:
                prices.append(value)
        if prices:
            return round(min(prices), 2)
    if platform == "made_in_china":
        usd_hits: list[float] = []
        for raw in re.findall(r"US\$\s*(\d+(?:\.\d+)?)", text, flags=re.I):
            try:
                value = float(raw)
            except Exception:
                continue
            if value > 0:
                usd_hits.append(value)
        if usd_hits:
            return round(min(usd_hits) * _usd_to_cny_rate(), 2)
    return _best_price_cny(text, platform)


def _extract_rating_value(text: str) -> float | None:
    for pattern in RATING_PATTERNS:
        found = re.search(pattern, text, flags=re.I)
        if found:
            try:
                value = float(found.group(1))
            except Exception:
                continue
            if 0.0 < value <= 5.0:
                return value
    return None


def _extract_review_count(text: str) -> int | None:
    for pattern in REVIEW_COUNT_PATTERNS:
        found = re.search(pattern, text, flags=re.I)
        if found:
            try:
                return int(found.group(1).replace(",", ""))
            except Exception:
                continue
    return None


def _normalize_category_name(raw: str) -> str:
    cleaned = _clean_text(raw)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" >-")
    return cleaned


def _extract_amazon_best_seller_rank(text: str) -> tuple[int | None, str | None]:
    section_patterns = [
        r"Best Sellers Rank.*?#([\d,]+)\s+in\s+([^<\n\r]+)",
        r"#([\d,]+)\s+in\s+([^<\n\r]+).*?Best Sellers Rank",
    ]
    for pattern in section_patterns:
        found = re.search(pattern, text, flags=re.I | re.S)
        if not found:
            continue
        try:
            rank = int(found.group(1).replace(",", ""))
        except Exception:
            continue
        category = _normalize_category_name(found.group(2))
        if rank > 0 and category:
            return rank, category
    return None, None


def _estimate_amazon_daily_orders_from_rank(rank: int | None, category: str | None) -> float | None:
    if not rank or rank <= 0:
        return None
    lowered = (category or "").lower()
    is_top_category = any(token in lowered for token in AMAZON_TOP_CATEGORIES)
    if is_top_category:
        if rank <= 100:
            return 120.0
        if rank <= 250:
            return 80.0
        if rank <= 500:
            return 50.0
        if rank <= 1000:
            return 30.0
        if rank <= 3000:
            return 15.0
        if rank <= 10000:
            return 5.0
        return 1.0
    if rank <= 10:
        return 20.0
    if rank <= 50:
        return 10.0
    if rank <= 200:
        return 5.0
    if rank <= 1000:
        return 2.0
    return 0.5


def _extract_monthly_orders(text: str) -> float | None:
    lowered = text.lower()
    values: list[float] = []
    for pattern in MONTHLY_ORDER_PATTERNS:
        for found in re.finditer(pattern, lowered, flags=re.I):
            try:
                raw = found.group(1).replace(",", "").strip().lower()
                whole = found.group(0).lower()
                base = float(raw[:-1]) if raw.endswith("k") else float(raw)
                if raw.endswith("k") or re.search(r"\d\s*k\+?", whole):
                    values.append(base * 1000.0)
                else:
                    values.append(base)
            except Exception:
                continue
    if values:
        return max(values)
    weekly_values: list[float] = []
    for pattern in WEEKLY_ORDER_PATTERNS:
        for found in re.finditer(pattern, lowered, flags=re.I):
            raw = found.group(1).replace(",", "").lower()
            try:
                value = float(raw[:-1]) * 1000 if raw.endswith("k") else float(raw)
            except Exception:
                continue
            weekly_values.append(value * (30.0 / 7.0))
    if weekly_values:
        return max(weekly_values)
    for pattern in SOLD_PATTERNS:
        for found in re.finditer(pattern, lowered, flags=re.I):
            raw = found.group(1).replace(",", "").lower()
            try:
                value = float(raw[:-1]) * 1000 if raw.endswith("k") else float(raw)
            except Exception:
                continue
            # Treat "sold" as lower-confidence rolling demand proxy over roughly 90 days.
            values.append(value / 3.0)
    if values:
        return max(values)
    return None


def _extract_temu_sold_count(text: str) -> float | None:
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*sold", text, flags=re.I)
    values: list[float] = []
    for raw in matches:
        try:
            values.append(float(raw))
        except Exception:
            continue
    return max(values) if values else None


def _extract_listing_age_days(text: str) -> int | None:
    for pattern in LISTING_AGE_PATTERNS:
        found = re.search(pattern, text, flags=re.I)
        if not found:
            continue
        try:
            raw = found.group(1).strip()
            if "/" in raw:
                if len(raw.split("/", 1)[0]) == 4:
                    dt = datetime.strptime(raw, "%Y/%m/%d").replace(tzinfo=timezone.utc)
                else:
                    dt = datetime.strptime(raw, "%m/%d/%Y").replace(tzinfo=timezone.utc)
            elif "-" in raw and raw[:4].isdigit():
                dt = datetime.strptime(raw, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            else:
                dt = datetime.strptime(raw, "%B %d, %Y").replace(tzinfo=timezone.utc)
        except Exception:
            continue
        return max(0, int((_utc_now() - dt).total_seconds() // 86400))
    return None


def _pain_point_score(text: str) -> float:
    lowered = text.lower()
    hits = sum(1 for pattern in PAIN_POINT_PATTERNS if re.search(pattern, lowered, flags=re.I))
    return min(100.0, hits * 12.5)


def _brand_hint(title: str) -> str:
    tokens = [token for token in re.findall(r"[A-Za-z0-9]+", title) if len(token) >= 2]
    cleaned = [token.lower() for token in tokens if token.lower() not in TITLE_STOPWORDS]
    if not cleaned:
        return "unknown"
    return cleaned[0]


def _competition_density_hints(candidates: list[DemandCandidate]) -> dict[str, dict[str, Any]]:
    by_market: dict[tuple[str, str], list[DemandCandidate]] = {}
    for item in candidates:
        key = (item.sell_platform.lower(), _query_key(item.query))
        by_market.setdefault(key, []).append(item)
    hints: dict[str, dict[str, Any]] = {}
    for rows in by_market.values():
        candidate_count = len(rows)
        head_rows = rows[: min(5, candidate_count)]
        head_count = len(head_rows)
        review_values = sorted(int(item.review_count or 0) for item in rows if item.review_count is not None)
        if review_values:
            median_reviews = review_values[len(review_values) // 2]
        else:
            median_reviews = 0
        brand_counts: dict[str, int] = {}
        normalized_titles = [_normalize_title(item.title) for item in rows]
        token_frequency: dict[str, int] = {}
        head_brand_counts: dict[str, int] = {}
        head_brand_reviews: dict[str, float] = {}
        head_phrase_counts: dict[str, int] = {}
        for item in rows:
            brand = _brand_hint(item.title)
            brand_counts[brand] = int(brand_counts.get(brand, 0) or 0) + 1
        for item in head_rows:
            brand = _brand_hint(item.title)
            head_brand_counts[brand] = int(head_brand_counts.get(brand, 0) or 0) + 1
            head_brand_reviews[brand] = float(head_brand_reviews.get(brand, 0.0) or 0.0) + float(item.review_count or 0)
            phrase = " ".join(_normalize_title(item.title).split()[:3]).strip()
            if phrase:
                head_phrase_counts[phrase] = int(head_phrase_counts.get(phrase, 0) or 0) + 1
        for title in normalized_titles:
            for token in set(title.split()):
                if len(token) < 4:
                    continue
                token_frequency[token] = int(token_frequency.get(token, 0) or 0) + 1
        dominant_brand_share = max(brand_counts.values()) / candidate_count if brand_counts and candidate_count else 0.0
        repeated_token_share = (
            max(token_frequency.values()) / candidate_count
            if token_frequency and candidate_count
            else 0.0
        )
        ad_density = sum(1 for item in rows if "sponsored" in _normalize_title(item.title)) / candidate_count if candidate_count else 0.0
        head_brand_share = max(head_brand_counts.values()) / head_count if head_brand_counts and head_count else 0.0
        total_head_reviews = sum(float(item.review_count or 0) for item in head_rows)
        head_review_share = (
            max(head_brand_reviews.values()) / total_head_reviews
            if head_brand_reviews and total_head_reviews > 0
            else head_brand_share
        )
        head_phrase_share = max(head_phrase_counts.values()) / head_count if head_phrase_counts and head_count else 0.0
        head_monopoly_score = min(
            100.0,
            head_brand_share * 42.0
            + head_review_share * 34.0
            + head_phrase_share * 16.0
            + ad_density * 8.0,
        )
        for item in rows:
            hints[item.candidate_id] = {
                "query_candidate_count": candidate_count,
                "query_median_reviews": median_reviews,
                "dominant_brand_share": round(dominant_brand_share, 4),
                "repeated_token_share": round(repeated_token_share, 4),
                "ad_density": round(ad_density, 4),
                "head_result_count": head_count,
                "head_brand_share": round(head_brand_share, 4),
                "head_review_share": round(head_review_share, 4),
                "head_phrase_share": round(head_phrase_share, 4),
                "head_monopoly_score": round(head_monopoly_score, 2),
            }
    return hints


def _competition_score(text: str, *, review_count: int | None = None, raw_signals: dict[str, Any] | None = None) -> float:
    lowered = text.lower()
    ad_hits = len(re.findall(r"sponsored", lowered))
    top_brand_hits = len(re.findall(r"brand", lowered))
    duplicate_hits = len(re.findall(r"best seller", lowered))
    base_penalty = min(60.0, ad_hits * 8 + top_brand_hits * 4 + duplicate_hits * 6)
    raw_signals = raw_signals or {}
    query_candidate_count = float(raw_signals.get("query_candidate_count", 0) or 0)
    query_median_reviews = float(raw_signals.get("query_median_reviews", 0) or 0)
    dominant_brand_share = float(raw_signals.get("dominant_brand_share", 0) or 0)
    repeated_token_share = float(raw_signals.get("repeated_token_share", 0) or 0)
    ad_density = float(raw_signals.get("ad_density", 0) or 0)
    head_monopoly_score = float(raw_signals.get("head_monopoly_score", 0) or 0)
    density_penalty = min(22.0, max(0.0, query_candidate_count - 3.0) * 3.0)
    review_penalty = min(18.0, query_median_reviews / 600.0)
    if query_candidate_count >= 3:
        concentration_penalty = min(20.0, dominant_brand_share * 22.0 + repeated_token_share * 12.0 + ad_density * 10.0)
    else:
        concentration_penalty = min(8.0, ad_density * 10.0)
    monopoly_penalty = min(22.0, head_monopoly_score * 0.22) if query_candidate_count >= 3 else 0.0
    review_tail_penalty = min(10.0, (float(review_count or 0) / 2500.0)) if review_count is not None else 0.0
    penalty = min(85.0, base_penalty + density_penalty + review_penalty + concentration_penalty + monopoly_penalty + review_tail_penalty)
    return max(12.0, 100.0 - penalty)


def _price_stability_from_history(history: list[float]) -> float:
    cleaned = [float(x) for x in history if x is not None]
    if len(cleaned) < 3:
        return 50.0
    avg = sum(cleaned) / len(cleaned)
    if avg <= 0:
        return 0.0
    variance = sum((x - avg) ** 2 for x in cleaned) / len(cleaned)
    cv = (variance ** 0.5) / avg
    return max(0.0, min(100.0, 100.0 - cv * 200.0))


def _selection_scorecard(
    *,
    estimated_daily_orders: float,
    demand_confidence: float,
    sellersprite_keyword_signal_score: float,
    sellersprite_product_signal_score: float,
    competition_score: float,
    review_count: int | None,
    listing_age_days: int | None,
    margin: float | None,
    conservative_margin: float | None,
    weight_grade: str,
    price_stability_score: float,
    restricted: bool,
    source_match_score: float,
    supplier_confidence_score: float,
    query_source_kind: str = "base",
    adaptive_query_bias: float = 0.0,
) -> tuple[float, str, str, dict[str, float]]:
    demand_acceleration_score = min(
        100.0,
        estimated_daily_orders * 2.1
        + min(18.0, demand_confidence * 0.12)
        + min(14.0, max(sellersprite_keyword_signal_score, sellersprite_product_signal_score) * 0.18),
    )
    if review_count is None:
        review_inverse_score = 55.0
    elif review_count <= 150:
        review_inverse_score = 95.0
    elif review_count <= 500:
        review_inverse_score = 82.0
    elif review_count <= 1500:
        review_inverse_score = 66.0
    elif review_count <= 5000:
        review_inverse_score = 42.0
    else:
        review_inverse_score = 22.0
    if listing_age_days is None:
        freshness_score = 45.0
    elif listing_age_days <= 180:
        freshness_score = 100.0
    elif listing_age_days <= 365:
        freshness_score = 88.0
    elif listing_age_days <= 730:
        freshness_score = 62.0
    else:
        freshness_score = 24.0
    competition_inverse_score = round(
        min(100.0, competition_score * 0.55 + review_inverse_score * 0.25 + freshness_score * 0.20),
        2,
    )
    margin_score = min(100.0, max(0.0, (margin or 0.0) * 100.0 * 1.6)) if margin is not None else 0.0
    conservative_score = min(100.0, max(0.0, (conservative_margin or 0.0) * 100.0 * 1.7)) if conservative_margin is not None else 0.0
    weight_score = {"A": 100.0, "B": 82.0, "C": 58.0, "D": 25.0}.get(weight_grade, 35.0)
    profit_fulfillment_score = round(
        min(100.0, margin_score * 0.38 + conservative_score * 0.32 + weight_score * 0.18 + price_stability_score * 0.12),
        2,
    )
    query_source_bonus = {
        "seed": 12.0,
        "shared": 8.0,
        "base": 0.0,
    }.get(query_source_kind, 0.0)
    trend_source_score = round(
        min(
            100.0,
            sellersprite_keyword_signal_score * 0.48
            + sellersprite_product_signal_score * 0.34
            + min(12.0, adaptive_query_bias * 3.0)
            + query_source_bonus,
        ),
        2,
    )
    compliance_score = 22.0 if restricted else 100.0
    if weight_grade == "D":
        compliance_score = min(compliance_score, 35.0)
    elif weight_grade == "C":
        compliance_score = min(compliance_score, 72.0)
    execution_resilience_score = round(
        min(
            100.0,
            source_match_score * 0.34
            + supplier_confidence_score * 0.30
            + weight_score * 0.24
            + price_stability_score * 0.08
            + (10.0 if query_source_kind in {"seed", "shared"} else 0.0),
        ),
        2,
    )
    final_score = round(
        demand_acceleration_score * 0.28
        + trend_source_score * 0.17
        + competition_inverse_score * 0.18
        + profit_fulfillment_score * 0.20
        + execution_resilience_score * 0.09
        + compliance_score * 0.08,
        2,
    )
    if final_score >= 78.0:
        grade = "strong"
    elif final_score >= 64.0:
        grade = "medium"
    else:
        grade = "watch"
    strengths: list[str] = []
    if demand_acceleration_score >= 72.0:
        strengths.append("accelerating_demand")
    if trend_source_score >= 56.0:
        strengths.append("validated_trend")
    if competition_inverse_score >= 68.0:
        strengths.append("competition_gap")
    if profit_fulfillment_score >= 68.0:
        strengths.append("profitfulfillment_ok")
    if execution_resilience_score >= 66.0:
        strengths.append("execution_resilient")
    if compliance_score >= 85.0:
        strengths.append("compliance_safe")
    thesis = "+".join(strengths[:3]) if strengths else "needs_more_validation"
    return final_score, grade, thesis, {
        "demand_acceleration_score": round(demand_acceleration_score, 2),
        "trend_source_score": trend_source_score,
        "competition_inverse_score": competition_inverse_score,
        "profit_fulfillment_score": profit_fulfillment_score,
        "execution_resilience_score": execution_resilience_score,
        "compliance_score": round(compliance_score, 2),
    }


def _platform_screen_profile(
    *,
    platform: str,
    estimated_daily_orders: float,
    demand_score: float,
    competition_score: float,
    differentiation_score: float,
    price_stability_score: float,
    margin: float | None,
    conservative_margin: float | None,
    listing_age_days: int | None,
    review_count: int | None,
    confidence: float,
    restricted: bool,
    sellersprite_keyword_signal_score: float = 0.0,
    sellersprite_product_signal_score: float = 0.0,
    adaptive_platform_bias: float = 0.0,
    adaptive_query_bias: float = 0.0,
) -> tuple[float, str, str]:
    margin_score = min(100.0, max(0.0, (margin or 0.0) * 100.0 * 1.5)) if margin is not None else 0.0
    conservative_score = min(100.0, max(0.0, (conservative_margin or 0.0) * 100.0 * 1.5)) if conservative_margin is not None else 0.0
    review_signal = min(100.0, (review_count or 0) / 250.0)
    freshness_score = 0.0 if listing_age_days is None else (100.0 if listing_age_days <= 365 else (75.0 if listing_age_days <= 730 else 20.0))
    platform = platform.lower()
    if platform == "amazon":
        fit = (
            demand_score * 0.28
            + competition_score * 0.20
            + margin_score * 0.18
            + conservative_score * 0.12
            + confidence * 0.10
            + freshness_score * 0.07
            + min(100.0, estimated_daily_orders * 1.5) * 0.05
            + sellersprite_keyword_signal_score * 0.08
            + sellersprite_product_signal_score * 0.10
        )
    elif platform == "walmart":
        fit = (
            demand_score * 0.24
            + competition_score * 0.18
            + margin_score * 0.18
            + conservative_score * 0.12
            + price_stability_score * 0.10
            + confidence * 0.10
            + freshness_score * 0.08
            + sellersprite_keyword_signal_score * 0.08
            + sellersprite_product_signal_score * 0.08
        )
    elif platform == "temu":
        fit = (
            demand_score * 0.18
            + differentiation_score * 0.22
            + margin_score * 0.18
            + conservative_score * 0.12
            + confidence * 0.12
            + review_signal * 0.10
            + freshness_score * 0.08
            + sellersprite_keyword_signal_score * 0.08
            + sellersprite_product_signal_score * 0.06
        )
    elif platform == "tiktok":
        virality_proxy = min(100.0, demand_score * 0.8 + differentiation_score * 0.6 + review_signal * 0.3)
        fit = (
            virality_proxy * 0.30
            + differentiation_score * 0.22
            + margin_score * 0.16
            + conservative_score * 0.10
            + confidence * 0.10
            + freshness_score * 0.06
            + competition_score * 0.06
            + sellersprite_keyword_signal_score * 0.18
            + sellersprite_product_signal_score * 0.08
        )
    else:
        fit = (
            demand_score * 0.25
            + competition_score * 0.20
            + differentiation_score * 0.15
            + margin_score * 0.15
            + conservative_score * 0.10
            + confidence * 0.10
            + freshness_score * 0.05
            + sellersprite_keyword_signal_score * 0.08
            + sellersprite_product_signal_score * 0.07
        )
    if restricted:
        fit = min(fit, 35.0)
    fit += adaptive_platform_bias
    fit += adaptive_query_bias
    fit = round(max(0.0, min(100.0, fit)), 2)
    if fit >= 80.0:
        return fit, "strong", "priority_test"
    if fit >= 65.0:
        return fit, "medium", "watch_and_validate"
    return fit, "watch", "watchlist_only"


def _unit_to_kg(value: float, unit: str) -> float:
    unit = unit.lower()
    if unit in {"g", "克"}:
        return value / 1000.0
    return value


def _extract_weight(text: str, platform: str = "") -> tuple[float | None, str]:
    lowered = text.lower()
    if platform == "made_in_china":
        contextual = re.search(
            r"(?:package gross weight|gross weight|net weight)[^0-9]{0,80}(\d+(?:\.\d+)?)\s*(kg|g|公斤|克)",
            lowered,
            flags=re.I,
        )
        if contextual:
            return _unit_to_kg(float(contextual.group(1)), contextual.group(2)), "A"
    if platform == "yiwugo":
        gross = re.search(r"g\.w\./ctn[^0-9]{0,40}(\d+(?:\.\d+)?)\s*(kg|g|公斤|克)", lowered, flags=re.I)
        qty = re.search(r"qty/ctn[^0-9]{0,40}(\d+(?:\.\d+)?)\s*piece", lowered, flags=re.I)
        if gross and qty:
            gross_kg = _unit_to_kg(float(gross.group(1)), gross.group(2))
            qty_value = float(qty.group(1))
            if qty_value > 0:
                return round(gross_kg / qty_value, 4), "B"
    for pattern in WEIGHT_PATTERNS:
        match = re.search(pattern, lowered, flags=re.I)
        if not match:
            continue
        value = _unit_to_kg(float(match.group(1)), match.group(2))
        if 0.01 <= value <= 30:
            return value, "C"
    return None, "D"


def _extract_sell_candidates(platform: str, text: str, query: str) -> list[DemandCandidate]:
    rows: list[DemandCandidate] = []
    seen: set[str] = set()
    if platform == "amazon":
        matches = list(re.finditer(r'data-asin="([A-Z0-9]{10})"', text, flags=re.I))
        for idx_match, match in enumerate(matches):
            asin = match.group(1)
            link = f"https://www.amazon.com/dp/{asin}"
            if link in seen:
                continue
            seen.add(link)
            start = match.start()
            end = matches[idx_match + 1].start() if idx_match + 1 < len(matches) else min(len(text), start + 25000)
            snippet = text[start:end]
            title_match = re.search(r'<span[^>]*>([^<]{20,240})</span>', snippet, flags=re.I)
            title = _clean_text(title_match.group(1)) if title_match else query.title()
            if title.lower() in {"amazon", "prime", "limited time deal"}:
                title = query.title()
            price = _best_price_cny(snippet, platform)
            monthly_orders = _extract_monthly_orders(snippet)
            amazon_rank, amazon_category = _extract_amazon_best_seller_rank(snippet)
            amazon_rank_daily_proxy = _estimate_amazon_daily_orders_from_rank(amazon_rank, amazon_category)
            estimated_daily_orders = (monthly_orders / 30.0) if monthly_orders else amazon_rank_daily_proxy
            rating_value = _extract_rating_value(snippet)
            review_count = _extract_review_count(snippet)
            rows.append(
                DemandCandidate(
                    candidate_id=f"{platform}-{asin.lower()}",
                    title=title,
                    sell_platform=platform,
                    sell_link=link,
                    sell_price_cny=price,
                    query=query,
                    extracted_at=_utc_now_iso(),
                    estimated_daily_orders=estimated_daily_orders,
                    rating_value=rating_value,
                    review_count=review_count,
                    demand_confidence=65.0 if monthly_orders else (45.0 if amazon_rank_daily_proxy else 35.0),
                    raw_signals={
                        "price_extracted": price is not None,
                        "monthly_orders": monthly_orders,
                        "amazon_best_seller_rank": amazon_rank,
                        "amazon_best_seller_category": amazon_category,
                        "amazon_rank_daily_proxy": amazon_rank_daily_proxy,
                        "source": "amazon_search",
                    },
                )
            )
            if len(rows) >= 8:
                break
        if rows:
            return rows
    patterns = {
        "temu": r"(https://www\.temu\.com/[^\s\"'<>]+-s\.html(?:\?[^\s\"'<>]+)?|/[^\s\"'<>]+-s\.html(?:\?[^\s\"'<>]+)?)",
        "amazon": r"(https://www\.amazon\.com/dp/[A-Z0-9]{10}|/dp/[A-Z0-9]{10})",
        "walmart": r"(https://www\.walmart\.com/ip/[^\s\"'<>]+|/ip/[^\s\"'<>]+)",
    }
    for raw_link in re.findall(patterns.get(platform, r"$^"), text, flags=re.I):
        link = raw_link if raw_link.startswith("http") else {
            "temu": "https://www.temu.com",
            "amazon": "https://www.amazon.com",
            "walmart": "https://www.walmart.com",
        }[platform] + raw_link
        if platform == "temu" and link.rstrip("/") == "https://www.temu.com":
            continue
        if platform == "walmart" and "/search?" in link:
            continue
        if link in seen:
            continue
        seen.add(link)
        title = _title_from_link(link, query.title())
        price = _best_price_cny(text, platform)
        monthly_orders = _extract_monthly_orders(text)
        rating_value = _extract_rating_value(text)
        review_count = _extract_review_count(text)
        rows.append(
            DemandCandidate(
                candidate_id=f"{platform}-{_slug(link)}",
                title=title,
                sell_platform=platform,
                sell_link=link,
                sell_price_cny=price,
                query=query,
                extracted_at=_utc_now_iso(),
                estimated_daily_orders=(monthly_orders / 30.0) if monthly_orders else None,
                rating_value=rating_value,
                review_count=review_count,
                demand_confidence=55.0 if monthly_orders else 20.0,
                raw_signals={"price_extracted": price is not None, "monthly_orders": monthly_orders},
            )
        )
        if len(rows) >= 5:
            break
    if not rows:
        fallback_price = _best_price_cny(text, platform)
        rows.append(
            DemandCandidate(
                candidate_id=f"{platform}-{_slug(query)}",
                title=query.title(),
                sell_platform=platform,
                sell_link=url_for_sell(platform, query),
                sell_price_cny=fallback_price,
                query=query,
                extracted_at=_utc_now_iso(),
                estimated_daily_orders=None,
                rating_value=_extract_rating_value(text),
                review_count=_extract_review_count(text),
                demand_confidence=10.0,
                raw_signals={"fallback_from_search_page": True},
            )
        )
    return rows


def url_for_sell(platform: str, query: str) -> str:
    return SELL_PLATFORM_SEARCH[platform](query)


def _extract_source_candidate(platform: str, text: str, query: str, fetch_tool: str) -> SourceCandidate:
    price_cny = _extract_platform_price_cny(text, platform)
    weight, grade = _extract_weight(text[:120000], platform)
    raw_link = ""
    if platform == "1688":
        for candidate_link in re.findall(r"(https?://[^\s\"']+offer[^\s\"']+|/offer/[^\s\"']+)", text, flags=re.I):
            raw_link = candidate_link
            break
    elif platform == "yiwugo":
        for candidate_link in re.findall(r"(https?://en\.yiwugo\.com/product/detail/[^\s\"']+|/product/detail/[^\s\"']+)", text, flags=re.I):
            raw_link = candidate_link
            break
    elif platform == "made_in_china":
        candidates = re.findall(r"https?://[^\s\"']+made-in-china\.com/product/[^\s\"']+", text, flags=re.I)
        if not candidates:
            candidates = re.findall(r"/product/[^\s\"']+", text, flags=re.I)
        for candidate_link in candidates:
            if any(skip in candidate_link for skip in ("productdirectory.do", "/products-search/", "/search/product", "/suggest/")):
                continue
            raw_link = candidate_link
            break
    if raw_link and not raw_link.startswith("http"):
        raw_link = {
            "1688": "https://detail.1688.com",
            "yiwugo": "https://en.yiwugo.com",
            "made_in_china": "https://www.made-in-china.com",
        }[platform] + raw_link
    blocked = any(re.search(p, text, flags=re.I) for p in BLOCK_PATTERNS.get(platform, []))
    match_score = 35.0
    if price_cny is not None:
        match_score += 20
    if weight is not None:
        match_score += 25
    if raw_link:
        match_score += 10
    normalized_query = _normalize_title(query)
    normalized_text = _normalize_title(text[:4000])
    overlap = len(set(normalized_query.split()) & set(normalized_text.split()))
    match_score += min(10, overlap * 2)
    title = _source_title_from_link(raw_link) if raw_link else query.title()
    if not title or title.lower() == query.lower():
        page_title = _extract_page_title(text[:8000])
        if page_title:
            title = page_title
    return SourceCandidate(
        platform=platform,
        link=raw_link or SOURCE_PLATFORM_SEARCH[platform](query),
        title=title or query.title(),
        price_cny=price_cny,
        weight_kg=weight,
        weight_grade=grade,
        fetch_tool=fetch_tool,
        match_score=min(100.0, match_score),
        blocked=blocked,
        notes="blocked" if blocked else "",
    )


def _extract_source_links(platform: str, text: str) -> list[str]:
    raw_links: list[str] = []
    if platform == "1688":
        raw_links.extend(re.findall(r"(https?://[^\s\"']+offer[^\s\"']+|/offer/[^\s\"']+)", text, flags=re.I))
    elif platform == "yiwugo":
        raw_links.extend(re.findall(r"(https?://en\.yiwugo\.com/product/detail/[^\s\"']+|/product/detail/[^\s\"']+)", text, flags=re.I))
    elif platform == "made_in_china":
        raw_links.extend(re.findall(r"https?://[^\s\"']+made-in-china\.com/product/[^\s\"']+", text, flags=re.I))
        raw_links.extend(re.findall(r"/product/[^\s\"']+", text, flags=re.I))
    resolved: list[str] = []
    for raw_link in raw_links:
        if any(skip in raw_link for skip in ("productdirectory.do", "/products-search/", "/search/product", "/suggest/")):
            continue
        link = raw_link
        if not link.startswith("http"):
            link = {
                "1688": "https://detail.1688.com",
                "yiwugo": "https://en.yiwugo.com",
                "made_in_china": "https://www.made-in-china.com",
            }[platform] + link
        if link not in resolved:
            resolved.append(link)
    return resolved[:5]


def _source_title_from_link(link: str) -> str:
    cleaned = re.sub(r"https?://", "", str(link or ""), flags=re.I)
    cleaned = re.sub(r"[?#].*$", "", cleaned)
    slug = cleaned.rstrip("/").split("/")[-1]
    slug = re.sub(r"\.html?$", "", slug, flags=re.I)
    slug = re.sub(r"[-_]+", " ", slug)
    return _clean_text(slug)


def _candidate_overlap_score(query: str, text: str) -> float:
    query_tokens = set(_normalize_title(query).split())
    text_tokens = set(_normalize_title(text).split())
    if not query_tokens or not text_tokens:
        return 0.0
    overlap = len(query_tokens & text_tokens)
    coverage = overlap / max(1, len(query_tokens))
    return min(20.0, coverage * 20.0)


def _match_tokens(text: str) -> set[str]:
    tokens = {
        token
        for token in re.findall(r"[a-z0-9]+", str(text or "").lower())
        if token not in MATCH_TOKEN_STOPWORDS and len(token) >= 2
    }
    raw_text = str(text or "")
    for phrase, canonical_tokens in CHINESE_TOKEN_NORMALIZERS.items():
        if phrase not in raw_text:
            continue
        for token in canonical_tokens:
            if token not in MATCH_TOKEN_STOPWORDS:
                tokens.add(token)
    return tokens


def _attribute_signature(text: str) -> dict[str, set[str]]:
    raw_text = str(text or "")
    lowered = raw_text.lower()
    tokens = _match_tokens(lowered)
    return {
        "colors": tokens & COLOR_TOKENS,
        "materials": tokens & MATERIAL_TOKENS,
        "features": tokens & FEATURE_TOKENS,
        "numbers": set(re.findall(r"\b\d+(?:\.\d+)?\b", lowered)),
        "dimensions": {
            _clean_text(match.group(0))
            for match in re.finditer(
                r"\b\d+(?:\.\d+)?\s*(?:inch|in|cm|mm|oz|ml|lb|lbs|g|kg)\b|"
                r"\d+(?:\.\d+)?\s*(?:厘米|毫米|克|公斤|千克|斤|英寸)|"
                r"\d+(?:\.\d+)?\s*[x×*]\s*\d+(?:\.\d+)?(?:\s*[x×*]\s*\d+(?:\.\d+)?)?\s*(?:cm|mm|in|inch|厘米|毫米|英寸)",
                raw_text,
                flags=re.I,
            )
        },
    }


def _set_overlap_ratio(left: set[str], right: set[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / max(1, len(left | right))


def _price_band_similarity(sell_price_cny: float | None, source_price_cny: float | None) -> float:
    if sell_price_cny is None or source_price_cny is None or sell_price_cny <= 0 or source_price_cny <= 0:
        return 45.0
    ratio = source_price_cny / sell_price_cny
    if 0.08 <= ratio <= 0.72:
        return 100.0
    if 0.04 <= ratio < 0.08:
        return 72.0
    if 0.72 < ratio <= 0.90:
        return 68.0
    if 0.02 <= ratio < 0.04:
        return 42.0
    if 0.90 < ratio <= 1.15:
        return 30.0
    return 12.0


def _weight_reasonableness_score(weight_kg: float | None) -> float:
    if weight_kg is None or weight_kg <= 0:
        return 35.0
    if weight_kg <= 1.5:
        return 100.0
    if weight_kg <= 3.0:
        return 82.0
    if weight_kg <= 6.0:
        return 58.0
    return 24.0


def _estimated_weight_proxy_kg(candidate: DemandCandidate, source: SourceCandidate) -> float | None:
    text = " ".join(part for part in [candidate.title, candidate.query, source.title] if part).lower()
    tokens = _match_tokens(text)
    if not tokens:
        return None
    weight = 0.85
    if tokens & {"hook", "clip", "clips", "pill", "cable", "tray"}:
        weight = min(weight, 0.38)
    if tokens & {"bag", "tote", "insert", "box", "bin", "basket", "holder", "container", "organizer"}:
        weight = max(weight, 0.72)
    if tokens & {"drawer", "rack", "shelf", "tier", "under", "sink", "wall", "mounted"}:
        weight = max(weight, 1.15)
    if tokens & {"metal", "steel", "iron"}:
        weight += 0.45
    elif tokens & {"bamboo", "wood", "wooden"}:
        weight += 0.22
    elif tokens & {"fabric", "cotton", "linen", "mesh"}:
        weight -= 0.12
    if tokens & {"set", "pack", "bundle"}:
        weight += 0.18
    return round(max(0.18, min(4.5, weight)), 3)


def _supplier_fuzzy_match(
    candidate: DemandCandidate,
    source: SourceCandidate,
    *,
    source_text: str = "",
) -> dict[str, Any]:
    sell_text = " ".join(part for part in [candidate.title, candidate.query] if part)
    source_blob = " ".join(part for part in [source.title, source_text] if part)
    sell_tokens = _match_tokens(sell_text)
    source_tokens = _match_tokens(source_blob)
    keyword_coverage = len(sell_tokens & source_tokens) / max(1, len(sell_tokens)) if sell_tokens else 0.0
    keyword_similarity = round(min(100.0, keyword_coverage * 100.0), 2)

    sell_signature = _attribute_signature(sell_text)
    source_signature = _attribute_signature(source_blob)
    attribute_ratios = [
        _set_overlap_ratio(sell_signature["colors"], source_signature["colors"]),
        _set_overlap_ratio(sell_signature["materials"], source_signature["materials"]),
        _set_overlap_ratio(sell_signature["features"], source_signature["features"]),
        _set_overlap_ratio(sell_signature["numbers"], source_signature["numbers"]),
        _set_overlap_ratio(sell_signature["dimensions"], source_signature["dimensions"]),
    ]
    present_ratios = [ratio for ratio in attribute_ratios if ratio > 0]
    attribute_similarity = round((sum(present_ratios) / len(present_ratios) if present_ratios else 0.0) * 100.0, 2)

    price_similarity = _price_band_similarity(candidate.sell_price_cny, source.price_cny)
    weight_reasonableness = _weight_reasonableness_score(source.weight_kg)
    link_bonus = 100.0 if source.link and source.link.startswith("http") else 35.0
    similarity_score = round(
        min(
            100.0,
            keyword_similarity * 0.42
            + attribute_similarity * 0.24
            + price_similarity * 0.18
            + weight_reasonableness * 0.10
            + link_bonus * 0.06,
        ),
        2,
    )
    if similarity_score >= 74:
        grade = "high"
    elif similarity_score >= 58:
        grade = "medium"
    elif similarity_score >= 42:
        grade = "weak"
    else:
        grade = "none"
    reasons: list[str] = []
    if keyword_similarity >= 68:
        reasons.append("keyword_overlap_strong")
    elif keyword_similarity >= 45:
        reasons.append("keyword_overlap_partial")
    if _set_overlap_ratio(sell_signature["materials"], source_signature["materials"]) > 0:
        reasons.append("material_match")
    if _set_overlap_ratio(sell_signature["features"], source_signature["features"]) > 0:
        reasons.append("feature_match")
    if _set_overlap_ratio(sell_signature["dimensions"], source_signature["dimensions"]) > 0 or _set_overlap_ratio(sell_signature["numbers"], source_signature["numbers"]) > 0:
        reasons.append("size_or_pack_match")
    if price_similarity >= 68:
        reasons.append("plausible_price_band")
    if weight_reasonableness >= 82:
        reasons.append("weight_reasonable")
    return {
        "keyword_similarity": keyword_similarity,
        "attribute_similarity": attribute_similarity,
        "price_band_similarity": round(price_similarity, 2),
        "supplier_similarity_score": similarity_score,
        "supplier_similarity_grade": grade,
        "supplier_similarity_reasons": reasons,
    }


def _with_supplier_similarity(
    candidate: DemandCandidate,
    source: SourceCandidate,
    *,
    source_text: str = "",
) -> SourceCandidate:
    fuzzy = _supplier_fuzzy_match(candidate, source, source_text=source_text)
    similarity_score = float(fuzzy.get("supplier_similarity_score", 0) or 0)
    grade = str(fuzzy.get("supplier_similarity_grade", "none") or "none")
    grade_bonus = {"high": 10.0, "medium": 5.0, "weak": 1.5}.get(grade, 0.0)
    composite_score = round(min(100.0, source.match_score * 0.58 + similarity_score * 0.42 + grade_bonus), 2)
    return SourceCandidate(
        platform=source.platform,
        link=source.link,
        title=source.title,
        price_cny=source.price_cny,
        weight_kg=source.weight_kg,
        weight_grade=source.weight_grade,
        fetch_tool=source.fetch_tool,
        match_score=composite_score,
        keyword_similarity=float(fuzzy.get("keyword_similarity", 0) or 0),
        attribute_similarity=float(fuzzy.get("attribute_similarity", 0) or 0),
        price_band_similarity=float(fuzzy.get("price_band_similarity", 0) or 0),
        supplier_similarity_score=similarity_score,
        supplier_similarity_grade=grade,
        supplier_similarity_reasons=list(fuzzy.get("supplier_similarity_reasons") or []),
        blocked=source.blocked,
        notes=source.notes,
    )


def _apply_supplier_weight_proxy(candidate: DemandCandidate, source: SourceCandidate) -> SourceCandidate:
    if source.weight_kg is not None:
        return source
    if source.price_cny is None or source.supplier_similarity_grade not in {"high", "medium"}:
        return source
    proxy_weight = _estimated_weight_proxy_kg(candidate, source)
    if proxy_weight is None:
        return source
    proxy_grade = "C" if source.supplier_similarity_grade == "high" else "D"
    notes = "|".join(part for part in [source.notes, "weight_proxy"] if part)
    return SourceCandidate(
        platform=source.platform,
        link=source.link,
        title=source.title,
        price_cny=source.price_cny,
        weight_kg=proxy_weight,
        weight_grade=proxy_grade,
        fetch_tool=source.fetch_tool,
        match_score=min(100.0, source.match_score + (4.0 if source.supplier_similarity_grade == "high" else 1.5)),
        keyword_similarity=source.keyword_similarity,
        attribute_similarity=source.attribute_similarity,
        price_band_similarity=source.price_band_similarity,
        supplier_similarity_score=source.supplier_similarity_score,
        supplier_similarity_grade=source.supplier_similarity_grade,
        supplier_similarity_reasons=list(source.supplier_similarity_reasons or []),
        weight_proxy_used=True,
        blocked=source.blocked,
        notes=notes,
    )


def _supplier_confidence_score(source: SourceCandidate) -> float:
    base = float(source.match_score or 0)
    if source.blocked:
        base -= 28.0
    if source.link and source.link.startswith("http"):
        base += 8.0
    if source.weight_grade == "A":
        base += 18.0
    elif source.weight_grade == "B":
        base += 10.0
    elif source.weight_grade == "C":
        base += 2.0
    else:
        base -= 18.0
    platform_bonus = {
        "made_in_china": 12.0,
        "yiwugo": 6.0,
        "1688": 2.0,
    }.get(source.platform, 0.0)
    base += platform_bonus
    base += float(source.supplier_similarity_score or 0) * 0.18
    if source.supplier_similarity_grade == "high":
        base += 10.0
    elif source.supplier_similarity_grade == "medium":
        base += 4.0
    if source.weight_proxy_used:
        base -= 8.0
    if source.price_cny is not None:
        base += 6.0
    return round(max(0.0, min(100.0, base)), 2)


def _enrich_source_candidate(platform: str, candidate: SourceCandidate) -> SourceCandidate:
    if not candidate.link or (candidate.weight_kg is not None and candidate.price_cny is not None):
        return candidate
    if platform not in {"made_in_china", "yiwugo", "1688"}:
        return candidate
    detail = fetch_best(platform, candidate.link, max_tools=2)
    detail_weight, detail_grade = _extract_weight(detail.text[:120000], platform)
    detail_price = _extract_platform_price_cny(detail.text, platform)
    notes = candidate.notes
    if detail.status == "blocked":
        notes = (notes + " | detail_blocked").strip(" |")
    return SourceCandidate(
        platform=candidate.platform,
        link=candidate.link,
        title=candidate.title,
        price_cny=detail_price if detail_price is not None else candidate.price_cny,
        weight_kg=detail_weight if detail_weight is not None else candidate.weight_kg,
        weight_grade=detail_grade if detail_weight is not None else candidate.weight_grade,
        fetch_tool=f"{candidate.fetch_tool}+detail:{detail.tool}",
        match_score=min(100.0, candidate.match_score + (12.0 if detail_weight is not None else 0.0)),
        keyword_similarity=candidate.keyword_similarity,
        attribute_similarity=candidate.attribute_similarity,
        price_band_similarity=candidate.price_band_similarity,
        supplier_similarity_score=candidate.supplier_similarity_score,
        supplier_similarity_grade=candidate.supplier_similarity_grade,
        supplier_similarity_reasons=list(candidate.supplier_similarity_reasons or []),
        weight_proxy_used=candidate.weight_proxy_used,
        blocked=candidate.blocked and detail.status == "blocked",
        notes=notes,
    )


def _best_source_for_platform(
    platform: str,
    candidate: DemandCandidate,
    *,
    max_tools: int | None = None,
    platform_bias: float = 0.0,
    fast: bool = False,
) -> tuple[SourceCandidate, list[dict[str, Any]]]:
    fetch_log: list[dict[str, Any]] = []
    best: SourceCandidate | None = None
    for query_variant in _source_query_variants_for_platform(candidate, platform):
        url = SOURCE_PLATFORM_SEARCH[platform](query_variant)
        result = fetch_best(platform, url, max_tools=max_tools, fast=fast)
        fetch_log.append(
            {
                "stage": "match",
                "platform": platform,
                "query": query_variant,
                "tool": result.tool,
                "status": result.status,
                "score": result.score,
                "platform_bias": platform_bias,
            }
        )
        row = _extract_source_candidate(platform, result.text, query_variant, result.tool)
        row.match_score = min(100.0, row.match_score + platform_bias)
        row = _with_supplier_similarity(candidate, row, source_text=result.text[:8000])
        if not fast:
            row = _enrich_source_candidate(platform, row)
            row = _with_supplier_similarity(candidate, row, source_text=result.text[:8000])
        row = _apply_supplier_weight_proxy(candidate, row)
        if best is None or row.match_score > best.match_score:
            best = row
        extra_links = _extract_source_links(platform, result.text)
        detail_cap = 0 if fast else 2
        for detail_link in extra_links[:detail_cap]:
            detail_result = fetch_best(platform, detail_link, max_tools=2 if max_tools is None else max_tools, fast=fast)
            fetch_log.append(
                {
                    "stage": "match_detail",
                    "platform": platform,
                    "query": query_variant,
                    "link": detail_link,
                    "tool": detail_result.tool,
                    "status": detail_result.status,
                    "score": detail_result.score,
                    "platform_bias": platform_bias,
                }
            )
            detail_row = SourceCandidate(
                platform=platform,
                link=detail_link,
                title=_source_title_from_link(detail_link) or query_variant.title(),
                price_cny=_extract_platform_price_cny(detail_result.text, platform),
                weight_kg=_extract_weight(detail_result.text[:120000], platform)[0],
                weight_grade=_extract_weight(detail_result.text[:120000], platform)[1],
                fetch_tool=f"detail:{detail_result.tool}",
                match_score=min(
                    100.0,
                    35.0
                    + (20.0 if _extract_platform_price_cny(detail_result.text, platform) is not None else 0.0)
                    + (25.0 if _extract_weight(detail_result.text[:120000], platform)[0] is not None else 0.0)
                    + _candidate_overlap_score(query_variant, _source_title_from_link(detail_link) + " " + detail_result.text[:4000]),
                    + platform_bias,
                ),
                blocked=detail_result.status == "blocked",
                notes="detail_probe",
            )
            detail_row = _with_supplier_similarity(candidate, detail_row, source_text=detail_result.text[:8000])
            detail_row = _apply_supplier_weight_proxy(candidate, detail_row)
            if best is None or detail_row.match_score > best.match_score:
                best = detail_row
        if row.match_score >= 75 and row.weight_grade in {"A", "B"} and row.supplier_similarity_grade in {"high", "medium"} and not row.blocked:
            break
    return best or SourceCandidate(platform=platform, link=SOURCE_PLATFORM_SEARCH[platform](candidate.query), title=candidate.title, price_cny=None, weight_kg=None, weight_grade="D", fetch_tool="none", match_score=0.0, blocked=True, notes="no_source_probe"), fetch_log


def _enrich_sell_candidate(candidate: DemandCandidate, *, max_tools: int | None = None) -> tuple[DemandCandidate, dict[str, Any]]:
    tool_plan = {
        "temu": ["agent_browser", "playwright_stealth", "curl_cffi", "crawl4ai"],
        "amazon": ["curl_cffi", "agent_browser", "playwright_stealth", "direct_http", "crawl4ai"],
        "walmart": ["agent_browser", "playwright_stealth", "curl_cffi", "direct_http", "crawl4ai"],
    }.get(candidate.sell_platform, [fetch_best(candidate.sell_platform, candidate.sell_link, max_tools=max_tools).tool])
    if max_tools is not None:
        tool_plan = tool_plan[:max_tools]
    best_price = candidate.sell_price_cny
    best_monthly_orders = candidate.estimated_daily_orders * 30.0 if candidate.estimated_daily_orders else None
    best_listing_age = candidate.listing_age_days
    best_rating = candidate.rating_value
    best_reviews = candidate.review_count
    best_title = candidate.title
    best_confidence = candidate.demand_confidence
    best_status = "unknown"
    best_tool = "none"
    best_amazon_rank = (candidate.raw_signals or {}).get("amazon_best_seller_rank")
    best_amazon_category = (candidate.raw_signals or {}).get("amazon_best_seller_category")
    best_amazon_rank_daily_proxy = (candidate.raw_signals or {}).get("amazon_rank_daily_proxy")
    for tool in tool_plan:
        fetcher = TOOL_FETCHERS.get(tool)
        if not fetcher:
            continue
        try:
            text, err = fetcher(candidate.sell_link)
        except Exception:
            continue
        detail = _score_fetch(candidate.sell_platform, tool, text, err)
        best_status = detail.status
        best_tool = detail.tool
        price = _best_price_cny(detail.text, candidate.sell_platform)
        if price is not None:
            best_price = price
        monthly_orders = _extract_monthly_orders(detail.text)
        if monthly_orders is None and candidate.sell_platform == "temu":
            sold_proxy = _extract_temu_sold_count(detail.text)
            if sold_proxy:
                monthly_orders = sold_proxy
        if monthly_orders is not None:
            best_monthly_orders = max(best_monthly_orders or 0.0, monthly_orders)
            best_confidence = max(best_confidence, 70.0 if candidate.sell_platform == "temu" else 65.0)
        elif candidate.sell_platform == "amazon":
            amazon_rank, amazon_category = _extract_amazon_best_seller_rank(detail.text)
            amazon_rank_daily_proxy = _estimate_amazon_daily_orders_from_rank(amazon_rank, amazon_category)
            if amazon_rank_daily_proxy is not None:
                inferred_monthly_orders = amazon_rank_daily_proxy * 30.0
                best_monthly_orders = max(best_monthly_orders or 0.0, inferred_monthly_orders)
                best_confidence = max(best_confidence, 48.0)
                best_amazon_rank = amazon_rank
                best_amazon_category = amazon_category
                best_amazon_rank_daily_proxy = amazon_rank_daily_proxy
        listing_age_days = _extract_listing_age_days(detail.text)
        if listing_age_days is not None:
            best_listing_age = listing_age_days
            best_confidence = max(best_confidence, 70.0)
        rating_value = _extract_rating_value(detail.text)
        if rating_value is not None:
            best_rating = rating_value
        review_count = _extract_review_count(detail.text)
        if review_count is not None:
            best_reviews = review_count
        title = _extract_page_title(detail.text)
        if title:
            best_title = title
        if best_monthly_orders is not None and best_listing_age is not None:
            break
    updated = DemandCandidate(
        candidate_id=candidate.candidate_id,
        title=best_title,
        sell_platform=candidate.sell_platform,
        sell_link=candidate.sell_link,
        sell_price_cny=best_price,
        query=candidate.query,
        extracted_at=candidate.extracted_at,
        estimated_daily_orders=(best_monthly_orders / 30.0) if best_monthly_orders else candidate.estimated_daily_orders,
        listing_age_days=best_listing_age,
        rating_value=best_rating,
        review_count=best_reviews,
        demand_confidence=max(candidate.demand_confidence, best_confidence),
        demand_refresh_priority=candidate.demand_refresh_priority,
        raw_signals={
            **candidate.raw_signals,
            "detail_tool": best_tool,
            "detail_status": best_status,
            "detail_monthly_orders": best_monthly_orders,
            "amazon_best_seller_rank": best_amazon_rank,
            "amazon_best_seller_category": best_amazon_category,
            "amazon_rank_daily_proxy": best_amazon_rank_daily_proxy,
        },
    )
    return updated, {"stage": "sell_detail", "platform": candidate.sell_platform, "query": candidate.query, "tool": best_tool, "status": best_status, "score": 0}


def _is_restricted(text: str) -> tuple[bool, list[str]]:
    lowered = text.lower()
    hits = [pattern for pattern in RESTRICTED_PATTERNS if re.search(pattern, lowered, flags=re.I)]
    return bool(hits), hits


def _compute_decision(candidate: DemandCandidate, sources: list[SourceCandidate]) -> ArbitrageDecision:
    reasons: list[str] = []
    sellersprite_keyword_signal_score = float(((candidate.raw_signals or {}).get("sellersprite_keyword_signal_score", 0) or 0))
    sellersprite_product_signal_score = float(((candidate.raw_signals or {}).get("sellersprite_product_signal_score", 0) or 0))
    adaptive_platform_bias = float(((candidate.raw_signals or {}).get("adaptive_platform_bias", 0) or 0))
    adaptive_query_bias = float(((candidate.raw_signals or {}).get("adaptive_query_bias", 0) or 0))
    query_source_kind = str(((candidate.raw_signals or {}).get("query_source_kind", "base")) or "base")
    thresholds = (candidate.raw_signals or {}).get("adaptive_thresholds") or {}
    restricted, hits = _is_restricted(candidate.title)
    if restricted:
        reasons.append(f"restricted:{','.join(hits[:4])}")

    def _source_rank(source: SourceCandidate) -> float:
        return (
            float(source.match_score or 0) * 0.48
            + float(source.supplier_similarity_score or 0) * 0.32
            + (8.0 if source.weight_kg is not None else 0.0)
            + (6.0 if source.price_cny is not None else 0.0)
            + {"high": 8.0, "medium": 4.0, "weak": 1.0}.get(source.supplier_similarity_grade, 0.0)
        )

    best_source = None
    for source in sorted(sources, key=_source_rank, reverse=True):
        if source.blocked:
            continue
        if source.price_cny is None:
            continue
        if best_source is None or _source_rank(source) > _source_rank(best_source):
            best_source = source
    if best_source is None:
        return ArbitrageDecision(
            product_name=candidate.title,
            buy_platform="",
            buy_link="",
            sell_platform=candidate.sell_platform,
            sell_link=candidate.sell_link,
            sell_price_cny=candidate.sell_price_cny,
            purchase_cost_cny=None,
            weight_kg=None,
            gross_profit_amount=None,
            gross_margin_rate=None,
            conservative_margin_rate=None,
            platform_fee_rate=None,
            estimated_daily_orders=candidate.estimated_daily_orders,
            listing_age_days=candidate.listing_age_days,
        demand_score=0.0,
        competition_score=0.0,
        differentiation_score=0.0,
        price_stability_score=0.0,
        platform_fit_score=0.0,
        platform_fit_label="watch",
        platform_recommendation="watchlist_only",
        launchability_score=0.0,
        confidence_score=0.0,
        weight_grade="D",
            qualified=False,
            reasons=["no_usable_source_match", *reasons],
        )
    if best_source.weight_kg is None:
        return ArbitrageDecision(
            product_name=candidate.title,
            buy_platform=best_source.platform,
            buy_link=best_source.link,
            sell_platform=candidate.sell_platform,
            sell_link=candidate.sell_link,
            sell_price_cny=candidate.sell_price_cny,
            purchase_cost_cny=best_source.price_cny,
            weight_kg=None,
            gross_profit_amount=None,
            gross_margin_rate=None,
            conservative_margin_rate=None,
            platform_fee_rate=PLATFORM_FEE_TABLE.get(candidate.sell_platform, DEFAULT_PLATFORM_FEE),
            estimated_daily_orders=candidate.estimated_daily_orders,
            listing_age_days=candidate.listing_age_days,
            demand_score=0.0,
            competition_score=0.0,
            differentiation_score=0.0,
            price_stability_score=0.0,
            platform_fit_score=0.0,
            platform_fit_label="watch",
            platform_recommendation="watchlist_only",
            launchability_score=0.0,
            confidence_score=min(79.0, best_source.match_score),
            weight_grade=best_source.weight_grade,
            qualified=False,
            supplier_similarity_score=best_source.supplier_similarity_score,
            supplier_similarity_grade=best_source.supplier_similarity_grade,
            reasons=["missing_trusted_weight", f"supplier_similarity:{best_source.supplier_similarity_grade}", *reasons],
        )
    if candidate.sell_price_cny is None:
        return ArbitrageDecision(
            product_name=candidate.title,
            buy_platform=best_source.platform,
            buy_link=best_source.link,
            sell_platform=candidate.sell_platform,
            sell_link=candidate.sell_link,
            sell_price_cny=None,
            purchase_cost_cny=best_source.price_cny,
            weight_kg=best_source.weight_kg,
            gross_profit_amount=None,
            gross_margin_rate=None,
            conservative_margin_rate=None,
            platform_fee_rate=PLATFORM_FEE_TABLE.get(candidate.sell_platform, DEFAULT_PLATFORM_FEE),
            estimated_daily_orders=candidate.estimated_daily_orders,
            listing_age_days=candidate.listing_age_days,
            demand_score=0.0,
            competition_score=0.0,
            differentiation_score=0.0,
            price_stability_score=0.0,
            platform_fit_score=0.0,
            platform_fit_label="watch",
            platform_recommendation="watchlist_only",
            launchability_score=0.0,
            confidence_score=min(79.0, best_source.match_score),
            weight_grade=best_source.weight_grade,
            qualified=False,
            supplier_similarity_score=best_source.supplier_similarity_score,
            supplier_similarity_grade=best_source.supplier_similarity_grade,
            reasons=["missing_sell_price", f"supplier_similarity:{best_source.supplier_similarity_grade}", *reasons],
        )
    platform_fee_rate = PLATFORM_FEE_TABLE.get(candidate.sell_platform, DEFAULT_PLATFORM_FEE)
    platform_fee_amount = candidate.sell_price_cny * platform_fee_rate
    logistics = 59 * best_source.weight_kg + 35 * best_source.weight_kg + 1.4
    gross_profit = candidate.sell_price_cny - best_source.price_cny - logistics - platform_fee_amount
    margin = gross_profit / candidate.sell_price_cny if candidate.sell_price_cny else None
    conservative_sell = candidate.sell_price_cny * 0.95
    conservative_purchase = best_source.price_cny * 1.05
    conservative_weight = best_source.weight_kg * 1.10
    conservative_fee = conservative_sell * platform_fee_rate
    conservative_profit = conservative_sell - conservative_purchase - (59 * conservative_weight) - (35 * conservative_weight) - conservative_fee - 1.4
    conservative_margin = conservative_profit / conservative_sell if conservative_sell else None
    estimated_daily_orders = candidate.estimated_daily_orders or 0.0
    listing_age_days = candidate.listing_age_days
    head_monopoly_score = float(((candidate.raw_signals or {}).get("head_monopoly_score", 0) or 0))
    head_result_count = int(((candidate.raw_signals or {}).get("head_result_count", 0) or 0))
    demand_score = min(100.0, estimated_daily_orders * 2.0) if estimated_daily_orders else min(60.0, candidate.demand_confidence)
    if estimated_daily_orders < 30:
        reasons.append("estimated_daily_orders_below_threshold")
    if listing_age_days is None:
        reasons.append("listing_age_unknown")
    elif listing_age_days > 730:
        reasons.append("listing_age_above_two_years")
    if head_result_count >= 3 and head_monopoly_score >= 72.0:
        reasons.append("head_links_monopolized")
    if best_source.weight_proxy_used:
        reasons.append("supplier_weight_proxy_used")
    if best_source.supplier_similarity_grade == "weak":
        reasons.append("supplier_match_weak_similarity")
    elif best_source.supplier_similarity_grade == "medium":
        reasons.append("supplier_match_medium_similarity")
    competition_score = _competition_score(
        candidate.title,
        review_count=candidate.review_count,
        raw_signals=candidate.raw_signals or {},
    )
    differentiation_score = _pain_point_score(candidate.title)
    price_history = (candidate.raw_signals or {}).get("observed_sell_prices", []) or []
    if candidate.sell_price_cny is not None:
        price_history = [*price_history, candidate.sell_price_cny]
    price_stability_score = _price_stability_from_history(price_history)
    launchability_score = round(
        demand_score * 0.25
        + (min(100.0, max(0.0, (margin or 0.0) * 100.0 * 1.5)) if margin is not None else 0.0) * 0.25
        + competition_score * 0.20
        + differentiation_score * 0.15
        + price_stability_score * 0.15,
        2,
    )
    confidence = min(
        100.0,
        25.0
        + best_source.match_score * 0.35
        + (20.0 if best_source.weight_grade in {"A", "B"} else 0.0)
        + (10.0 if not restricted else 0.0)
        + (10.0 if margin is not None and margin >= 0.30 else 0.0)
        + (10.0 if conservative_margin is not None and conservative_margin >= 0.30 else 0.0),
    )
    supplier_confidence_score = _supplier_confidence_score(best_source)
    selection_score, selection_grade, selection_thesis, selection_components = _selection_scorecard(
        estimated_daily_orders=estimated_daily_orders,
        demand_confidence=candidate.demand_confidence,
        sellersprite_keyword_signal_score=sellersprite_keyword_signal_score,
        sellersprite_product_signal_score=sellersprite_product_signal_score,
        competition_score=competition_score,
        review_count=candidate.review_count,
        listing_age_days=listing_age_days,
        margin=margin,
        conservative_margin=conservative_margin,
        weight_grade=best_source.weight_grade,
        price_stability_score=price_stability_score,
        restricted=restricted,
        source_match_score=best_source.match_score,
        supplier_confidence_score=supplier_confidence_score,
        query_source_kind=query_source_kind,
        adaptive_query_bias=adaptive_query_bias,
    )
    platform_fit_score, platform_fit_label, platform_recommendation = _platform_screen_profile(
        platform=candidate.sell_platform,
        estimated_daily_orders=estimated_daily_orders,
        demand_score=demand_score,
        competition_score=competition_score,
        differentiation_score=differentiation_score,
        price_stability_score=price_stability_score,
        margin=margin,
        conservative_margin=conservative_margin,
        listing_age_days=listing_age_days,
        review_count=candidate.review_count,
        confidence=confidence,
        restricted=restricted,
        sellersprite_keyword_signal_score=sellersprite_keyword_signal_score,
        sellersprite_product_signal_score=sellersprite_product_signal_score,
        adaptive_platform_bias=adaptive_platform_bias,
        adaptive_query_bias=adaptive_query_bias,
    )
    qualified = (
        not restricted
        and estimated_daily_orders >= float(thresholds.get("order_floor", 30.0) or 30.0)
        and listing_age_days is not None
        and listing_age_days <= int(thresholds.get("listing_age_ceiling", 730) or 730)
        and margin is not None
        and conservative_margin is not None
        and margin >= float(thresholds.get("margin_floor", 0.30) or 0.30)
        and conservative_margin >= float(thresholds.get("conservative_margin_floor", 0.30) or 0.30)
        and best_source.weight_grade in {"A", "B"}
        and launchability_score >= float(thresholds.get("launchability_floor", 70.0) or 70.0)
        and confidence >= float(thresholds.get("confidence_floor", 80.0) or 80.0)
        and platform_fit_score >= float(thresholds.get("platform_fit_floor", 65.0) or 65.0)
        and (head_result_count < 3 or head_monopoly_score < 72.0)
    )
    decision_bucket, priority_reason = _decision_bucket(
        qualified=qualified,
        platform_fit_score=platform_fit_score,
        confidence=confidence,
        launchability_score=launchability_score,
        estimated_daily_orders=estimated_daily_orders,
        listing_age_days=listing_age_days,
        margin=margin,
        conservative_margin=conservative_margin,
        sellersprite_keyword_signal_score=sellersprite_keyword_signal_score,
        sellersprite_product_signal_score=sellersprite_product_signal_score,
        selection_score=selection_score,
        selection_grade=selection_grade,
        reasons=reasons,
    )
    if not qualified and not reasons:
        reasons.append("below_threshold_or_low_confidence")
    return ArbitrageDecision(
        product_name=candidate.title,
        buy_platform=best_source.platform,
        buy_link=best_source.link,
        sell_platform=candidate.sell_platform,
        sell_link=candidate.sell_link,
        sell_price_cny=round(candidate.sell_price_cny, 2),
        purchase_cost_cny=round(best_source.price_cny, 2),
        weight_kg=round(best_source.weight_kg, 4),
        gross_profit_amount=round(gross_profit, 2),
        gross_margin_rate=round(margin, 4) if margin is not None else None,
        conservative_margin_rate=round(conservative_margin, 4) if conservative_margin is not None else None,
        platform_fee_rate=platform_fee_rate,
        estimated_daily_orders=round(estimated_daily_orders, 2) if estimated_daily_orders else None,
        listing_age_days=listing_age_days,
        demand_score=round(demand_score, 2),
        competition_score=round(competition_score, 2),
        differentiation_score=round(differentiation_score, 2),
        price_stability_score=round(price_stability_score, 2),
        platform_fit_score=platform_fit_score,
        platform_fit_label=platform_fit_label,
        platform_recommendation=platform_recommendation,
        launchability_score=launchability_score,
        confidence_score=round(confidence, 2),
        weight_grade=best_source.weight_grade,
        qualified=qualified,
        selection_score=selection_score,
        selection_grade=selection_grade,
        selection_thesis=selection_thesis,
        query_source_kind=query_source_kind,
        trend_source_score=float(selection_components.get("trend_source_score", 0) or 0),
        execution_resilience_score=float(selection_components.get("execution_resilience_score", 0) or 0),
        supplier_confidence_score=supplier_confidence_score,
        supplier_similarity_score=best_source.supplier_similarity_score,
        supplier_similarity_grade=best_source.supplier_similarity_grade,
        head_monopoly_score=round(head_monopoly_score, 2),
        decision_bucket=decision_bucket,
        priority_reason=priority_reason,
        reasons=[
            *reasons,
            f"supplier_similarity_reasons:{','.join(best_source.supplier_similarity_reasons or [])}",
            f"selection_thesis:{selection_thesis}",
            f"selection_components:{json.dumps(selection_components, ensure_ascii=False)}",
        ],
    )


def _refresh_priority(row: dict[str, Any]) -> float:
    monthly_orders = float(((row.get("raw_signals", {}) or {}).get("monthly_orders", 0) or 0) or 0)
    detail_monthly_orders = float(((row.get("raw_signals", {}) or {}).get("detail_monthly_orders", 0) or 0) or 0)
    best_monthly_orders = max(monthly_orders, detail_monthly_orders)
    estimated_daily_orders = float(row.get("estimated_daily_orders", 0) or 0)
    review_count = float(row.get("review_count", 0) or 0)
    listing_age_days = row.get("listing_age_days")
    amazon_rank_daily_proxy = float(((row.get("raw_signals", {}) or {}).get("amazon_rank_daily_proxy", 0) or 0) or 0)
    score = 0.0
    if best_monthly_orders:
        score += min(60.0, best_monthly_orders / 20.0)
    elif estimated_daily_orders:
        score += min(50.0, estimated_daily_orders * 1.5)
    elif amazon_rank_daily_proxy:
        score += min(45.0, amazon_rank_daily_proxy * 1.5)
    score += min(20.0, review_count / 5000.0)
    if listing_age_days is None:
        score += 15.0
    elif listing_age_days <= 730:
        score += 10.0
    return round(score, 2)


def _ensure_refresh_priority(row: dict[str, Any]) -> float:
    priority = row.get("demand_refresh_priority")
    if priority is None:
        priority = _refresh_priority(row)
        row["demand_refresh_priority"] = priority
    return float(priority or 0.0)


def _load_state() -> dict[str, Any]:
    default = {"runs": [], "candidates": {}, "last_discovery_at": "", "last_match_at": "", "last_report_date": ""}
    if not STATE_PATH.exists():
        return default
    try:
        payload = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return default
    if not isinstance(payload, dict):
        return default
    if "candidates" not in payload:
        payload["candidates"] = {}
    payload.setdefault("runs", [])
    payload.setdefault("last_discovery_at", "")
    payload.setdefault("last_match_at", "")
    payload.setdefault("last_report_date", "")
    payload.pop("recent_candidates", None)
    for row in (payload.get("candidates") or {}).values():
        if not isinstance(row, dict):
            continue
        row.setdefault("raw_signals", {})
        row["demand_refresh_priority"] = _ensure_refresh_priority(row)
        decision = row.get("decision")
        if isinstance(decision, dict):
            row["decision"] = _decision_payload(decision)
    return payload


def _save_state(payload: dict[str, Any]) -> None:
    STATE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _decision_payload(payload: dict[str, Any]) -> dict[str, Any]:
    merged = dict(payload)
    for key, value in DECISION_DEFAULTS.items():
        merged.setdefault(key, value)
    merged.setdefault("reasons", [])
    return {key: merged.get(key) for key in ArbitrageDecision.__dataclass_fields__.keys()}


def _decision_from_payload(payload: dict[str, Any]) -> ArbitrageDecision:
    return ArbitrageDecision(**_decision_payload(payload))


def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def _sheet_name_for_platform(platform: str) -> str:
    mapping = {
        "amazon": "Amazon",
        "walmart": "Walmart",
        "temu": "Temu",
        "tiktok": "TikTok",
    }
    return mapping.get(platform.lower(), platform.title()[:31])


def _platform_sheet_rows(decisions: list[ArbitrageDecision], platform: str) -> list[ArbitrageDecision]:
    return [item for item in decisions if item.sell_platform.lower() == platform.lower()]


def _parse_json_stdout(raw: str) -> dict[str, Any]:
    text = str(raw or "").strip()
    if not text:
        return {}
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            return {}
    return {}


def _sellersprite_fetch(
    *,
    save_prefix: str,
    url: str | None = None,
    api_path: str | None = None,
    base_url: str | None = None,
    wait_seconds: float = 6.0,
) -> dict[str, Any]:
    cmd = [str(VENV_PY), str(SELLERSPRITE_FETCHER), "--save-prefix", save_prefix, "--wait-seconds", str(wait_seconds)]
    if api_path:
        cmd.extend(["--api-path", api_path])
        if base_url:
            cmd.extend(["--base-url", base_url])
        elif url:
            cmd.extend(["--base-url", url])
    elif url:
        cmd.extend(["--url", url])
    else:
        return {"ok": False, "error": "missing_url_or_api_path"}
    proc = _run(cmd, timeout=120)
    payload = _parse_json_stdout(proc.stdout)
    if not payload:
        return {
            "ok": False,
            "returncode": proc.returncode,
            "stdout": proc.stdout.strip(),
            "stderr": proc.stderr.strip(),
        }
    payload["ok"] = proc.returncode == 0
    payload["returncode"] = proc.returncode
    payload["stderr"] = proc.stderr.strip()
    return payload


def _sellersprite_probe_status(parsed: dict[str, Any]) -> tuple[str, str]:
    body = parsed.get("body")
    if isinstance(body, dict):
        code = str(body.get("code", "")).strip()
        if code and code != "0":
            return "error", code
        return "ok", code or "ok"
    if isinstance(body, list):
        return "ok", f"rows:{len(body)}"
    return "unknown", ""


def _clean_seed_phrase(value: str) -> str:
    phrase = re.sub(r"\s+", " ", str(value or "").strip())
    phrase = phrase.replace("“", '"').replace("”", '"').replace("’", "'")
    phrase = re.sub(r"[^A-Za-z0-9&+/' -]", " ", phrase)
    phrase = re.sub(r"\s+", " ", phrase).strip(" -")
    return phrase


def _valid_sellersprite_keyword(row: dict[str, Any]) -> bool:
    keyword = _clean_seed_phrase(str(row.get("keyword", "") or ""))
    if not keyword:
        return False
    tokens = [token for token in re.split(r"[^A-Za-z0-9]+", keyword.lower()) if token]
    if len(tokens) < 2:
        return False
    if len(keyword) < 8 or len(keyword) > 80:
        return False
    if any(token in {"richard", "高级会员", "会员", "seller", "sprite"} for token in tokens):
        return False
    monthly = str(row.get("monthly_searches", "") or "").replace(",", "").strip()
    if monthly and monthly.isdigit() and int(monthly) < 3000:
        return False
    return True


def _sellersprite_seed_queries(summary: dict[str, Any], limit: int = 6) -> list[str]:
    seeds: list[str] = []
    seen: set[str] = set()
    keyword_probe = summary.get("keyword_result_probe") or {}
    for row in (keyword_probe.get("top_keywords") or []):
        if not isinstance(row, dict) or not _valid_sellersprite_keyword(row):
            continue
        keyword = _clean_seed_phrase(str(row.get("keyword", "") or ""))
        normalized = keyword.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        seeds.append(keyword)
        if len(seeds) >= limit:
            break
    return seeds


def _sellersprite_product_watchlist(summary: dict[str, Any], limit: int = 5) -> list[dict[str, Any]]:
    watchlist: list[dict[str, Any]] = []
    product_probe = summary.get("product_result_probe") or {}
    for row in (product_probe.get("top_products") or []):
        if not isinstance(row, dict):
            continue
        name = _clean_seed_phrase(str(row.get("product_name", "") or ""))
        if len(name) < 8:
            continue
        watchlist.append(
            {
                "rank": row.get("rank"),
                "product_name": name,
                "asin": str(row.get("asin", "") or ""),
                "monthly_sales_hint": str(row.get("monthly_sales_hint", "") or ""),
                "sales_amount_hint": str(row.get("sales_amount_hint", "") or ""),
                "price_hint": str(row.get("price_hint", "") or ""),
                "review_count_hint": str(row.get("review_count_hint", "") or ""),
                "rating_hint": str(row.get("rating_hint", "") or ""),
                "listing_date_hint": str(row.get("listing_date_hint", "") or ""),
            }
        )
        if len(watchlist) >= limit:
            break
    return watchlist


def _merge_queries(base_queries: list[str], sellersprite_queries: list[str], limit: int | None = None) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for query in [*(base_queries or []), *(sellersprite_queries or [])]:
        cleaned = _clean_seed_phrase(query)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        merged.append(cleaned)
        if limit is not None and len(merged) >= limit:
            break
    return merged


def _compact_number_to_float(value: str | None) -> float:
    raw = str(value or "").strip().upper().replace(",", "")
    if not raw:
        return 0.0
    multiplier = 1.0
    if raw.endswith("K+"):
        raw = raw[:-2]
        multiplier = 1000.0
    elif raw.endswith("M+"):
        raw = raw[:-2]
        multiplier = 1000000.0
    elif raw.endswith("B+"):
        raw = raw[:-2]
        multiplier = 1000000000.0
    elif raw.endswith("K"):
        raw = raw[:-1]
        multiplier = 1000.0
    elif raw.endswith("M"):
        raw = raw[:-1]
        multiplier = 1000000.0
    elif raw.endswith("B"):
        raw = raw[:-1]
        multiplier = 1000000000.0
    try:
        return float(raw) * multiplier
    except Exception:
        return 0.0


def _token_set(value: str) -> set[str]:
    return {token for token in re.split(r"[^a-z0-9]+", _normalize_title(value).lower()) if token}


def _overlap_ratio(a: str, b: str) -> float:
    a_tokens = _token_set(a)
    b_tokens = _token_set(b)
    if not a_tokens or not b_tokens:
        return 0.0
    overlap = len(a_tokens & b_tokens)
    return overlap / float(max(len(a_tokens), len(b_tokens)))


def _listing_days_from_raw_date(value: str) -> int | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%B %d, %Y"):
        try:
            dt = datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except Exception:
            continue
        return max(0, int((_utc_now() - dt).total_seconds() // 86400))
    return None


def _sellersprite_candidate_hints(summary: dict[str, Any], candidate: DemandCandidate) -> dict[str, Any]:
    keyword_probe = summary.get("keyword_result_probe") or {}
    product_watchlist = summary.get("product_watchlist") or []
    query_phrase = candidate.query
    title_phrase = candidate.title
    keyword_matches = 0
    best_keyword_monthly_searches = 0.0
    best_keyword_overlap = 0.0
    for row in (keyword_probe.get("top_keywords") or []):
        if not isinstance(row, dict) or not _valid_sellersprite_keyword(row):
            continue
        keyword = str(row.get("keyword", "") or "")
        overlap = max(_overlap_ratio(query_phrase, keyword), _overlap_ratio(title_phrase, keyword))
        if overlap < 0.34:
            continue
        keyword_matches += 1
        best_keyword_overlap = max(best_keyword_overlap, overlap)
        best_keyword_monthly_searches = max(
            best_keyword_monthly_searches,
            _compact_number_to_float(str(row.get("monthly_searches", "") or "")),
        )
    product_matches = 0
    best_product_monthly_sales = 0.0
    best_product_revenue = 0.0
    best_product_rating = 0.0
    freshest_listing_days = None
    for row in product_watchlist:
        if not isinstance(row, dict):
            continue
        product_name = str(row.get("product_name", "") or "")
        overlap = max(_overlap_ratio(query_phrase, product_name), _overlap_ratio(title_phrase, product_name))
        if overlap < 0.28:
            continue
        product_matches += 1
        best_product_monthly_sales = max(best_product_monthly_sales, _compact_number_to_float(row.get("monthly_sales_hint")))
        best_product_revenue = max(best_product_revenue, _compact_number_to_float(row.get("sales_amount_hint")))
        try:
            best_product_rating = max(best_product_rating, float(str(row.get("rating_hint", "") or "0")))
        except Exception:
            pass
        listing_date = str(row.get("listing_date_hint", "") or "")
        listing_days = _listing_days_from_raw_date(listing_date) or _extract_listing_age_days(listing_date)
        if listing_days is not None:
            freshest_listing_days = listing_days if freshest_listing_days is None else min(freshest_listing_days, listing_days)
    keyword_signal_score = min(
        100.0,
        keyword_matches * 18.0
        + min(45.0, best_keyword_monthly_searches / 2500.0)
        + best_keyword_overlap * 20.0,
    )
    product_signal_score = min(
        100.0,
        product_matches * 18.0
        + min(35.0, best_product_monthly_sales / 4000.0)
        + min(20.0, best_product_revenue / 50000.0),
    )
    product_daily_orders_proxy = round(best_product_monthly_sales / 30.0, 2) if best_product_monthly_sales > 0 else 0.0
    return {
        "sellersprite_keyword_matches": keyword_matches,
        "sellersprite_keyword_monthly_searches": round(best_keyword_monthly_searches, 2),
        "sellersprite_keyword_overlap": round(best_keyword_overlap, 4),
        "sellersprite_keyword_signal_score": round(keyword_signal_score, 2),
        "sellersprite_product_matches": product_matches,
        "sellersprite_product_monthly_sales": round(best_product_monthly_sales, 2),
        "sellersprite_product_daily_orders_proxy": product_daily_orders_proxy,
        "sellersprite_product_revenue": round(best_product_revenue, 2),
        "sellersprite_product_rating": round(best_product_rating, 2),
        "sellersprite_product_freshest_listing_days": freshest_listing_days,
        "sellersprite_product_signal_score": round(product_signal_score, 2),
    }


def _decision_bucket(
    *,
    qualified: bool,
    platform_fit_score: float,
    confidence: float,
    launchability_score: float,
    estimated_daily_orders: float,
    listing_age_days: int | None,
    margin: float | None,
    conservative_margin: float | None,
    sellersprite_keyword_signal_score: float,
    sellersprite_product_signal_score: float,
    selection_score: float,
    selection_grade: str,
    reasons: list[str],
) -> tuple[str, str]:
    if qualified:
        return "qualified", "passed_strong_thresholds"
    combined_signal = max(sellersprite_keyword_signal_score, sellersprite_product_signal_score)
    if selection_score >= 74.0 and selection_grade in {"strong", "medium"}:
        return "near_miss", "high_selection_score_needs_validation"
    if (
        platform_fit_score >= 60.0
        and confidence >= 70.0
        and launchability_score >= 58.0
        and margin is not None
        and margin >= 0.30
        and conservative_margin is not None
        and conservative_margin >= 0.25
        and estimated_daily_orders >= 15.0
        and (listing_age_days is None or listing_age_days <= 900)
    ):
        if combined_signal >= 45.0:
            return "near_miss", "strong_sellersprite_signal_needs_validation"
        return "near_miss", "close_to_threshold"
    if combined_signal >= 55.0 or selection_score >= 62.0:
        return "watchlist", "high_signal_watchlist"
    if any(reason in {"listing_age_unknown", "estimated_daily_orders_below_threshold"} for reason in reasons):
        return "watchlist", "demand_or_age_evidence_gap"
    return "watchlist", "low_confidence_or_margin"


def _state_decisions(state: dict[str, Any]) -> list[ArbitrageDecision]:
    decisions: list[ArbitrageDecision] = []
    for payload in (state.get("candidates") or {}).values():
        decision = payload.get("decision")
        if isinstance(decision, dict):
            try:
                decisions.append(_decision_from_payload(decision))
            except Exception:
                continue
    return decisions


def _adaptive_profile_from_state(state: dict[str, Any]) -> dict[str, Any]:
    decisions = _state_decisions(state)
    failure_categories = _decision_failure_categories(decisions)
    primary_blocker = next(iter(failure_categories.keys()), "none")
    keyword_multiplier = 1.0
    product_multiplier = 1.0
    if primary_blocker in {"estimated_daily_orders_below_threshold", "listing_age_unknown"}:
        keyword_multiplier += 0.18
        product_multiplier += 0.12
    elif primary_blocker in {"no_usable_source_match", "missing_trusted_weight"}:
        keyword_multiplier += 0.05
        product_multiplier += 0.15
    platform_biases: dict[str, float] = {}
    for platform in ["amazon", "walmart", "temu", "tiktok"]:
        rows = [item for item in decisions if item.sell_platform.lower() == platform]
        if not rows:
            platform_biases[platform] = 0.0
            continue
        near_miss = sum(1 for item in rows if item.decision_bucket == "near_miss")
        qualified = sum(1 for item in rows if item.decision_bucket == "qualified")
        bias = min(8.0, near_miss * 1.5 + qualified * 0.8)
        platform_biases[platform] = round(bias, 2)
    source_biases: dict[str, float] = {}
    source_scores: dict[str, float] = {}
    for source in ["made_in_china", "1688", "yiwugo"]:
        rows = [item for item in decisions if str(item.buy_platform or "").lower() == source]
        qualified = sum(1 for item in rows if item.decision_bucket == "qualified")
        near_miss = sum(1 for item in rows if item.decision_bucket == "near_miss")
        watchlist = sum(1 for item in rows if item.decision_bucket == "watchlist")
        bias = min(10.0, qualified * 2.0 + near_miss * 1.0 - watchlist * 0.3)
        if primary_blocker == "missing_trusted_weight" and source == "made_in_china":
            bias += 1.5
        if primary_blocker == "no_usable_source_match" and source in {"made_in_china", "1688"}:
            bias += 1.0
        source_biases[source] = round(max(-4.0, bias), 2)
        source_scores[source] = qualified * 3.0 + near_miss * 1.5 - watchlist * 0.2
    source_order = sorted(source_scores.keys(), key=lambda key: (-source_scores[key], -source_biases[key], key))
    query_feedback = _query_feedback_from_state(state)
    query_rows = list(query_feedback.get("queries") or [])
    query_biases = {
        str(row.get("query_key", "")): round(float(row.get("query_bias", 0) or 0), 2)
        for row in query_rows
        if str(row.get("query_key", "")).strip()
    }
    query_order = [str(row.get("query_key", "")).strip() for row in query_rows if str(row.get("query_key", "")).strip()]
    query_focus = [str(row.get("query", "")).strip() for row in query_rows[:3] if str(row.get("query", "")).strip()]
    return {
        "status": "adaptive" if decisions else "cold_start",
        "decision_samples": len(decisions),
        "primary_blocker": primary_blocker,
        "failure_categories": failure_categories,
        "keyword_signal_multiplier": round(keyword_multiplier, 2),
        "product_signal_multiplier": round(product_multiplier, 2),
        "platform_biases": platform_biases,
        "source_biases": source_biases,
        "source_order": source_order,
        "query_biases": query_biases,
        "query_order": query_order,
        "query_focus": query_focus,
    }


def _load_adaptive_history() -> dict[str, Any]:
    default = {"entries": []}
    return _read_json(ADAPTIVE_HISTORY_PATH, default)


def _load_query_history() -> dict[str, Any]:
    default = {"entries": []}
    return _read_json(QUERY_HISTORY_PATH, default)


def _query_key(value: str) -> str:
    cleaned = _clean_seed_phrase(value)
    if cleaned:
        return cleaned
    lowered = _normalize_title(value)
    return lowered or _slug(value).replace("-", " ")


def _query_feedback_from_state(state: dict[str, Any]) -> dict[str, Any]:
    query_map: dict[str, dict[str, Any]] = {}
    for payload in (state.get("candidates") or {}).values():
        query = str(payload.get("query", "")).strip()
        if not query:
            continue
        query_key = _query_key(query)
        if not query_key:
            continue
        stats = query_map.setdefault(
            query_key,
            {
                "query": query,
                "query_key": query_key,
                "candidate_count": 0,
                "qualified_count": 0,
                "near_miss_count": 0,
                "watchlist_count": 0,
                "refresh_priority_total": 0.0,
                "sellersprite_keyword_signal_total": 0.0,
                "sellersprite_product_signal_total": 0.0,
                "head_monopoly_total": 0.0,
                "failure_categories": {},
                "platform_counts": {},
            },
        )
        stats["candidate_count"] += 1
        refresh_priority = float(payload.get("demand_refresh_priority", 0) or 0)
        raw_signals = payload.get("raw_signals") or {}
        keyword_signal = float(raw_signals.get("sellersprite_keyword_signal_score", 0) or 0)
        product_signal = float(raw_signals.get("sellersprite_product_signal_score", 0) or 0)
        head_monopoly_score = float(raw_signals.get("head_monopoly_score", 0) or 0)
        stats["refresh_priority_total"] += refresh_priority
        stats["sellersprite_keyword_signal_total"] += keyword_signal
        stats["sellersprite_product_signal_total"] += product_signal
        stats["head_monopoly_total"] += head_monopoly_score
        platform = str(payload.get("sell_platform", "")).lower().strip() or "unknown"
        platform_counts = stats["platform_counts"]
        platform_counts[platform] = int(platform_counts.get(platform, 0) or 0) + 1
        decision = payload.get("decision") or {}
        bucket = str(decision.get("decision_bucket", "")).strip() or "watchlist"
        if bool(decision.get("qualified")) or bucket == "qualified":
            stats["qualified_count"] += 1
        elif bucket == "near_miss":
            stats["near_miss_count"] += 1
        else:
            stats["watchlist_count"] += 1
        for reason in (decision.get("reasons") or []):
            reason = str(reason).strip()
            if not reason:
                continue
            failure_categories = stats["failure_categories"]
            failure_categories[reason] = int(failure_categories.get(reason, 0) or 0) + 1
    query_rows: list[dict[str, Any]] = []
    for stats in query_map.values():
        candidate_count = int(stats.get("candidate_count", 0) or 0)
        if candidate_count <= 0:
            continue
        avg_refresh_priority = float(stats.get("refresh_priority_total", 0) or 0) / candidate_count
        avg_keyword_signal = float(stats.get("sellersprite_keyword_signal_total", 0) or 0) / candidate_count
        avg_product_signal = float(stats.get("sellersprite_product_signal_total", 0) or 0) / candidate_count
        avg_head_monopoly = float(stats.get("head_monopoly_total", 0) or 0) / candidate_count
        qualified_count = int(stats.get("qualified_count", 0) or 0)
        near_miss_count = int(stats.get("near_miss_count", 0) or 0)
        watchlist_count = int(stats.get("watchlist_count", 0) or 0)
        score = (
            qualified_count * 4.0
            + near_miss_count * 2.2
            + watchlist_count * 0.5
            + min(2.0, avg_refresh_priority / 30.0)
            + min(2.0, max(avg_keyword_signal, avg_product_signal) / 28.0)
            - min(2.5, avg_head_monopoly / 35.0)
        )
        query_bias = max(-1.5, min(6.0, score - 1.0))
        top_platform = "unknown"
        if stats["platform_counts"]:
            top_platform = sorted(
                stats["platform_counts"].keys(),
                key=lambda key: (-int(stats["platform_counts"][key] or 0), key),
            )[0]
        query_rows.append(
            {
                "query": stats["query"],
                "query_key": stats["query_key"],
                "score": round(score, 2),
                "query_bias": round(query_bias, 2),
                "candidate_count": candidate_count,
                "qualified_count": qualified_count,
                "near_miss_count": near_miss_count,
                "watchlist_count": watchlist_count,
                "avg_refresh_priority": round(avg_refresh_priority, 2),
                "avg_sellersprite_keyword_signal": round(avg_keyword_signal, 2),
                "avg_sellersprite_product_signal": round(avg_product_signal, 2),
                "avg_head_monopoly_score": round(avg_head_monopoly, 2),
                "top_platform": top_platform,
                "top_failure_category": next(iter((stats.get("failure_categories") or {}).keys()), "none"),
            }
        )
    query_rows.sort(
        key=lambda row: (
            -float(row.get("score", 0) or 0),
            -int(row.get("qualified_count", 0) or 0),
            -int(row.get("near_miss_count", 0) or 0),
            str(row.get("query_key", "")),
        )
    )
    return {
        "status": "adaptive" if query_rows else "cold_start",
        "queries": query_rows[:15],
        "primary_query": str((query_rows[0] if query_rows else {}).get("query", "none")),
    }


def _persist_query_history(query_feedback: dict[str, Any]) -> dict[str, Any]:
    payload = _load_query_history()
    entries = list(payload.get("entries") or [])
    top_queries = list(query_feedback.get("queries") or [])[:5]
    latest_top_score = float((top_queries[0] if top_queries else {}).get("score", 0) or 0)
    entry = {
        "captured_at": _utc_now_iso(),
        "primary_query": query_feedback.get("primary_query", "none"),
        "query_count": len(query_feedback.get("queries") or []),
        "latest_top_score": round(latest_top_score, 2),
        "top_queries": top_queries,
    }
    entries.append(entry)
    entries = entries[-40:]
    payload["entries"] = entries
    _write_json_file(QUERY_HISTORY_PATH, payload)
    latest = entries[-1] if entries else entry
    prev = entries[-2] if len(entries) >= 2 else None
    delta = round(float(latest.get("latest_top_score", 0) or 0) - float((prev or {}).get("latest_top_score", 0) or 0), 2)
    if not prev:
        direction = "new"
    elif delta > 1.0:
        direction = "up"
    elif delta < -1.0:
        direction = "down"
    else:
        direction = "stable"
    return {
        "entries_total": len(entries),
        "latest": latest,
        "trend": {
            "direction": direction,
            "delta": delta,
        },
        "path": str(QUERY_HISTORY_PATH),
    }


def _load_market_history() -> dict[str, Any]:
    default = {"entries": []}
    return _read_json(MARKET_HISTORY_PATH, default)


def _market_feedback_from_state(state: dict[str, Any]) -> dict[str, Any]:
    market_map: dict[str, dict[str, Any]] = {}
    for payload in (state.get("candidates") or {}).values():
        query = str(payload.get("query", "")).strip()
        platform = str(payload.get("sell_platform", "")).lower().strip()
        if not query or not platform:
            continue
        market_key = f"{platform}:{_query_key(query)}"
        row = market_map.setdefault(
            market_key,
            {
                "market_key": market_key,
                "platform": platform,
                "query": query,
                "candidate_count": 0,
                "competition_scores": [],
                "head_monopoly_scores": [],
                "review_counts": [],
                "listing_ages": [],
                "qualified_count": 0,
                "near_miss_count": 0,
            },
        )
        row["candidate_count"] += 1
        signals = payload.get("raw_signals") or {}
        if signals.get("head_monopoly_score") is not None:
            row["head_monopoly_scores"].append(float(signals.get("head_monopoly_score") or 0))
        if signals.get("query_median_reviews") is not None:
            row["review_counts"].append(float(signals.get("query_median_reviews") or 0))
        if payload.get("review_count") is not None:
            row["review_counts"].append(float(payload.get("review_count") or 0))
        if payload.get("listing_age_days") is not None:
            row["listing_ages"].append(float(payload.get("listing_age_days") or 0))
        if payload.get("decision"):
            decision = _decision_payload(payload.get("decision") or {})
            row["competition_scores"].append(float(decision.get("competition_score", 0) or 0))
            if bool(decision.get("qualified")) or str(decision.get("decision_bucket", "")) == "qualified":
                row["qualified_count"] += 1
            elif str(decision.get("decision_bucket", "")) == "near_miss":
                row["near_miss_count"] += 1
    rows: list[dict[str, Any]] = []
    for value in market_map.values():
        competition_avg = round(sum(value["competition_scores"]) / len(value["competition_scores"]), 2) if value["competition_scores"] else 0.0
        head_monopoly_avg = round(sum(value["head_monopoly_scores"]) / len(value["head_monopoly_scores"]), 2) if value["head_monopoly_scores"] else 0.0
        review_median = 0.0
        if value["review_counts"]:
            ordered = sorted(value["review_counts"])
            review_median = round(ordered[len(ordered) // 2], 2)
        listing_age_avg = round(sum(value["listing_ages"]) / len(value["listing_ages"]), 2) if value["listing_ages"] else 0.0
        pressure_score = round(
            min(
                100.0,
                max(0.0, 100.0 - competition_avg) * 0.35
                + min(20.0, head_monopoly_avg / 5.0)
                + min(35.0, review_median / 250.0)
                + min(20.0, listing_age_avg / 90.0)
                - value["qualified_count"] * 4.0
                - value["near_miss_count"] * 2.0,
            ),
            2,
        )
        rows.append(
            {
                "market_key": value["market_key"],
                "platform": value["platform"],
                "query": value["query"],
                "candidate_count": value["candidate_count"],
                "competition_score_avg": competition_avg,
                "head_monopoly_avg": head_monopoly_avg,
                "review_median": review_median,
                "listing_age_avg": listing_age_avg,
                "qualified_count": value["qualified_count"],
                "near_miss_count": value["near_miss_count"],
                "pressure_score": pressure_score,
            }
        )
    rows.sort(key=lambda item: (-float(item.get("pressure_score", 0) or 0), str(item.get("market_key", ""))))
    return {
        "status": "adaptive" if rows else "cold_start",
        "markets": rows[:20],
        "primary_market": str((rows[0] if rows else {}).get("market_key", "none")),
    }


def _persist_market_history(market_feedback: dict[str, Any]) -> dict[str, Any]:
    payload = _load_market_history()
    entries = list(payload.get("entries") or [])
    top_markets = list(market_feedback.get("markets") or [])[:5]
    latest_pressure = float((top_markets[0] if top_markets else {}).get("pressure_score", 0) or 0)
    entry = {
        "captured_at": _utc_now_iso(),
        "primary_market": market_feedback.get("primary_market", "none"),
        "market_count": len(market_feedback.get("markets") or []),
        "latest_pressure_score": round(latest_pressure, 2),
        "top_markets": top_markets,
    }
    entries.append(entry)
    entries = entries[-40:]
    payload["entries"] = entries
    _write_json_file(MARKET_HISTORY_PATH, payload)
    latest = entries[-1] if entries else entry
    prev = entries[-2] if len(entries) >= 2 else None
    delta = round(float(latest.get("latest_pressure_score", 0) or 0) - float((prev or {}).get("latest_pressure_score", 0) or 0), 2)
    if not prev:
        direction = "new"
    elif delta > 1.0:
        direction = "up"
    elif delta < -1.0:
        direction = "down"
    else:
        direction = "stable"
    return {
        "entries_total": len(entries),
        "latest": latest,
        "trend": {
            "direction": direction,
            "delta": delta,
        },
        "path": str(MARKET_HISTORY_PATH),
    }


def _load_threshold_history() -> dict[str, Any]:
    default = {"entries": []}
    return _read_json(THRESHOLD_HISTORY_PATH, default)


def _adaptive_thresholds_from_state(state: dict[str, Any]) -> dict[str, Any]:
    decisions = _state_decisions(state)
    total = len(decisions)
    qualified = sum(1 for item in decisions if item.qualified)
    near_miss = sum(1 for item in decisions if item.decision_bucket == "near_miss")
    failure_categories = _decision_failure_categories(decisions)
    primary_blocker = next(iter(failure_categories.keys()), "none")
    order_floor = 30.0
    listing_age_ceiling = 730
    margin_floor = 0.30
    conservative_margin_floor = 0.30
    launchability_floor = 70.0
    confidence_floor = 80.0
    platform_fit_floor = 65.0
    if total >= 8 and qualified == 0 and near_miss >= 4:
        order_floor = 26.0
        margin_floor = 0.42
        conservative_margin_floor = 0.40
        launchability_floor = 67.0
        confidence_floor = 77.0
        platform_fit_floor = 62.0
    if primary_blocker == "estimated_daily_orders_below_threshold":
        order_floor = min(order_floor, 26.0)
    if primary_blocker == "listing_age_unknown":
        listing_age_ceiling = 900
    if primary_blocker == "missing_trusted_weight":
        confidence_floor = max(confidence_floor, 82.0)
        platform_fit_floor = max(platform_fit_floor, 66.0)
    if qualified >= 3:
        order_floor = max(order_floor, 30.0)
        margin_floor = max(margin_floor, 0.30)
        conservative_margin_floor = max(conservative_margin_floor, 0.30)
    return {
        "status": "adaptive" if decisions else "cold_start",
        "decision_samples": total,
        "qualified_total": qualified,
        "near_miss_total": near_miss,
        "primary_blocker": primary_blocker,
        "order_floor": round(order_floor, 2),
        "listing_age_ceiling": int(listing_age_ceiling),
        "margin_floor": round(margin_floor, 4),
        "conservative_margin_floor": round(conservative_margin_floor, 4),
        "launchability_floor": round(launchability_floor, 2),
        "confidence_floor": round(confidence_floor, 2),
        "platform_fit_floor": round(platform_fit_floor, 2),
    }


def _persist_threshold_history(thresholds: dict[str, Any]) -> dict[str, Any]:
    payload = _load_threshold_history()
    entries = list(payload.get("entries") or [])
    entry = {
        "captured_at": _utc_now_iso(),
        "primary_blocker": thresholds.get("primary_blocker", "none"),
        "decision_samples": thresholds.get("decision_samples", 0),
        "qualified_total": thresholds.get("qualified_total", 0),
        "near_miss_total": thresholds.get("near_miss_total", 0),
        "order_floor": thresholds.get("order_floor", 30.0),
        "listing_age_ceiling": thresholds.get("listing_age_ceiling", 730),
        "margin_floor": thresholds.get("margin_floor", 0.30),
        "conservative_margin_floor": thresholds.get("conservative_margin_floor", 0.30),
        "launchability_floor": thresholds.get("launchability_floor", 70.0),
        "confidence_floor": thresholds.get("confidence_floor", 80.0),
        "platform_fit_floor": thresholds.get("platform_fit_floor", 65.0),
    }
    entries.append(entry)
    entries = entries[-40:]
    payload["entries"] = entries
    _write_json_file(THRESHOLD_HISTORY_PATH, payload)
    latest = entries[-1] if entries else entry
    prev = entries[-2] if len(entries) >= 2 else None
    delta = round(float(latest.get("order_floor", 30.0) or 30.0) - float((prev or {}).get("order_floor", 30.0) or 30.0), 2)
    if not prev:
        direction = "new"
    elif delta > 0.5:
        direction = "up"
    elif delta < -0.5:
        direction = "down"
    else:
        direction = "stable"
    return {
        "entries_total": len(entries),
        "latest": latest,
        "trend": {
            "direction": direction,
            "delta": delta,
        },
        "path": str(THRESHOLD_HISTORY_PATH),
    }


def _apply_query_adaptive_order(queries: list[str], adaptive_profile: dict[str, Any], *, limit: int | None = None) -> list[str]:
    query_biases = adaptive_profile.get("query_biases") or {}
    query_order = list(adaptive_profile.get("query_order") or [])
    query_order_index = {str(key): index for index, key in enumerate(query_order)}
    rows: list[tuple[str, str, float]] = []
    seen: set[str] = set()
    for query in queries:
        query_key = _query_key(query)
        if not query_key or query_key in seen:
            continue
        seen.add(query_key)
        rows.append((query, query_key, float(query_biases.get(query_key, 0) or 0)))
    rows.sort(key=lambda row: (-row[2], query_order_index.get(row[1], 999), row[0]))
    ordered = [row[0] for row in rows]
    if limit is not None:
        ordered = ordered[:limit]
    return ordered


def _adaptive_strength_score(profile: dict[str, Any]) -> float:
    keyword_multiplier = float(profile.get("keyword_signal_multiplier", 1.0) or 1.0)
    product_multiplier = float(profile.get("product_signal_multiplier", 1.0) or 1.0)
    decision_samples = float(profile.get("decision_samples", 0) or 0)
    biases = profile.get("platform_biases") or {}
    bias_strength = sum(abs(float(value or 0)) for value in biases.values())
    source_biases = profile.get("source_biases") or {}
    source_bias_strength = sum(abs(float(value or 0)) for value in source_biases.values())
    query_biases = profile.get("query_biases") or {}
    query_bias_strength = sum(abs(float(value or 0)) for value in query_biases.values())
    return round(
        min(
            100.0,
            40.0
            + decision_samples * 1.5
            + (keyword_multiplier - 1.0) * 45.0
            + (product_multiplier - 1.0) * 40.0
            + min(12.0, bias_strength)
            + min(10.0, source_bias_strength),
            + min(10.0, query_bias_strength),
        ),
        2,
    )


def _persist_adaptive_history(profile: dict[str, Any]) -> dict[str, Any]:
    payload = _load_adaptive_history()
    entries = list(payload.get("entries") or [])
    entry = {
        "captured_at": _utc_now_iso(),
        "primary_blocker": profile.get("primary_blocker", "none"),
        "decision_samples": int(profile.get("decision_samples", 0) or 0),
        "keyword_signal_multiplier": float(profile.get("keyword_signal_multiplier", 1.0) or 1.0),
        "product_signal_multiplier": float(profile.get("product_signal_multiplier", 1.0) or 1.0),
        "platform_biases": profile.get("platform_biases", {}) or {},
        "source_biases": profile.get("source_biases", {}) or {},
        "source_order": profile.get("source_order", []) or [],
        "query_biases": profile.get("query_biases", {}) or {},
        "query_order": profile.get("query_order", []) or [],
        "query_focus": profile.get("query_focus", []) or [],
        "strength_score": _adaptive_strength_score(profile),
    }
    entries.append(entry)
    entries = entries[-40:]
    payload["entries"] = entries
    _write_json_file(ADAPTIVE_HISTORY_PATH, payload)
    latest = entries[-1] if entries else entry
    prev = entries[-2] if len(entries) >= 2 else None
    delta = round(float(latest.get("strength_score", 0) or 0) - float((prev or {}).get("strength_score", 0) or 0), 2)
    if not prev:
        direction = "new"
    elif delta > 1.5:
        direction = "up"
    elif delta < -1.5:
        direction = "down"
    else:
        direction = "stable"
    return {
        "entries_total": len(entries),
        "latest": latest,
        "trend": {
            "direction": direction,
            "delta": delta,
        },
        "path": str(ADAPTIVE_HISTORY_PATH),
    }


def _collect_sellersprite_summary(queries: list[str], *, fast: bool = False) -> dict[str, Any]:
    pages: dict[str, Any] = {}
    page_map = {
        "product_research": SELLERSPRITE_PRODUCT_RESEARCH_URL,
        "keyword_research": SELLERSPRITE_KEYWORD_RESEARCH_URL,
    }
    if fast:
        page_map = {"keyword_research": SELLERSPRITE_KEYWORD_RESEARCH_URL}
    for label, url in page_map.items():
        pages[label] = _sellersprite_fetch(
            save_prefix=f"sellersprite-{label}",
            url=url,
            wait_seconds=1.5 if fast else 6.0,
        )
    available_platforms = sorted(
        {
            str(platform).strip().lower()
            for payload in pages.values()
            for platform in (((payload.get("parsed") or {}).get("platforms") or []))
            if str(platform).strip()
        }
    )
    latest_month = ""
    for payload in pages.values():
        months = ((payload.get("parsed") or {}).get("months") or [])
        if months:
            latest_month = str(months[0]).strip()
            break
    api_probes: list[dict[str, Any]] = []
    probe_month = latest_month or (_ny_now().strftime("%Y-%m"))
    if not fast:
        for query in queries[:2]:
            api_path = f"/v2/keyword-stat/gkdata?station=US&keyword={quote_plus(query)}&month={probe_month}"
            probe = _sellersprite_fetch(
                save_prefix=f"sellersprite-keyword-probe-{_slug(query)}",
                api_path=api_path,
                base_url=SELLERSPRITE_KEYWORD_RESEARCH_URL,
                wait_seconds=3.0,
            )
            parsed = probe.get("parsed") or {}
            state, detail = _sellersprite_probe_status(parsed if isinstance(parsed, dict) else {})
            body = parsed.get("body") if isinstance(parsed, dict) else {}
            api_probes.append(
                {
                    "keyword": query,
                    "month": probe_month,
                    "status": state,
                    "detail": detail,
                    "api_status": parsed.get("api_status"),
                    "api_ok": parsed.get("api_ok"),
                    "api_path": api_path,
                    "message": body.get("message") if isinstance(body, dict) else "",
                    "code": body.get("code") if isinstance(body, dict) else "",
                    "json_path": probe.get("json_path", ""),
                }
            )
    page_ok_total = sum(1 for payload in pages.values() if payload.get("ok"))
    api_ok_total = sum(1 for row in api_probes if row.get("status") == "ok")
    status = "unavailable"
    if page_ok_total and api_ok_total:
        status = "connected"
    elif page_ok_total:
        status = "partial"
    submit_keyword = _slug(queries[0]).replace("-", " ") if queries else "organizer"
    if len(submit_keyword.strip()) < 3:
        submit_keyword = "organizer"
    submit_probe: dict[str, Any] = {}
    product_probe: dict[str, Any] = {}
    if not fast:
        submit_probe = _parse_json_stdout(
            _run(
                [
                    str(VENV_PY),
                    str(SELLERSPRITE_KEYWORD_SUBMIT_PROBE),
                    "--keyword",
                    submit_keyword,
                    "--save-prefix",
                    f"sellersprite-submit-{_slug(submit_keyword)}",
                    "--wait-seconds",
                    "4",
                ],
                timeout=120,
            ).stdout
        )
        product_probe = _parse_json_stdout(
            _run(
                [
                    str(VENV_PY),
                    str(SELLERSPRITE_PRODUCT_SUBMIT_PROBE),
                    "--save-prefix",
                    "sellersprite-product-default",
                    "--wait-seconds",
                    "6",
                ],
                timeout=120,
            ).stdout
        )
    keyword_metrics = (((pages.get("keyword_research") or {}).get("parsed") or {}).get("filter_metrics") or [])
    product_metrics = (((pages.get("product_research") or {}).get("parsed") or {}).get("filter_metrics") or [])
    summary = {
        "status": status,
        "latest_month": probe_month,
        "available_platforms": available_platforms,
        "page_ok_total": page_ok_total,
        "api_ok_total": api_ok_total,
        "mode": "fast" if fast else "full",
        "pages": pages,
        "api_probes": api_probes,
        "keyword_metrics_available": keyword_metrics,
        "product_metrics_available": product_metrics,
        "keyword_result_probe": {
            "keyword": submit_probe.get("keyword", submit_keyword),
            "title": submit_probe.get("title", ""),
            "final_url": submit_probe.get("final_url", ""),
            "result_count": ((submit_probe.get("parsed") or {}).get("result_count", 0)),
            "top_keywords": ((submit_probe.get("parsed") or {}).get("top_keywords", [])[:5]),
            "json_path": submit_probe.get("json_path", ""),
        },
        "product_result_probe": {
            "title": product_probe.get("title", ""),
            "final_url": product_probe.get("final_url", ""),
            "product_rows_detected": ((product_probe.get("parsed") or {}).get("product_rows_detected", 0)),
            "top_products": ((product_probe.get("parsed") or {}).get("top_products", [])[:5]),
            "json_path": product_probe.get("json_path", ""),
        },
    }
    summary["seed_queries"] = _sellersprite_seed_queries(summary)
    summary["product_watchlist"] = _sellersprite_product_watchlist(summary)
    return summary


def _write_excel(run_id: str, decisions: list[ArbitrageDecision], summary: dict[str, Any]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Qualified_All"
    ws.append(["产品名称", "目标采购平台", "采购链接", "目标售卖平台", "售卖链接", "平台适配分", "平台建议"])
    style_header(ws)
    for item in decisions:
        if not item.qualified:
            continue
        ws.append([item.product_name, item.buy_platform, item.buy_link, item.sell_platform, item.sell_link, item.platform_fit_score, item.platform_recommendation])
    autosize(ws)

    near_ws = wb.create_sheet("Near_Miss")
    near_ws.append(["产品名称", "售卖平台", "来源类型", "平台适配分", "精选分", "精选标签", "趋势源分", "执行分", "货源可信分", "货源相似分", "货源相似标签", "头部垄断分", "可做分", "置信度", "优先原因", "原因"])
    style_header(near_ws)
    for item in decisions:
        if item.decision_bucket != "near_miss":
            continue
        near_ws.append([
            item.product_name,
            item.sell_platform,
            item.query_source_kind,
            item.platform_fit_score,
            item.selection_score,
            item.selection_grade,
            item.trend_source_score,
            item.execution_resilience_score,
            item.supplier_confidence_score,
            item.supplier_similarity_score,
            item.supplier_similarity_grade,
            item.head_monopoly_score,
            item.launchability_score,
            item.confidence_score,
            item.priority_reason,
            " | ".join(item.reasons),
        ])
    autosize(near_ws)

    watch_ws = wb.create_sheet("Watchlist")
    watch_ws.append(["产品名称", "售卖平台", "来源类型", "平台适配分", "精选分", "精选标签", "趋势源分", "执行分", "货源可信分", "货源相似分", "货源相似标签", "头部垄断分", "可做分", "置信度", "优先原因", "原因"])
    style_header(watch_ws)
    for item in decisions:
        if item.decision_bucket != "watchlist":
            continue
        watch_ws.append([
            item.product_name,
            item.sell_platform,
            item.query_source_kind,
            item.platform_fit_score,
            item.selection_score,
            item.selection_grade,
            item.trend_source_score,
            item.execution_resilience_score,
            item.supplier_confidence_score,
            item.supplier_similarity_score,
            item.supplier_similarity_grade,
            item.head_monopoly_score,
            item.launchability_score,
            item.confidence_score,
            item.priority_reason,
            " | ".join(item.reasons),
        ])
    autosize(watch_ws)

    for platform in ["amazon", "walmart", "temu", "tiktok"]:
        platform_ws = wb.create_sheet(_sheet_name_for_platform(platform))
        platform_ws.append([
            "产品名称", "目标采购平台", "采购链接", "目标售卖平台", "售卖链接", "售价(RMB)", "采购成本(RMB)", "毛利率",
            "保守毛利率", "估算日单量", "上架天数", "精选分", "精选标签", "精选论点", "趋势源分", "执行分", "货源可信分", "货源相似分", "货源相似标签", "头部垄断分", "可做分", "平台适配分", "平台适配标签", "平台建议", "结果分层", "优先原因", "是否入选", "原因"
        ])
        style_header(platform_ws)
        platform_rows = _platform_sheet_rows(decisions, platform)
        if platform == "tiktok" and not platform_rows:
            platform_ws.append(["", "", "", "tiktok", "", "", "", "", "", "", "", "", "", "", "watchlist_only", "no_current_tiktok_candidates", "trend intake pending"])
        for item in platform_rows:
            platform_ws.append([
                item.product_name,
                item.buy_platform,
                item.buy_link,
                item.sell_platform,
                item.sell_link,
                item.sell_price_cny,
                item.purchase_cost_cny,
                item.gross_margin_rate,
                item.conservative_margin_rate,
                item.estimated_daily_orders,
                item.listing_age_days,
                item.selection_score,
                item.selection_grade,
                item.selection_thesis,
                item.trend_source_score,
                item.execution_resilience_score,
                item.supplier_confidence_score,
                item.supplier_similarity_score,
                item.supplier_similarity_grade,
                item.head_monopoly_score,
                item.launchability_score,
                item.platform_fit_score,
                item.platform_fit_label,
                item.platform_recommendation,
                item.decision_bucket,
                item.priority_reason,
                "yes" if item.qualified else "no",
                " | ".join(item.reasons),
            ])
        autosize(platform_ws)

    audit = wb.create_sheet("Audit")
    headers = [
        "产品名称", "采购平台", "采购链接", "售卖平台", "售卖链接", "售价(RMB)", "采购成本(RMB)", "重量(kg)",
        "毛利额", "毛利率", "保守毛利率", "平台佣金率", "估算日单量", "上架天数", "需求分", "竞争分",
        "差异化分", "价格稳定分", "精选分", "精选标签", "精选论点", "趋势源分", "执行分", "货源可信分", "货源相似分", "货源相似标签", "头部垄断分", "平台适配分", "平台适配标签", "平台建议", "结果分层", "优先原因", "综合可做分", "置信度", "重量等级", "是否入选", "原因"
    ]
    audit.append(headers)
    style_header(audit)
    for item in decisions:
        audit.append([
            item.product_name,
            item.buy_platform,
            item.buy_link,
            item.sell_platform,
            item.sell_link,
            item.sell_price_cny,
            item.purchase_cost_cny,
            item.weight_kg,
            item.gross_profit_amount,
            item.gross_margin_rate,
            item.conservative_margin_rate,
            item.platform_fee_rate,
            item.estimated_daily_orders,
            item.listing_age_days,
            item.demand_score,
            item.competition_score,
            item.differentiation_score,
            item.price_stability_score,
            item.selection_score,
            item.selection_grade,
            item.selection_thesis,
            item.trend_source_score,
            item.execution_resilience_score,
            item.supplier_confidence_score,
            item.supplier_similarity_score,
            item.supplier_similarity_grade,
            item.head_monopoly_score,
            item.platform_fit_score,
            item.platform_fit_label,
            item.platform_recommendation,
            item.decision_bucket,
            item.priority_reason,
            item.launchability_score,
            item.confidence_score,
            item.weight_grade,
            "yes" if item.qualified else "no",
            " | ".join(item.reasons),
        ])
    autosize(audit)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["字段", "内容"])
    style_header(summary_ws)
    for key, value in summary.items():
        summary_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    autosize(summary_ws)

    query_feedback = summary.get("query_feedback") or {}
    if query_feedback:
        q_ws = wb.create_sheet("Query_Insights")
        q_ws.append([
            "Query",
            "Score",
            "Bias",
            "Candidates",
            "Qualified",
            "Near Miss",
            "Watchlist",
            "Avg Refresh Priority",
            "Avg SS Keyword Signal",
            "Avg SS Product Signal",
            "Avg Head Monopoly",
            "Top Platform",
            "Top Failure Category",
        ])
        style_header(q_ws)
        for row in (query_feedback.get("queries") or [])[:20]:
            q_ws.append([
                row.get("query", ""),
                row.get("score", 0),
                row.get("query_bias", 0),
                row.get("candidate_count", 0),
                row.get("qualified_count", 0),
                row.get("near_miss_count", 0),
                row.get("watchlist_count", 0),
                row.get("avg_refresh_priority", 0),
                row.get("avg_sellersprite_keyword_signal", 0),
                row.get("avg_sellersprite_product_signal", 0),
                row.get("avg_head_monopoly_score", 0),
                row.get("top_platform", ""),
                row.get("top_failure_category", ""),
            ])
        autosize(q_ws)

    market_feedback = summary.get("market_feedback") or {}
    if market_feedback:
        m_ws = wb.create_sheet("Market_Pressure")
        m_ws.append([
            "Market Key",
            "Platform",
            "Query",
            "Pressure Score",
            "Competition Avg",
            "Head Monopoly Avg",
            "Review Median",
            "Listing Age Avg",
            "Candidates",
            "Qualified",
            "Near Miss",
        ])
        style_header(m_ws)
        for row in (market_feedback.get("markets") or [])[:25]:
            m_ws.append([
                row.get("market_key", ""),
                row.get("platform", ""),
                row.get("query", ""),
                row.get("pressure_score", 0),
                row.get("competition_score_avg", 0),
                row.get("head_monopoly_avg", 0),
                row.get("review_median", 0),
                row.get("listing_age_avg", 0),
                row.get("candidate_count", 0),
                row.get("qualified_count", 0),
                row.get("near_miss_count", 0),
            ])
        autosize(m_ws)

    sellersprite = summary.get("sellersprite") or {}
    if sellersprite:
        ss_ws = wb.create_sheet("SellerSprite")
        ss_ws.append(["类型", "名称", "状态", "细节", "附加信息"])
        style_header(ss_ws)
        for query in (sellersprite.get("seed_queries") or []):
            ss_ws.append(["seed_query", query, "ok", "", ""])
        for page_name, payload in (sellersprite.get("pages") or {}).items():
            parsed = payload.get("parsed") or {}
            detail = json.dumps(
                {
                    "page_kind": parsed.get("page_kind", ""),
                    "months": parsed.get("months", []),
                    "platforms": parsed.get("platforms", []),
                    "metrics": parsed.get("filter_metrics", []),
                },
                ensure_ascii=False,
            )
            ss_ws.append(["page", page_name, "ok" if payload.get("ok") else "error", detail, payload.get("url", "")])
        for probe in (sellersprite.get("api_probes") or []):
            detail = json.dumps(
                {
                    "month": probe.get("month", ""),
                    "api_status": probe.get("api_status", ""),
                    "message": probe.get("message", ""),
                    "code": probe.get("code", ""),
                },
                ensure_ascii=False,
            )
            ss_ws.append(["api_probe", probe.get("keyword", ""), probe.get("status", ""), detail, probe.get("api_path", "")])
        keyword_result_probe = sellersprite.get("keyword_result_probe") or {}
        if keyword_result_probe:
            ss_ws.append(
                [
                    "result_probe",
                    keyword_result_probe.get("keyword", ""),
                    "ok" if keyword_result_probe.get("result_count", 0) else "watch",
                    json.dumps(
                        {
                            "result_count": keyword_result_probe.get("result_count", 0),
                            "top_keywords": keyword_result_probe.get("top_keywords", []),
                        },
                        ensure_ascii=False,
                    ),
                    keyword_result_probe.get("final_url", ""),
                ]
            )
        product_result_probe = sellersprite.get("product_result_probe") or {}
        if product_result_probe:
            ss_ws.append(
                [
                    "product_probe",
                    "default_product_research",
                    "ok" if product_result_probe.get("product_rows_detected", 0) else "watch",
                    json.dumps(
                        {
                            "product_rows_detected": product_result_probe.get("product_rows_detected", 0),
                            "top_products": product_result_probe.get("top_products", []),
                        },
                        ensure_ascii=False,
                    ),
                    product_result_probe.get("final_url", ""),
                ]
            )
        for row in (sellersprite.get("product_watchlist") or []):
            ss_ws.append(
                [
                    "product_watch",
                    row.get("product_name", ""),
                    "watch",
                    json.dumps(
                        {
                            "rank": row.get("rank"),
                            "asin": row.get("asin", ""),
                            "monthly_sales_hint": row.get("monthly_sales_hint", ""),
                            "sales_amount_hint": row.get("sales_amount_hint", ""),
                            "price_hint": row.get("price_hint", ""),
                            "review_count_hint": row.get("review_count_hint", ""),
                            "rating_hint": row.get("rating_hint", ""),
                            "listing_date_hint": row.get("listing_date_hint", ""),
                        },
                        ensure_ascii=False,
                    ),
                    "",
                ]
            )
        autosize(ss_ws)

    path = OUT_DIR / f"cross-market-arbitrage-{run_id}.xlsx"
    wb.save(path)
    return path


def _write_markdown(run_id: str, decisions: list[ArbitrageDecision], summary: dict[str, Any]) -> Path:
    governance = summary.get("governance") or {}
    platform_summary = summary.get("platform_summary") or {}
    sellersprite = summary.get("sellersprite") or {}
    adaptive_profile = summary.get("adaptive_profile") or {}
    adaptive_history = summary.get("adaptive_history") or {}
    adaptive_thresholds = summary.get("adaptive_thresholds") or {}
    threshold_history = summary.get("threshold_history") or {}
    query_feedback = summary.get("query_feedback") or {}
    query_history = summary.get("query_history") or {}
    market_feedback = summary.get("market_feedback") or {}
    market_history = summary.get("market_history") or {}
    market_feedback = summary.get("market_feedback") or {}
    market_history = summary.get("market_history") or {}
    lines = [
        "# Cross-market arbitrage run",
        "",
        f"- Run ID: `{run_id}`",
        f"- Generated at: `{summary['generated_at']}`",
        f"- Qualified count: `{summary['qualified_count']}`",
        f"- Discovery candidates: `{summary['discovery_candidate_count']}`",
        f"- Governance status: `{governance.get('status', 'unknown')}`",
        "",
        "## Platform Summary",
        "",
    ]
    for platform, payload in platform_summary.items():
        lines.append(
            f"- `{platform}`: candidates=`{payload.get('candidate_count', 0)}`, qualified=`{payload.get('qualified_count', 0)}`, near_miss=`{payload.get('near_miss_count', 0)}`, strong_fit=`{payload.get('strong_fit_count', 0)}`, watch=`{payload.get('watch_count', 0)}`"
        )
    lines.extend([
        "",
        "## Governance",
        "",
        f"- Primary blocker: `{governance.get('primary_blocker', 'none')}`",
        f"- Next actions: `{', '.join(governance.get('next_actions', []) or ['none'])}`",
        f"- Failure categories: `{json.dumps(governance.get('failure_categories', {}), ensure_ascii=False)}`",
        "",
        "## Adaptive Profile",
        "",
        f"- Status: `{adaptive_profile.get('status', 'cold_start')}`",
        f"- Decision samples: `{adaptive_profile.get('decision_samples', 0)}`",
        f"- Primary blocker: `{adaptive_profile.get('primary_blocker', 'none')}`",
        f"- Keyword multiplier: `{adaptive_profile.get('keyword_signal_multiplier', 1.0)}`",
        f"- Product multiplier: `{adaptive_profile.get('product_signal_multiplier', 1.0)}`",
        f"- Platform biases: `{json.dumps(adaptive_profile.get('platform_biases', {}), ensure_ascii=False)}`",
        f"- Source biases: `{json.dumps(adaptive_profile.get('source_biases', {}), ensure_ascii=False)}`",
        f"- Source order: `{', '.join(adaptive_profile.get('source_order', []) or ['none'])}`",
        f"- Query focus: `{', '.join(adaptive_profile.get('query_focus', []) or ['none'])}`",
        f"- Query biases: `{json.dumps(adaptive_profile.get('query_biases', {}), ensure_ascii=False)}`",
        f"- History trend: `{((adaptive_history.get('trend') or {}).get('direction', 'stable'))}` delta=`{((adaptive_history.get('trend') or {}).get('delta', 0))}` entries=`{adaptive_history.get('entries_total', 0)}`",
        "",
        "## Adaptive Thresholds",
        "",
        f"- Status: `{adaptive_thresholds.get('status', 'cold_start')}`",
        f"- Primary blocker: `{adaptive_thresholds.get('primary_blocker', 'none')}`",
        f"- Order floor: `{adaptive_thresholds.get('order_floor', 30.0)}`",
        f"- Listing age ceiling: `{adaptive_thresholds.get('listing_age_ceiling', 730)}`",
        f"- Margin floor: `{adaptive_thresholds.get('margin_floor', 0.30)}` / conservative=`{adaptive_thresholds.get('conservative_margin_floor', 0.30)}`",
        f"- Launchability / confidence / platform fit: `{adaptive_thresholds.get('launchability_floor', 70.0)}` / `{adaptive_thresholds.get('confidence_floor', 80.0)}` / `{adaptive_thresholds.get('platform_fit_floor', 65.0)}`",
        f"- History trend: `{((threshold_history.get('trend') or {}).get('direction', 'stable'))}` delta=`{((threshold_history.get('trend') or {}).get('delta', 0))}` entries=`{threshold_history.get('entries_total', 0)}`",
        "",
        "## Query Feedback",
        "",
        f"- Status: `{query_feedback.get('status', 'cold_start')}`",
        f"- Primary query: `{query_feedback.get('primary_query', 'none')}`",
        f"- History trend: `{((query_history.get('trend') or {}).get('direction', 'stable'))}` delta=`{((query_history.get('trend') or {}).get('delta', 0))}` entries=`{query_history.get('entries_total', 0)}`",
        "",
        "## Market Pressure",
        "",
        f"- Status: `{market_feedback.get('status', 'cold_start')}`",
        f"- Primary market: `{market_feedback.get('primary_market', 'none')}`",
        f"- History trend: `{((market_history.get('trend') or {}).get('direction', 'stable'))}` delta=`{((market_history.get('trend') or {}).get('delta', 0))}` entries=`{market_history.get('entries_total', 0)}`",
        "",
        "## SellerSprite",
        "",
        f"- Status: `{sellersprite.get('status', 'unavailable')}`",
        f"- Latest month: `{sellersprite.get('latest_month', '')}`",
        f"- Available platforms: `{', '.join(sellersprite.get('available_platforms', []) or ['none'])}`",
        f"- Seed queries: `{', '.join((sellersprite.get('seed_queries') or [])[:6]) or 'none'}`",
        f"- Keyword metrics: `{', '.join((sellersprite.get('keyword_metrics_available', []) or [])[:8]) or 'none'}`",
        f"- Product metrics: `{', '.join((sellersprite.get('product_metrics_available', []) or [])[:8]) or 'none'}`",
        f"- Result probe: keyword=`{((sellersprite.get('keyword_result_probe') or {}).get('keyword', ''))}` count=`{((sellersprite.get('keyword_result_probe') or {}).get('result_count', 0))}`",
        f"- Product probe: rows=`{((sellersprite.get('product_result_probe') or {}).get('product_rows_detected', 0))}`",
        "",
        "### API Probes",
        "",
    ])
    for probe in (sellersprite.get("api_probes") or []):
        lines.append(
            f"- `{probe.get('keyword', '')}`: status=`{probe.get('status', '')}` detail=`{probe.get('detail', '')}` api_status=`{probe.get('api_status', '')}` code=`{probe.get('code', '')}`"
        )
    lines.extend([
        "",
        "### Result Probe",
        "",
    ])
    for row in (((sellersprite.get("keyword_result_probe") or {}).get("top_keywords") or [])[:5]):
        lines.append(
            f"- `#{row.get('rank', '')}` `{row.get('keyword', '')}` / `{row.get('keyword_cn', '')}` / search=`{row.get('monthly_searches', '')}`"
        )
    lines.extend([
        "",
        "### Product Probe",
        "",
    ])
    for row in (((sellersprite.get("product_result_probe") or {}).get("top_products") or [])[:5]):
        lines.append(
            f"- `#{row.get('rank', '')}` `{row.get('product_name', '')}` / price_hint=`{row.get('price_hint', '')}`"
        )
    if sellersprite.get("product_watchlist"):
        lines.extend([
            "",
            "### Product Watchlist",
            "",
        ])
        for row in (sellersprite.get("product_watchlist") or [])[:5]:
            lines.append(
                f"- `#{row.get('rank', '')}` `{row.get('product_name', '')}` / asin=`{row.get('asin', '')}` / sales=`{row.get('monthly_sales_hint', '')}` / revenue=`{row.get('sales_amount_hint', '')}` / price=`{row.get('price_hint', '')}` / rating=`{row.get('rating_hint', '')}` / reviews=`{row.get('review_count_hint', '')}` / listed=`{row.get('listing_date_hint', '')}`"
            )
    if query_feedback:
        lines.extend([
            "",
            "### Query Insights",
            "",
        ])
        for row in (query_feedback.get("queries") or [])[:8]:
            lines.append(
                f"- `{row.get('query', '')}` / score=`{row.get('score', 0)}` / bias=`{row.get('query_bias', 0)}` / qualified=`{row.get('qualified_count', 0)}` / near_miss=`{row.get('near_miss_count', 0)}` / monopoly_avg=`{row.get('avg_head_monopoly_score', 0)}` / top_platform=`{row.get('top_platform', '')}`"
            )
    if market_feedback:
        lines.extend([
            "",
            "### Market Pressure",
            "",
        ])
        for row in (market_feedback.get("markets") or [])[:8]:
            lines.append(
                f"- `{row.get('market_key', '')}` / pressure=`{row.get('pressure_score', 0)}` / competition_avg=`{row.get('competition_score_avg', 0)}` / monopoly_avg=`{row.get('head_monopoly_avg', 0)}` / review_median=`{row.get('review_median', 0)}` / listing_age_avg=`{row.get('listing_age_avg', 0)}`"
            )
    lines.extend([
        "",
        "## Qualified",
        "",
    ])
    qualified = [item for item in decisions if item.qualified]
    near_miss = [item for item in decisions if item.decision_bucket == "near_miss"]
    watchlist = [item for item in decisions if item.decision_bucket == "watchlist"]
    if not qualified:
        lines.append("- No candidates passed the current strong thresholds.")
    for item in qualified:
        lines.extend([
            f"### {item.product_name}",
            f"- Buy: `{item.buy_platform}` -> {item.buy_link}",
            f"- Sell: `{item.sell_platform}` -> {item.sell_link}",
            f"- Estimated daily orders: `{item.estimated_daily_orders}`",
            f"- Listing age days: `{item.listing_age_days}`",
            f"- Margin: `{item.gross_margin_rate}`",
            f"- Conservative margin: `{item.conservative_margin_rate}`",
            f"- Selection score: `{item.selection_score}` / `{item.selection_grade}` / `{item.selection_thesis}`",
            f"- Trend source / execution / supplier / supplier similarity / monopoly: `{item.trend_source_score}` / `{item.execution_resilience_score}` / `{item.supplier_confidence_score}` / `{item.supplier_similarity_score}` (`{item.supplier_similarity_grade}`) / `{item.head_monopoly_score}`",
            f"- Platform fit: `{item.platform_fit_score}` / `{item.platform_fit_label}` / `{item.platform_recommendation}`",
            f"- Launchability: `{item.launchability_score}`",
            f"- Confidence: `{item.confidence_score}`",
            "",
        ])
    lines.extend(["## Near Miss", ""])
    if not near_miss:
        lines.append("- No near-miss candidates this run.")
    for item in near_miss[:10]:
        lines.extend([
            f"### {item.product_name}",
            f"- Platform: `{item.sell_platform}`",
            f"- Priority reason: `{item.priority_reason}`",
            f"- Selection score: `{item.selection_score}` / `{item.selection_grade}` / `{item.selection_thesis}`",
            f"- Trend source / execution / supplier / supplier similarity / monopoly: `{item.trend_source_score}` / `{item.execution_resilience_score}` / `{item.supplier_confidence_score}` / `{item.supplier_similarity_score}` (`{item.supplier_similarity_grade}`) / `{item.head_monopoly_score}`",
            f"- Platform fit: `{item.platform_fit_score}` / `{item.platform_fit_label}` / `{item.platform_recommendation}`",
            f"- Launchability: `{item.launchability_score}` / confidence=`{item.confidence_score}`",
            f"- Reasons: `{', '.join(item.reasons or ['none'])}`",
            "",
        ])
    lines.extend(["## Watchlist", ""])
    if not watchlist:
        lines.append("- No watchlist candidates this run.")
    for item in watchlist[:10]:
        lines.extend([
            f"- `{item.product_name}` / `{item.sell_platform}` / reason=`{item.priority_reason}` / selection=`{item.selection_score}` / fit=`{item.platform_fit_score}` / launchability=`{item.launchability_score}`",
        ])
    lines.extend(["## Audit notes", ""])
    lines.extend([
        "- 1688 is still treated as high-risk / often blocked under automated browsing.",
        "- Yiwugo is currently supplementary, not a primary trusted weight source.",
        "- Made-in-China is currently the most stable public source among the tested source-side platforms.",
    ])
    path = OUT_DIR / f"cross-market-arbitrage-{run_id}.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def _write_json(run_id: str, payload: dict[str, Any]) -> Path:
    path = OUT_DIR / f"cross-market-arbitrage-{run_id}.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _state_candidate_window(state: dict[str, Any], *, hours: int = REPORT_WINDOW_HOURS) -> list[dict[str, Any]]:
    cutoff = _utc_now() - timedelta(hours=hours)
    rows: list[dict[str, Any]] = []
    for payload in (state.get("candidates") or {}).values():
        discovered_at = payload.get("discovered_at") or payload.get("updated_at") or ""
        dt = parse_dt(discovered_at)
        if dt and dt >= cutoff:
            rows.append(payload)
    return rows


def parse_dt(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(timezone.utc)
    except Exception:
        return None


def _needs_rematch(row: dict[str, Any]) -> bool:
    last_matched = parse_dt(str(row.get("last_matched_at", "")))
    if last_matched is None:
        return True
    updated_at = parse_dt(str(row.get("updated_at", "")))
    if updated_at and updated_at > last_matched:
        return True
    return not bool(row.get("decision"))


def _send_to_telegram(*, chat_id: str, text: str, media_paths: list[Path]) -> list[dict[str, Any]]:
    deliveries: list[dict[str, Any]] = []
    text_cmd = [OPENCLAW_BIN, "message", "send", "--channel", "telegram", "--target", chat_id, "--message", text, "--json"]
    text_proc = _run(text_cmd, timeout=60)
    deliveries.append({"kind": "text", "returncode": text_proc.returncode, "stdout": text_proc.stdout.strip(), "stderr": text_proc.stderr.strip()})
    for media_path in media_paths:
        cmd = [OPENCLAW_BIN, "message", "send", "--channel", "telegram", "--target", chat_id, "--media", str(media_path), "--force-document", "--json"]
        proc = _run(cmd, timeout=120)
        deliveries.append({"kind": "media", "path": str(media_path), "returncode": proc.returncode, "stdout": proc.stdout.strip(), "stderr": proc.stderr.strip()})
    return deliveries


def _summary_text(summary: dict[str, Any], decisions: list[ArbitrageDecision]) -> str:
    governance = summary.get("governance") or {}
    sellersprite = summary.get("sellersprite") or {}
    adaptive_profile = summary.get("adaptive_profile") or {}
    adaptive_history = summary.get("adaptive_history") or {}
    adaptive_thresholds = summary.get("adaptive_thresholds") or {}
    threshold_history = summary.get("threshold_history") or {}
    query_feedback = summary.get("query_feedback") or {}
    query_history = summary.get("query_history") or {}
    market_feedback = summary.get("market_feedback") or {}
    market_history = summary.get("market_history") or {}
    qualified = [item for item in decisions if item.qualified]
    lines = [
        f"Cross-market arbitrage run completed.",
        f"Run ID: {summary['run_id']}",
        f"过去24小时候选数: {summary['discovery_candidate_count']}",
        f"通过强规则候选数: {summary['qualified_count']}",
        f"治理状态: {governance.get('status', 'unknown')}",
    ]
    primary_blocker = governance.get("primary_blocker")
    if primary_blocker and primary_blocker != "none":
        lines.append(f"主阻塞: {primary_blocker}")
    next_actions = governance.get("next_actions", []) or []
    if next_actions:
        lines.append(f"下一步: {', '.join(next_actions[:3])}")
    if sellersprite:
        lines.append(
            f"SellerSprite: {sellersprite.get('status', 'unavailable')} / month={sellersprite.get('latest_month', '')} / api_ok={sellersprite.get('api_ok_total', 0)} / seed_queries={len(sellersprite.get('seed_queries') or [])} / keyword_results={((sellersprite.get('keyword_result_probe') or {}).get('result_count', 0))} / product_rows={((sellersprite.get('product_result_probe') or {}).get('product_rows_detected', 0))}"
        )
    if adaptive_profile:
        lines.append(
            f"Adaptive: blocker={adaptive_profile.get('primary_blocker', 'none')} / keyword_x={adaptive_profile.get('keyword_signal_multiplier', 1.0)} / product_x={adaptive_profile.get('product_signal_multiplier', 1.0)}"
        )
        lines.append(
            f"Source adaptive: order={','.join(adaptive_profile.get('source_order', []) or ['none'])} / biases={json.dumps(adaptive_profile.get('source_biases', {}), ensure_ascii=False)}"
        )
        lines.append(
            f"Query adaptive: focus={','.join(adaptive_profile.get('query_focus', []) or ['none'])} / biases={json.dumps(adaptive_profile.get('query_biases', {}), ensure_ascii=False)}"
        )
    if adaptive_history:
        lines.append(
            f"Adaptive trend: {((adaptive_history.get('trend') or {}).get('direction', 'stable'))} / delta={((adaptive_history.get('trend') or {}).get('delta', 0))} / entries={adaptive_history.get('entries_total', 0)}"
        )
    if adaptive_thresholds:
        lines.append(
            f"Adaptive thresholds: order>={adaptive_thresholds.get('order_floor', 30.0)} / age<={adaptive_thresholds.get('listing_age_ceiling', 730)} / margin>={adaptive_thresholds.get('margin_floor', 0.30)} / fit>={adaptive_thresholds.get('platform_fit_floor', 65.0)}"
        )
    if threshold_history:
        lines.append(
            f"Threshold trend: {((threshold_history.get('trend') or {}).get('direction', 'stable'))} / delta={((threshold_history.get('trend') or {}).get('delta', 0))} / entries={threshold_history.get('entries_total', 0)}"
        )
    if query_feedback:
        lines.append(
            f"Query feedback: primary={query_feedback.get('primary_query', 'none')} / tracked={len(query_feedback.get('queries') or [])} / monopoly_focus={round(float(((query_feedback.get('queries') or [{}])[0] or {}).get('avg_head_monopoly_score', 0) or 0), 2) if (query_feedback.get('queries') or []) else 0}"
        )
    if query_history:
        lines.append(
            f"Query trend: {((query_history.get('trend') or {}).get('direction', 'stable'))} / delta={((query_history.get('trend') or {}).get('delta', 0))} / entries={query_history.get('entries_total', 0)}"
        )
    if market_feedback:
        lines.append(
            f"Market pressure: primary={market_feedback.get('primary_market', 'none')} / tracked={len(market_feedback.get('markets') or [])} / monopoly={round(float(((market_feedback.get('markets') or [{}])[0] or {}).get('head_monopoly_avg', 0) or 0), 2) if (market_feedback.get('markets') or []) else 0}"
        )
    if market_history:
        lines.append(
            f"Market trend: {((market_history.get('trend') or {}).get('direction', 'stable'))} / delta={((market_history.get('trend') or {}).get('delta', 0))} / entries={market_history.get('entries_total', 0)}"
        )
    if qualified:
        top = qualified[:3]
        for item in top:
            lines.append(
                f"- {item.product_name}: 买 {item.buy_platform} / 卖 {item.sell_platform} / 毛利率 {item.gross_margin_rate:.2%}"
            )
    else:
        lines.append("- 本轮没有候选通过强阈值；请看附件里的审计和证据。")
    near_miss = [item for item in decisions if item.decision_bucket == "near_miss"]
    if near_miss:
        lines.append(f"- Near miss: {near_miss[0].product_name} / {near_miss[0].sell_platform} / {near_miss[0].priority_reason}")
    return "\n".join(lines)


def _decision_failure_categories(decisions: list[ArbitrageDecision]) -> dict[str, int]:
    categories: dict[str, int] = {}
    for item in decisions:
        if item.qualified:
            continue
        reason = "unknown_failure"
        for candidate in item.reasons or []:
            normalized = str(candidate or "").strip()
            if normalized:
                reason = normalized
                break
        categories[reason] = categories.get(reason, 0) + 1
    return dict(sorted(categories.items(), key=lambda row: (-row[1], row[0])))


def _build_governance_summary(rows: list[dict[str, Any]], decisions: list[ArbitrageDecision]) -> dict[str, Any]:
    failure_categories = _decision_failure_categories(decisions)
    primary_blocker = next(iter(failure_categories.keys()), "none")
    if any(item.qualified for item in decisions):
        status = "healthy"
        next_actions = ["continue_discovery_cadence", "continue_matching_cadence"]
    elif rows and primary_blocker == "no_usable_source_match":
        status = "attention_required"
        next_actions = ["strengthen_source_matching", "increase_source_detail_extraction", "review_source_site_health"]
    elif rows and primary_blocker in {"listing_age_unknown", "estimated_daily_orders_below_threshold"}:
        status = "attention_required"
        next_actions = ["strengthen_demand_signal_extraction", "improve_listing_age_capture", "improve_order_estimation"]
    else:
        status = "watching"
        next_actions = ["continue_discovery_cadence", "continue_matching_cadence", "inspect_recent_failures"]
    return {
        "status": status,
        "primary_blocker": primary_blocker,
        "failure_categories": failure_categories,
        "next_actions": next_actions,
    }


def _platform_summary(decisions: list[ArbitrageDecision]) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for platform in ["amazon", "walmart", "temu", "tiktok"]:
        rows = [item for item in decisions if item.sell_platform.lower() == platform]
        summary[platform] = {
            "candidate_count": len(rows),
            "qualified_count": sum(1 for item in rows if item.qualified),
            "near_miss_count": sum(1 for item in rows if item.decision_bucket == "near_miss"),
            "strong_fit_count": sum(1 for item in rows if item.platform_fit_label == "strong"),
            "watch_count": sum(1 for item in rows if item.platform_recommendation == "watchlist_only"),
        }
    return summary


def run_once(*, test: bool = False) -> dict[str, Any]:
    previous_timeout_cap = os.environ.get("CROSS_MARKET_TIMEOUT_CAP")
    if test:
        os.environ["CROSS_MARKET_TIMEOUT_CAP"] = "12"
    run_id = _utc_now().strftime("%Y%m%dT%H%M%SZ")
    try:
        state = _load_state()
        adaptive_profile = _adaptive_profile_from_state(state)
        adaptive_history = _persist_adaptive_history(adaptive_profile)
        query_feedback = _query_feedback_from_state(state)
        query_history = _persist_query_history(query_feedback)
        market_feedback = _market_feedback_from_state(state)
        market_history = _persist_market_history(market_feedback)
        adaptive_thresholds = _adaptive_thresholds_from_state(state)
        threshold_history = _persist_threshold_history(adaptive_thresholds)
        base_queries = DEFAULT_DISCOVERY_QUERIES[:2] if test else DEFAULT_DISCOVERY_QUERIES
        sellersprite = _collect_sellersprite_summary(base_queries, fast=test)
        seed_query_keys = {_query_key(item) for item in (sellersprite.get("seed_queries") or []) if _query_key(item)}
        base_query_keys = {_query_key(item) for item in base_queries if _query_key(item)}
        query_limit = 2 if test else 10
        queries = _apply_query_adaptive_order(
            _merge_queries(base_queries, sellersprite.get("seed_queries") or [], limit=query_limit),
            adaptive_profile,
            limit=query_limit,
        )
        sell_platforms = ["amazon", "temu"] if test else ["temu", "amazon", "walmart"]
        source_platforms = list((adaptive_profile.get("source_order") or []) or ["made_in_china", "1688", "yiwugo"])
        if test:
            test_source_platforms = [platform for platform in source_platforms if platform != "yiwugo"]
            source_platforms = (test_source_platforms or source_platforms)[:2]
        max_tools = 1 if test else None

        fetch_log: list[dict[str, Any]] = []
        demand_candidates: list[DemandCandidate] = []

        for query in queries:
            for platform in sell_platforms:
                result = fetch_best(platform, url_for_sell(platform, query), max_tools=max_tools, fast=test)
                fetch_log.append({"platform": platform, "query": query, "tool": result.tool, "status": result.status, "score": result.score})
                demand_candidates.extend(_extract_sell_candidates(platform, result.text, query))

        deduped: dict[tuple[str, str], DemandCandidate] = {}
        for item in demand_candidates:
            query_key = _query_key(item.query)
            item.raw_signals = {
                **(item.raw_signals or {}),
                "query_source_kind": (
                    "shared" if query_key in seed_query_keys and query_key in base_query_keys
                    else "seed" if query_key in seed_query_keys
                    else "base"
                ),
            }
            key = (item.sell_platform, _normalize_title(item.title))
            if key not in deduped:
                deduped[key] = item
        demand_candidates = list(deduped.values())
        competition_hints = _competition_density_hints(demand_candidates)
        for item in demand_candidates:
            if item.candidate_id in competition_hints:
                item.raw_signals = {**(item.raw_signals or {}), **competition_hints[item.candidate_id]}
        if test:
            demand_candidates = demand_candidates[:3]

        source_matches: dict[str, list[SourceCandidate]] = {}
        for candidate in demand_candidates:
            source_rows: list[SourceCandidate] = []
            for platform in source_platforms:
                source_bias = float(((adaptive_profile.get("source_biases") or {}).get(platform, 0) or 0))
                best_row, best_log = _best_source_for_platform(
                    platform,
                    candidate,
                    max_tools=max_tools,
                    platform_bias=source_bias,
                    fast=test,
                )
                fetch_log.extend(best_log)
                source_rows.append(best_row)
            source_matches[candidate.candidate_id] = source_rows

        decisions = [_compute_decision(item, source_matches.get(item.candidate_id, [])) for item in demand_candidates]
        summary = {
            "generated_at": _utc_now_iso(),
            "mode": "test" if test else "normal",
            "discovery_candidate_count": len(demand_candidates),
            "qualified_count": sum(1 for item in decisions if item.qualified),
            "platform_summary": _platform_summary(decisions),
            "adaptive_profile": adaptive_profile,
            "adaptive_history": adaptive_history,
            "adaptive_thresholds": adaptive_thresholds,
            "threshold_history": threshold_history,
            "query_feedback": query_feedback,
            "query_history": query_history,
            "market_feedback": market_feedback,
            "market_history": market_history,
            "sellersprite": sellersprite,
            "base_queries": base_queries,
            "queries": queries,
            "timing": {
                "discovery_interval_seconds": DISCOVERY_INTERVAL_SECONDS,
                "matching_interval_seconds": MATCH_INTERVAL_SECONDS,
                "report_hour_new_york": REPORT_HOUR,
            },
        }
        excel_path = _write_excel(run_id, decisions, summary)
        md_path = _write_markdown(run_id, decisions, summary)
        json_path = _write_json(
            run_id,
            {
                "summary": summary,
                "fetch_log": fetch_log,
                "decisions": [asdict(item) for item in decisions],
                "source_matches": {key: [asdict(row) for row in rows] for key, rows in source_matches.items()},
            },
        )

        state.setdefault("runs", []).append(
            {
                "run_id": run_id,
                "generated_at": summary["generated_at"],
                "qualified_count": summary["qualified_count"],
                "excel_path": str(excel_path),
                "markdown_path": str(md_path),
                "json_path": str(json_path),
            }
        )
        state["runs"] = state["runs"][-40:]
        _save_state(state)

        return {
            "run_id": run_id,
            "summary": summary,
            "excel_path": str(excel_path),
            "markdown_path": str(md_path),
            "json_path": str(json_path),
        }
    finally:
        if previous_timeout_cap is None:
            os.environ.pop("CROSS_MARKET_TIMEOUT_CAP", None)
        else:
            os.environ["CROSS_MARKET_TIMEOUT_CAP"] = previous_timeout_cap


def _discover_cycle(state: dict[str, Any], *, test: bool = False) -> tuple[dict[str, Any], list[DemandCandidate], list[dict[str, Any]]]:
    adaptive_profile = _adaptive_profile_from_state(state)
    adaptive_thresholds = _adaptive_thresholds_from_state(state)
    base_queries = DEFAULT_DISCOVERY_QUERIES[:1] if test else DEFAULT_DISCOVERY_QUERIES
    sellersprite = _collect_sellersprite_summary(base_queries, fast=test)
    seed_query_keys = {_query_key(item) for item in (sellersprite.get("seed_queries") or []) if _query_key(item)}
    base_query_keys = {_query_key(item) for item in base_queries if _query_key(item)}
    query_limit = 2 if test else 10
    queries = _apply_query_adaptive_order(
        _merge_queries(base_queries, sellersprite.get("seed_queries") or [], limit=query_limit),
        adaptive_profile,
        limit=query_limit,
    )
    sell_platforms = ["amazon", "temu"] if test else ["temu", "amazon", "walmart"]
    max_tools = 1 if test else None
    fetch_log: list[dict[str, Any]] = []
    candidates: list[DemandCandidate] = []
    for query in queries:
        for platform in sell_platforms:
            result = fetch_best(platform, url_for_sell(platform, query), max_tools=max_tools, fast=test)
            fetch_log.append({"stage": "discover", "platform": platform, "query": query, "tool": result.tool, "status": result.status, "score": result.score})
            candidates.extend(_extract_sell_candidates(platform, result.text, query))
    enriched: list[DemandCandidate] = []
    for item in candidates:
        updated, detail_log = _enrich_sell_candidate(item, max_tools=max_tools)
        fetch_log.append(detail_log)
        hints = _sellersprite_candidate_hints(sellersprite, updated)
        hints["sellersprite_keyword_signal_score"] = round(
            min(
                100.0,
                float(hints.get("sellersprite_keyword_signal_score", 0) or 0)
                * float(adaptive_profile.get("keyword_signal_multiplier", 1.0) or 1.0),
            ),
            2,
        )
        hints["sellersprite_product_signal_score"] = round(
            min(
                100.0,
                float(hints.get("sellersprite_product_signal_score", 0) or 0)
                * float(adaptive_profile.get("product_signal_multiplier", 1.0) or 1.0),
            ),
            2,
        )
        hints["adaptive_platform_bias"] = float(
            ((adaptive_profile.get("platform_biases") or {}).get(updated.sell_platform.lower(), 0) or 0)
        )
        hints["adaptive_query_bias"] = float(
            ((adaptive_profile.get("query_biases") or {}).get(_query_key(updated.query), 0) or 0)
        )
        query_key = _query_key(updated.query)
        hints["query_source_kind"] = (
            "shared" if query_key in seed_query_keys and query_key in base_query_keys
            else "seed" if query_key in seed_query_keys
            else "base"
        )
        hints["adaptive_thresholds"] = adaptive_thresholds
        updated.raw_signals = {**(updated.raw_signals or {}), **hints}
        sellersprite_daily_proxy = float(hints.get("sellersprite_product_daily_orders_proxy", 0) or 0)
        if sellersprite_daily_proxy > 0:
            current_daily_orders = float(updated.estimated_daily_orders or 0)
            if current_daily_orders <= 0:
                updated.estimated_daily_orders = round(sellersprite_daily_proxy * 0.65, 2)
                updated.raw_signals["estimated_daily_orders_source"] = "sellersprite_product_proxy"
            elif current_daily_orders < 20 and sellersprite_daily_proxy > current_daily_orders:
                blended = max(current_daily_orders, min(sellersprite_daily_proxy * 0.55, current_daily_orders * 1.8))
                updated.estimated_daily_orders = round(blended, 2)
                updated.raw_signals["estimated_daily_orders_source"] = "blended_sellersprite_product_proxy"
        if hints.get("sellersprite_product_freshest_listing_days") is not None and updated.listing_age_days is None:
            updated.listing_age_days = int(hints["sellersprite_product_freshest_listing_days"])
        hint_boost = max(
            float(hints.get("sellersprite_keyword_signal_score", 0) or 0) * 0.12,
            float(hints.get("sellersprite_product_signal_score", 0) or 0) * 0.16,
        )
        if hint_boost:
            updated.demand_confidence = round(min(95.0, updated.demand_confidence + hint_boost), 2)
        enriched.append(updated)
    candidates = enriched
    competition_hints = _competition_density_hints(candidates)
    for item in candidates:
        if item.candidate_id in competition_hints:
            item.raw_signals = {**(item.raw_signals or {}), **competition_hints[item.candidate_id]}
    bucket = state.setdefault("candidates", {})
    for item in candidates:
        row = bucket.get(item.candidate_id, {})
        previous_snapshot = {
            "title": row.get("title"),
            "sell_price_cny": row.get("sell_price_cny"),
            "estimated_daily_orders": row.get("estimated_daily_orders"),
            "listing_age_days": row.get("listing_age_days"),
            "rating_value": row.get("rating_value"),
            "review_count": row.get("review_count"),
        }
        price_history = list(row.get("observed_sell_prices", []) or [])
        if item.sell_price_cny is not None:
            price_history.append(item.sell_price_cny)
            price_history = price_history[-20:]
        row.update(asdict(item))
        row["updated_at"] = _utc_now_iso()
        row["discovered_at"] = row.get("discovered_at") or item.extracted_at
        row["observed_sell_prices"] = price_history
        row.setdefault("source_matches", [])
        current_snapshot = {
            "title": row.get("title"),
            "sell_price_cny": row.get("sell_price_cny"),
            "estimated_daily_orders": row.get("estimated_daily_orders"),
            "listing_age_days": row.get("listing_age_days"),
            "rating_value": row.get("rating_value"),
            "review_count": row.get("review_count"),
        }
        if previous_snapshot != current_snapshot:
            row.pop("last_matched_at", None)
            row.pop("decision", None)
            row["source_matches"] = []
        row["demand_refresh_priority"] = _refresh_priority(row)
        bucket[item.candidate_id] = row
    state["last_discovery_at"] = _utc_now_iso()
    state["last_sellersprite_seed_queries"] = list(sellersprite.get("seed_queries") or [])
    state["last_effective_queries"] = list(queries)
    state["last_adaptive_profile"] = adaptive_profile
    state["last_query_feedback"] = _query_feedback_from_state(state)
    return state, candidates, fetch_log


def _match_cycle(state: dict[str, Any], *, test: bool = False) -> tuple[dict[str, Any], list[ArbitrageDecision], list[dict[str, Any]]]:
    adaptive_profile = _adaptive_profile_from_state(state)
    source_platforms = list((adaptive_profile.get("source_order") or []) or ["made_in_china", "1688", "yiwugo"])
    if test:
        source_platforms = source_platforms[:2]
    max_tools = 1 if test else None
    fetch_log: list[dict[str, Any]] = []
    decisions: list[ArbitrageDecision] = []
    ordered_candidates = sorted(
        (state.get("candidates") or {}).items(),
        key=lambda item: (-_ensure_refresh_priority(item[1] or {}), str(item[0])),
    )
    for candidate_id, payload in ordered_candidates:
        if not _needs_rematch(payload):
            continue
        candidate = DemandCandidate(**{k: payload[k] for k in DemandCandidate.__dataclass_fields__.keys() if k in payload})
        rows: list[SourceCandidate] = []
        for platform in source_platforms:
            source_bias = float(((adaptive_profile.get("source_biases") or {}).get(platform, 0) or 0))
            best_row, best_log = _best_source_for_platform(
                platform,
                candidate,
                max_tools=max_tools,
                platform_bias=source_bias,
                fast=test,
            )
            fetch_log.extend(best_log)
            rows.append(best_row)
        payload["source_matches"] = [asdict(row) for row in rows]
        payload["last_matched_at"] = _utc_now_iso()
        decision = _compute_decision(candidate, rows)
        payload["decision"] = asdict(decision)
        decisions.append(decision)
    state["last_match_at"] = _utc_now_iso()
    return state, decisions, fetch_log


def _report_cycle(
    state: dict[str, Any],
    *,
    chat_id: str | None,
    force_send: bool = False,
    scheduler_policy: dict[str, Any] | None = None,
) -> dict[str, Any]:
    run_id = _utc_now().strftime("%Y%m%dT%H%M%SZ")
    scheduler_policy = scheduler_policy or {}
    rows = _state_candidate_window(state, hours=REPORT_WINDOW_HOURS)
    decisions: list[ArbitrageDecision] = []
    fetch_log: list[dict[str, Any]] = []
    for payload in rows:
        decision = payload.get("decision")
        if decision:
            decisions.append(_decision_from_payload(decision))
    summary = {
        "run_id": run_id,
        "generated_at": _utc_now_iso(),
        "mode": "scheduled",
        "discovery_candidate_count": len(rows),
        "qualified_count": sum(1 for item in decisions if item.qualified),
        "platform_summary": _platform_summary(decisions),
        "adaptive_profile": _adaptive_profile_from_state(state),
        "adaptive_history": _persist_adaptive_history(_adaptive_profile_from_state(state)),
        "adaptive_thresholds": _adaptive_thresholds_from_state(state),
        "threshold_history": _persist_threshold_history(_adaptive_thresholds_from_state(state)),
        "query_feedback": _query_feedback_from_state(state),
        "query_history": _persist_query_history(_query_feedback_from_state(state)),
        "market_feedback": _market_feedback_from_state(state),
        "market_history": _persist_market_history(_market_feedback_from_state(state)),
        "sellersprite": _collect_sellersprite_summary(DEFAULT_DISCOVERY_QUERIES),
        "base_queries": DEFAULT_DISCOVERY_QUERIES,
        "queries": state.get("last_effective_queries") or DEFAULT_DISCOVERY_QUERIES,
        "timing": {
            "discovery_interval_seconds": int(scheduler_policy.get("discovery_interval_seconds", DISCOVERY_INTERVAL_SECONDS) or DISCOVERY_INTERVAL_SECONDS),
            "matching_interval_seconds": int(scheduler_policy.get("matching_interval_seconds", MATCH_INTERVAL_SECONDS) or MATCH_INTERVAL_SECONDS),
            "report_hour_new_york": int(scheduler_policy.get("report_hour_new_york", REPORT_HOUR) or REPORT_HOUR),
        },
        "governance": _build_governance_summary(rows, decisions),
        "scheduler_policy": scheduler_policy,
    }
    excel_path = _write_excel(run_id, decisions, summary)
    md_path = _write_markdown(run_id, decisions, summary)
    payload = {
        "summary": summary,
        "decisions": [asdict(item) for item in decisions],
        "fetch_log": fetch_log,
        "deliveries": [],
    }
    json_path = _write_json(run_id, payload)
    state.setdefault("runs", []).append(
        {
            "run_id": run_id,
            "generated_at": summary["generated_at"],
            "qualified_count": summary["qualified_count"],
            "excel_path": str(excel_path),
            "markdown_path": str(md_path),
            "json_path": str(json_path),
        }
    )
    state["runs"] = state["runs"][-40:]
    state["last_report_date"] = _ny_now().date().isoformat()
    deliveries = []
    if chat_id and force_send:
        deliveries = _send_to_telegram(chat_id=chat_id, text=_summary_text(summary, decisions), media_paths=[md_path, json_path, excel_path])
    summary["memory_writeback"] = record_memory_writeback(
        "project-cross-market-arbitrage",
        source="cross_market_arbitrage_cycle",
        summary={
            "attention_required": bool((summary.get("governance", {}) or {}).get("primary_blocker")),
            "state_patch": {},
            "governance_patch": {},
            "next_actions": [
                str(item).strip()
                for item in ((summary.get("governance", {}) or {}).get("next_actions", []) or [])
                if str(item).strip()
            ],
            "warnings": [str((summary.get("governance", {}) or {}).get("primary_blocker", "")).strip()] if str((summary.get("governance", {}) or {}).get("primary_blocker", "")).strip() else [],
            "errors": [],
            "decisions": ["cross_market_arbitrage_cycle_completed"],
            "memory_targets": ["project", "runtime"],
            "memory_reasons": ["cross_market_arbitrage_cycle", "project_arbitrage_feedback"],
        },
    )
    payload["summary"] = summary
    payload["deliveries"] = deliveries
    _write_json(run_id, payload)
    _write_json_file(
        LATEST_REPORT_PATH,
        {
            "run_id": run_id,
            "generated_at": summary["generated_at"],
            "qualified_count": summary["qualified_count"],
            "governance": summary.get("governance", {}) or {},
            "sellersprite": summary.get("sellersprite", {}) or {},
            "deliveries": deliveries,
            "excel_path": str(excel_path),
            "markdown_path": str(md_path),
            "json_path": str(json_path),
            "query_feedback": summary.get("query_feedback", {}) or {},
        },
    )
    _save_state(state)
    return {"run_id": run_id, "summary": summary, "excel_path": str(excel_path), "markdown_path": str(md_path), "json_path": str(json_path), "deliveries": deliveries}


def _discovery_due(state: dict[str, Any], interval_seconds: int = DISCOVERY_INTERVAL_SECONDS) -> bool:
    last = parse_dt(state.get("last_discovery_at", ""))
    return last is None or (_utc_now() - last).total_seconds() >= interval_seconds


def _match_due(state: dict[str, Any], interval_seconds: int = MATCH_INTERVAL_SECONDS) -> bool:
    last = parse_dt(state.get("last_match_at", ""))
    has_unmatched = any(_needs_rematch(row) for row in (state.get("candidates") or {}).values())
    return has_unmatched and (last is None or (_utc_now() - last).total_seconds() >= interval_seconds)


def _report_due(state: dict[str, Any], report_hour: int = REPORT_HOUR) -> bool:
    ny_now = _ny_now()
    if ny_now.hour < report_hour:
        return False
    return state.get("last_report_date") != ny_now.date().isoformat()


def _seconds_until_next_discovery() -> int:
    state = _load_state()
    runs = state.get("runs", []) or []
    if not runs:
        return 0
    try:
        last = datetime.fromisoformat(runs[-1]["generated_at"])
    except Exception:
        return 0
    elapsed = (_utc_now() - last).total_seconds()
    return max(0, int(DISCOVERY_INTERVAL_SECONDS - elapsed))


def run_daemon() -> int:
    while True:
        state = _load_state()
        scheduler_policy = _load_scheduler_policy()
        scheduler_state = _load_scheduler_state()
        execution_flags = _execution_flags_from_scheduler_policy(scheduler_policy)
        discovery_interval = int(scheduler_policy.get("discovery_interval_seconds", DISCOVERY_INTERVAL_SECONDS) or DISCOVERY_INTERVAL_SECONDS)
        matching_interval = int(scheduler_policy.get("matching_interval_seconds", MATCH_INTERVAL_SECONDS) or MATCH_INTERVAL_SECONDS)
        report_hour = int(scheduler_policy.get("report_hour_new_york", REPORT_HOUR) or REPORT_HOUR)
        loop_sleep_seconds = int(scheduler_policy.get("loop_sleep_seconds", 300) or 300)
        changed = False
        effective_discovery = False
        effective_match = False
        effective_report = False
        if execution_flags["allow_discovery"] and _discovery_due(state, interval_seconds=discovery_interval):
            state, _, _ = _discover_cycle(state, test=False)
            changed = True
            effective_discovery = True
        if execution_flags["allow_match"] and _match_due(state, interval_seconds=matching_interval):
            state, _, _ = _match_cycle(state, test=False)
            changed = True
            effective_match = True
        if changed:
            _save_state(state)
        if execution_flags["allow_report"] and _report_due(state, report_hour=report_hour):
            _report_cycle(
                state,
                chat_id=os.environ.get("CROSS_MARKET_TELEGRAM_CHAT", DEFAULT_TELEGRAM_TARGET),
                force_send=True,
                scheduler_policy=scheduler_policy,
            )
            effective_report = True
        scheduler_state_after = {
            "updated_at": _utc_now_iso(),
            "last_mode": scheduler_policy.get("recommended_mode", ""),
            "last_repair_focus": str(scheduler_policy.get("repair_focus", "")).strip(),
            "last_repair_mode": str(scheduler_policy.get("repair_mode", "")).strip(),
            "loop_sleep_seconds": loop_sleep_seconds,
            "discovery_interval_seconds": discovery_interval,
            "matching_interval_seconds": matching_interval,
            "report_hour_new_york": report_hour,
            "last_allow_discovery": execution_flags["allow_discovery"],
            "last_allow_match": execution_flags["allow_match"],
            "last_allow_report": execution_flags["allow_report"],
            "last_effective_discovery": effective_discovery,
            "last_effective_match": effective_match,
            "last_effective_report": effective_report,
            "last_discovery_started_at": _utc_now_iso() if effective_discovery else scheduler_state.get("last_discovery_started_at", ""),
            "last_match_started_at": _utc_now_iso() if effective_match else scheduler_state.get("last_match_started_at", ""),
            "last_report_started_at": _utc_now_iso() if effective_report else scheduler_state.get("last_report_started_at", ""),
        }
        _write_scheduler_state(scheduler_state_after)
        time.sleep(loop_sleep_seconds)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["once", "daemon"], default="once")
    parser.add_argument("--test", action="store_true")
    parser.add_argument("--send-telegram", action="store_true")
    parser.add_argument("--telegram-chat", default=DEFAULT_TELEGRAM_TARGET)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    if args.mode == "daemon":
        return run_daemon()
    scheduler_policy = _load_scheduler_policy()
    scheduler_state_before = _load_scheduler_state()
    execution_flags = _execution_flags_from_scheduler_policy(scheduler_policy)
    if not args.force and scheduler_policy and not bool(scheduler_policy.get("start_tasks", True)):
        payload = {
            "ran": False,
            "reason": "scheduler_policy_start_tasks_false",
            "scheduler_policy": scheduler_policy,
            "execution_flags": execution_flags,
            "scheduler_state_before": scheduler_state_before,
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0
    if args.test:
        payload = run_once(test=True)
        payload["scheduler_policy"] = scheduler_policy
        payload["execution_flags"] = execution_flags
        payload["scheduler_state_before"] = scheduler_state_before
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return 0
    state = _load_state()
    discover_log = []
    match_log = []
    decisions = []
    if execution_flags["allow_discovery"]:
        state, _, discover_log = _discover_cycle(state, test=args.test)
    if execution_flags["allow_match"]:
        state, decisions, match_log = _match_cycle(state, test=args.test)
    _save_state(state)
    payload = _report_cycle(
        state,
        chat_id=args.telegram_chat if args.send_telegram else None,
        force_send=args.send_telegram,
        scheduler_policy=scheduler_policy,
    )
    payload["fetch_log_count"] = len(discover_log) + len(match_log)
    payload["qualified_preview"] = [asdict(item) for item in decisions if item.qualified][:3]
    payload["scheduler_policy"] = scheduler_policy
    payload["execution_flags"] = execution_flags
    payload["scheduler_state_before"] = scheduler_state_before
    scheduler_state_after = {
        "updated_at": _utc_now_iso(),
        "last_mode": scheduler_policy.get("recommended_mode", ""),
        "last_repair_focus": str(scheduler_policy.get("repair_focus", "")).strip(),
        "last_repair_mode": str(scheduler_policy.get("repair_mode", "")).strip(),
        "loop_sleep_seconds": int(scheduler_policy.get("loop_sleep_seconds", 300) or 300),
        "discovery_interval_seconds": int(scheduler_policy.get("discovery_interval_seconds", DISCOVERY_INTERVAL_SECONDS) or DISCOVERY_INTERVAL_SECONDS),
        "matching_interval_seconds": int(scheduler_policy.get("matching_interval_seconds", MATCH_INTERVAL_SECONDS) or MATCH_INTERVAL_SECONDS),
        "report_hour_new_york": int(scheduler_policy.get("report_hour_new_york", REPORT_HOUR) or REPORT_HOUR),
        "last_allow_discovery": execution_flags["allow_discovery"],
        "last_allow_match": execution_flags["allow_match"],
        "last_allow_report": execution_flags["allow_report"],
        "last_effective_discovery": execution_flags["allow_discovery"],
        "last_effective_match": execution_flags["allow_match"],
        "last_effective_report": True,
        "last_discovery_started_at": _utc_now_iso() if execution_flags["allow_discovery"] else scheduler_state_before.get("last_discovery_started_at", ""),
        "last_match_started_at": _utc_now_iso() if execution_flags["allow_match"] else scheduler_state_before.get("last_match_started_at", ""),
        "last_report_started_at": _utc_now_iso(),
    }
    payload["scheduler_state_after"] = scheduler_state_after
    _write_scheduler_state(scheduler_state_after)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
