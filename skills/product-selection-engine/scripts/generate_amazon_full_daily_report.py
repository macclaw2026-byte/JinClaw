#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Users/mac_claw/.openclaw/workspace')
RAW = ROOT / 'data' / 'amazon-premium-wholesale' / 'raw_candidates.json'
OUT_DIR = ROOT / 'output'
OUT_DIR.mkdir(parents=True, exist_ok=True)

SUMMARY_SHEET = '日报摘要'
AMAZON_SHEET = '亚马逊全量结果'
SUMMARY_HEADERS = ['字段', '内容']
AMAZON_HEADERS = [
    '排名', '产品名称', '细分方向', '来源组合', '价格', '评分', '评论数', '月销量提示', '竞品链接',
    '需求强度', '搜索意图', '痛点清晰度', '复购需求', '评分类代理', '价格带强度', '列表密度',
    '中腰部存活信号', '竞争拥挤度', '产品简单度', '利润可行性', '生产可行性', '推荐原因'
]


def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-') or 'candidate'


def normalize_title_for_dedupe(text: str) -> str:
    text = text.lower()
    text = re.sub(r'\b\d+\s*(pack|pcs|piece|pieces|count)\b', ' ', text)
    text = re.sub(r'\b(extra large|large|medium|small|black|white|gray|grey|clear|natural)\b', ' ', text)
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    tokens = [t for t in text.split() if t not in {
        'for','and','with','the','pack','pcs','piece','pieces','organizer','storage','tray','box','set','kitchen','bathroom','office','bedroom'
    }]
    return ' '.join(tokens)


def infer_family_key(product_name: str, sub_niche: str) -> str:
    base = normalize_title_for_dedupe(product_name)
    family_patterns = [
        'drawer divider', 'silverware organizer', 'under sink organizer', 'cable management box', 'drawer organizer', 'clear drawer'
    ]
    family_prefix = ''
    for pattern in family_patterns:
        if all(part in base for part in pattern.split()):
            family_prefix = pattern.replace(' ', '-')
            break
    if not family_prefix:
        family_prefix = slugify(sub_niche or product_name)
    meaningful = [
        t for t in base.split()
        if t not in {'adjustable', 'expandable', 'pull', 'out', 'clear', 'plastic', 'bamboo', 'kitchen', 'bathroom', 'office', 'pack'}
    ]
    core = '-'.join(meaningful[:3]) if meaningful else family_prefix
    return f'{family_prefix}::{core}'


def score(item: dict) -> float:
    demand = item['demand_features']
    market = item['marketplace_features']
    business = item['business_features']
    diff = item['differentiation_features']
    demand_score = (demand['trend_strength'] + demand['search_intent_strength'] + demand['customer_pain_clarity'] + demand['repeat_need_strength']) / 4 * 100
    proxy = (market['review_depth_proxy'] + market['price_ladder_strength'] + market['listing_density'] + market['mid_tier_survival_signal']) / 4 * 100
    survivability = ((1 - market['competition_crowding']) + market['mid_tier_survival_signal'] + diff['clear_angle_exists'] + diff['competitor_positioning_gap']) / 4 * 100
    simplicity = (business['simplicity'] + business['manufacturability'] + (1 - business['fragility_risk']) + (1 - business['return_risk'])) / 4 * 100
    margin = business['margin_viability'] * 100
    differentiation = (diff['clear_angle_exists'] + diff['tool_gap_signal'] + diff['improvement_room'] + diff['competitor_positioning_gap']) / 4 * 100
    return round(demand_score*0.25 + proxy*0.20 + survivability*0.20 + simplicity*0.15 + margin*0.10 + differentiation*0.10, 1)


def load_and_dedupe() -> tuple[list[dict], dict]:
    payload = json.loads(RAW.read_text())
    candidates = payload.get('candidates', [])
    grouped: dict[str, dict] = {}
    suppressed = 0
    for item in candidates:
        flags = item.get('category_flags', {})
        if any(flags.get(k) for k in ['excluded', 'brand_risk', 'regulated', 'body_contact_formula', 'ingestible']):
            continue
        key = infer_family_key(item.get('product_name', ''), item.get('sub_niche', ''))
        item['_family_key'] = key
        item['_score'] = score(item)
        existing = grouped.get(key)
        if not existing or item['_score'] > existing['_score']:
            if existing:
                suppressed += 1
            grouped[key] = item
        else:
            suppressed += 1
    deduped = sorted(grouped.values(), key=lambda x: x['_score'], reverse=True)
    stats = {
        'raw_candidate_count': len(candidates),
        'qualified_after_filters': len(grouped),
        'suppressed_by_family_dedupe': suppressed,
        'queries': payload.get('queries', []),
        'generated_at': payload.get('generated_at', ''),
    }
    return deduped, stats


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def style_header(ws, row=1):
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def build_wb(rows: list[dict], stats: dict) -> Workbook:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = SUMMARY_SHEET
    ws1.append(SUMMARY_HEADERS)
    style_header(ws1)
    summary_rows = [
        ['报告生成时间', datetime.now().astimezone().isoformat()],
        ['原始候选总数', stats['raw_candidate_count']],
        ['过滤并家族去重后总数', stats['qualified_after_filters']],
        ['同族压缩数量', stats['suppressed_by_family_dedupe']],
        ['查询池', ', '.join(stats['queries'])],
        ['数据集时间', stats['generated_at']],
        ['说明', 'This nightly report includes all currently qualified candidates after filtering and deduplication, with no 20-item cap.'],
    ]
    for row in summary_rows:
        ws1.append(row)
    autosize(ws1)

    ws2 = wb.create_sheet(AMAZON_SHEET)
    ws2.append(AMAZON_HEADERS)
    style_header(ws2)
    for idx, item in enumerate(rows, start=1):
        pf = item.get('public_fields', {})
        d = item.get('demand_features', {})
        m = item.get('marketplace_features', {})
        b = item.get('business_features', {})
        ws2.append([
            idx,
            item.get('product_name', ''),
            item.get('sub_niche', ''),
            ', '.join(item.get('source_mix', [])),
            pf.get('price', ''),
            pf.get('rating', ''),
            pf.get('review_count', ''),
            pf.get('bought_past_month', ''),
            (item.get('competitor_links') or [''])[0],
            d.get('trend_strength', ''),
            d.get('search_intent_strength', ''),
            d.get('customer_pain_clarity', ''),
            d.get('repeat_need_strength', ''),
            m.get('review_depth_proxy', ''),
            m.get('price_ladder_strength', ''),
            m.get('listing_density', ''),
            m.get('mid_tier_survival_signal', ''),
            m.get('competition_crowding', ''),
            b.get('simplicity', ''),
            b.get('margin_viability', ''),
            b.get('manufacturability', ''),
            item.get('why_fit', ''),
        ])
    ws2.freeze_panes = 'A2'
    autosize(ws2)
    return wb


def main():
    rows, stats = load_and_dedupe()
    date_tag = datetime.now().strftime('%Y%m%d')
    out_path = OUT_DIR / f'亚马逊精品铺货完整夜报-{date_tag}.xlsx'
    wb = build_wb(rows, stats)
    wb.save(out_path)
    print(out_path)


if __name__ == '__main__':
    main()
