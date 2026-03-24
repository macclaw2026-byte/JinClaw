#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/Users/mac_claw/.openclaw/workspace')
LATEST = ROOT / 'output' / 'amazon-premium-wholesale' / 'latest.json'
OUT_DIR = ROOT / 'output'
OUT_DIR.mkdir(parents=True, exist_ok=True)

SUMMARY_SHEET = '日报摘要'
AMAZON_SHEET = '亚马逊精品铺货'

SUMMARY_HEADERS = ['字段', '内容']
AMAZON_HEADERS = [
    '排名', '产品名称', '细分方向', '来源组合', '预估日销区间', '证据等级',
    '总分', '需求分', '竞争存活分', '产品简单度', '推荐动作', '去重状态', '去重说明', '竞品链接', '适配原因'
]


def load_latest() -> dict:
    return json.loads(LATEST.read_text())


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            try:
                value = '' if cell.value is None else str(cell.value)
            except Exception:
                value = ''
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def style_header(ws, row=1):
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def build_workbook(payload: dict) -> Workbook:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = SUMMARY_SHEET
    ws_summary.append(SUMMARY_HEADERS)
    style_header(ws_summary)

    summary_rows = [
        ['报告生成时间', payload.get('run_at', '')],
        ['输入模式', payload.get('input_mode', '')],
        ['输入文件', payload.get('input_path', '')],
        ['原始候选数', payload.get('pre_dedupe_count', '')],
        ['家族去重后候选数', payload.get('post_family_dedupe_count', '')],
        ['最终候选数', payload.get('candidate_count', '')],
        ['报告目标', 'Amazon premium wholesale shortlist'],
        ['备注', 'Chinese file name / sheet names / headers; English body content preserved where applicable.'],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    autosize(ws_summary)

    ws_amazon = wb.create_sheet(AMAZON_SHEET)
    ws_amazon.append(AMAZON_HEADERS)
    style_header(ws_amazon)

    candidates = payload.get('candidates', [])
    for idx, item in enumerate(candidates, start=1):
        scores = item.get('scores', {})
        competitor_link = ''
        links = item.get('competitor_links') or []
        if links:
            competitor_link = links[0]
        ws_amazon.append([
            idx,
            item.get('product_name', ''),
            item.get('sub_niche', ''),
            ', '.join(item.get('source_mix', [])),
            item.get('sales_range', ''),
            item.get('evidence_grade', ''),
            scores.get('total_score', ''),
            scores.get('demand_score', ''),
            scores.get('competition_survivability_score', ''),
            scores.get('simplicity_score', ''),
            item.get('recommendation', ''),
            item.get('novelty_status', ''),
            item.get('dedupe_reason', ''),
            competitor_link,
            item.get('why_fit', ''),
        ])
    autosize(ws_amazon)
    ws_amazon.freeze_panes = 'A2'
    return wb


def main() -> int:
    payload = load_latest()
    run_at = payload.get('run_at')
    try:
        dt = datetime.fromisoformat(run_at)
        date_tag = dt.strftime('%Y%m%d')
    except Exception:
        date_tag = datetime.now().strftime('%Y%m%d')
    out_path = OUT_DIR / f'亚马逊精品铺货每日报告-{date_tag}.xlsx'
    wb = build_workbook(payload)
    wb.save(out_path)
    print(out_path)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
